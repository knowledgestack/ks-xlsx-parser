"""
Top-level workbook parser orchestrating the full parse pipeline.

This is the main entry point for parsing an Excel workbook. It coordinates
sheet parsing, table extraction, chart extraction, formula dependency
building, and metadata extraction into a complete WorkbookDTO.
"""

from __future__ import annotations

import io
import logging
import time
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
from pathlib import Path

import xxhash
from openpyxl import load_workbook

from ..models.common import ParseError, Severity
from ..models.workbook import (
    ExternalLink,
    NamedRangeDTO,
    WorkbookDTO,
    WorkbookProperties,
)
from .sheet_parser import SheetParser
from .table_parser import TableParser

logger = logging.getLogger(__name__)


class WorkbookParser:
    """
    Main parser for Excel .xlsx workbooks.

    Orchestrates the full parse pipeline:
    1. Load workbook and compute file hash
    2. Parse each sheet (cells, merges, formatting)
    3. Extract tables, charts, shapes
    4. Extract named ranges and metadata
    5. Build formula dependency graph
    6. Finalize all IDs and hashes

    Usage:
        parser = WorkbookParser(path="workbook.xlsx")
        result = parser.parse()
    """

    def __init__(
        self,
        path: str | Path | None = None,
        content: bytes | None = None,
        filename: str | None = None,
        max_cells_per_sheet: int = 2_000_000,
        parallel: bool = False,
        max_workers: int = 4,
    ):
        """
        Args:
            path: Path to the .xlsx file.
            content: Raw bytes of the .xlsx file (alternative to path).
            filename: Display filename (used when content is provided).
            max_cells_per_sheet: Safety limit per sheet.
            parallel: Whether to parse sheets in parallel (experimental).
            max_workers: Number of parallel workers.
        """
        if path is None and content is None:
            raise ValueError("Either path or content must be provided")

        self._path = Path(path) if path else None
        self._content = content
        self._filename = filename or (self._path.name if self._path else "unknown.xlsx")
        self._max_cells = max_cells_per_sheet
        self._parallel = parallel
        self._max_workers = max_workers

    def parse(self) -> WorkbookDTO:
        """
        Execute the full parse pipeline and return a WorkbookDTO.

        Returns:
            A fully populated WorkbookDTO with all sheets, tables,
            charts, named ranges, and metadata.
        """
        start_time = time.monotonic()
        logger.info("Starting workbook parse: %s", self._filename)

        # Compute file hash
        file_bytes = self._read_bytes()
        workbook_hash = xxhash.xxh64(file_bytes).hexdigest()

        # Initialize result DTO
        result = WorkbookDTO(
            filename=self._filename,
            file_path=str(self._path) if self._path else None,
            workbook_hash=workbook_hash,
        )

        # Detect macros
        has_macros = self._filename.lower().endswith(".xlsm")
        if has_macros:
            result.errors.append(ParseError(
                severity=Severity.WARNING,
                stage="load",
                message="Macro-enabled workbook detected (.xlsm). Macros will not be executed.",
            ))

        # Load workbook twice: with formulas and with computed values
        load_source = self._path or io.BytesIO(self._content)
        try:
            wb_formula = load_workbook(
                load_source,
                data_only=False,
                read_only=False,
            )
        except Exception as e:
            result.errors.append(ParseError(
                severity=Severity.ERROR,
                stage="load",
                message=f"Failed to load workbook: {e}",
            ))
            return result

        wb_computed = None
        load_source_2 = self._path or io.BytesIO(self._content)
        try:
            wb_computed = load_workbook(
                load_source_2,
                data_only=True,
                read_only=False,
            )
        except Exception as e:
            result.errors.append(ParseError(
                severity=Severity.WARNING,
                stage="load",
                message=f"Failed to load workbook in data_only mode: {e}",
            ))

        # Extract workbook properties
        result.properties = self._extract_properties(wb_formula, has_macros)

        # Parse sheets
        for idx, ws_name in enumerate(wb_formula.sheetnames):
            ws = wb_formula[ws_name]
            computed_ws = wb_computed[ws_name] if wb_computed and ws_name in wb_computed.sheetnames else None

            try:
                sheet_parser = SheetParser(
                    ws=ws,
                    sheet_index=idx,
                    computed_ws=computed_ws,
                    max_cells=self._max_cells,
                )
                sheet_dto = sheet_parser.parse()
                result.sheets.append(sheet_dto)

                # Extract tables from this sheet
                table_parser = TableParser(ws, ws_name)
                tables = table_parser.parse_all()
                result.tables.extend(tables)

            except Exception as e:
                logger.error("Failed to parse sheet '%s': %s", ws_name, e, exc_info=True)
                result.errors.append(ParseError(
                    severity=Severity.ERROR,
                    stage="parse",
                    message=f"Failed to parse sheet: {e}",
                    sheet_name=ws_name,
                ))

        # Extract named ranges
        result.named_ranges = self._extract_named_ranges(wb_formula)

        # Extract external links
        result.external_links = self._extract_external_links(wb_formula)

        # Extract charts via OOXML parsing
        try:
            from ..charts.chart_extractor import ChartExtractor
            chart_extractor = ChartExtractor(
                self._path or self._content, wb_formula.sheetnames
            )
            result.charts = chart_extractor.extract_all()
        except Exception as e:
            logger.warning("Chart extraction failed: %s", e)
            result.errors.append(ParseError(
                severity=Severity.WARNING,
                stage="parse",
                message=f"Chart extraction failed: {e}",
            ))

        # Build dependency graph
        try:
            from ..formula.dependency_builder import DependencyBuilder
            dep_builder = DependencyBuilder(result.sheets, result.named_ranges)
            result.dependency_graph = dep_builder.build()
        except Exception as e:
            logger.warning("Dependency graph building failed: %s", e)
            result.errors.append(ParseError(
                severity=Severity.WARNING,
                stage="normalize",
                message=f"Dependency graph building failed: {e}",
            ))

        # Detect pivot tables (presence only)
        try:
            result.pivot_table_ranges = self._detect_pivots(wb_formula)
        except Exception:
            pass

        # Close workbooks
        wb_formula.close()
        if wb_computed:
            wb_computed.close()

        # Finalize all IDs and hashes
        result.finalize()

        elapsed = (time.monotonic() - start_time) * 1000
        result.parse_duration_ms = elapsed
        logger.info(
            "Workbook parse complete: %s — %d sheets, %d cells, %d tables, %.0fms",
            self._filename,
            result.total_sheets,
            result.total_cells,
            len(result.tables),
            elapsed,
        )

        return result

    def _read_bytes(self) -> bytes:
        """Read the raw file bytes for hashing."""
        if self._content:
            return self._content
        if self._path:
            return self._path.read_bytes()
        raise ValueError("No file source available")

    def _extract_properties(self, wb, has_macros: bool) -> WorkbookProperties:
        """Extract workbook-level properties from openpyxl."""
        props = wb.properties
        calc = wb.calculation
        return WorkbookProperties(
            creator=props.creator if props else None,
            last_modified_by=props.lastModifiedBy if props else None,
            created=props.created if props else None,
            modified=props.modified if props else None,
            title=props.title if props else None,
            subject=props.subject if props else None,
            description=props.description if props else None,
            keywords=props.keywords if props else None,
            category=props.category if props else None,
            calc_mode=calc.calcMode if calc else None,
            iterate_enabled=bool(calc.iterate) if calc else False,
            iterate_count=calc.iterateCount if calc and calc.iterateCount else None,
            has_macros=has_macros,
            has_vba_project=has_macros,
        )

    def _extract_named_ranges(self, wb) -> list[NamedRangeDTO]:
        """Extract defined names / named ranges from the workbook."""
        named_ranges = []
        if not wb.defined_names:
            return named_ranges

        # openpyxl 3.1.x uses a dict-like DefinedNameDict
        for name, defn in wb.defined_names.items():
            try:
                value = defn.value  # e.g., "Sheet1!$A$1:$B$10"
                scope_sheet = None

                # Determine scope
                if hasattr(defn, "localSheetId") and defn.localSheetId is not None:
                    scope_idx = int(defn.localSheetId)
                    if scope_idx < len(wb.sheetnames):
                        scope_sheet = wb.sheetnames[scope_idx]

                is_hidden = bool(defn.hidden) if hasattr(defn, "hidden") else False

                named_ranges.append(NamedRangeDTO(
                    name=name,
                    ref_string=value or "",
                    scope_sheet=scope_sheet,
                    is_hidden=is_hidden,
                ))
            except Exception as e:
                logger.warning("Failed to parse defined name: %s", e)

        return named_ranges

    def _extract_external_links(self, wb) -> list[ExternalLink]:
        """Detect external workbook links."""
        links = []
        # openpyxl doesn't fully expose external links;
        # we detect them from defined names with [1], [2], etc. references
        if wb.defined_names:
            for _name, defn in wb.defined_names.items():
                if defn.value and "[" in defn.value:
                    links.append(ExternalLink(
                        link_index=len(links),
                        target_path=defn.value,
                        link_type="workbook",
                    ))
        return links

    def _detect_pivots(self, wb) -> list[dict]:
        """Detect pivot table presence and extract basic metadata."""
        pivots = []
        for ws in wb.worksheets:
            if hasattr(ws, "_pivots") and ws._pivots:
                for pivot in ws._pivots:
                    pivots.append({
                        "sheet_name": ws.title,
                        "name": getattr(pivot, "name", "unknown"),
                        "location": str(getattr(pivot, "location", "")),
                    })
        return pivots
