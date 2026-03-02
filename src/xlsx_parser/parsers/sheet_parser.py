"""
Sheet-level parsing: extract all data from a single worksheet.

Orchestrates cell parsing, merge resolution, conditional formatting,
data validation, and sheet property extraction. Designed for
independent execution to support parallel sheet processing.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from ..models.cell import CellDTO
from ..models.common import CellCoord, CellRange, ParseError, Severity
from ..models.sheet import (
    ConditionalFormatRule,
    DataValidationRule,
    MergedRegion,
    SheetDTO,
    SheetProperties,
)
from .cell_parser import CellParser

logger = logging.getLogger(__name__)


class SheetParser:
    """
    Parses a single openpyxl Worksheet into a SheetDTO.

    Handles cell extraction, merge resolution, property extraction,
    conditional formatting, data validations, and used-range detection.
    Designed for single-sheet processing; one instance per sheet.
    """

    def __init__(
        self,
        ws: OpenpyxlWorksheet,
        sheet_index: int,
        computed_ws: OpenpyxlWorksheet | None = None,
        max_cells: int = 2_000_000,
    ):
        """
        Args:
            ws: The openpyxl worksheet (with formulas).
            sheet_index: 0-based sheet index in the workbook.
            computed_ws: The same sheet opened with data_only=True for computed values.
            max_cells: Safety limit on cell count to prevent memory issues.
        """
        self._ws = ws
        self._sheet_index = sheet_index
        self._computed_ws = computed_ws
        self._max_cells = max_cells
        self._sheet_name = ws.title
        self._cell_parser = CellParser(self._sheet_name)

    def parse(self) -> SheetDTO:
        """
        Parse the worksheet into a SheetDTO.

        Returns a fully populated SheetDTO with cells, merges,
        properties, conditional formats, and data validations.
        """
        logger.info("Parsing sheet: %s (index=%d)", self._sheet_name, self._sheet_index)

        sheet = SheetDTO(
            sheet_name=self._sheet_name,
            sheet_index=self._sheet_index,
        )

        # Extract properties first
        sheet.properties = self._extract_properties()

        # Extract merged regions
        sheet.merged_regions = self._extract_merges()

        # Build merge lookup for cell parsing
        merge_masters = self._build_merge_lookup(sheet.merged_regions)

        # Extract cells
        self._extract_cells(sheet, merge_masters)

        # Extract row heights and column widths
        self._extract_dimensions(sheet)

        # Extract hidden rows/cols
        self._extract_hidden(sheet)

        # Extract conditional formatting
        sheet.conditional_format_rules = self._extract_conditional_formats()

        # Extract data validations
        sheet.data_validations = self._extract_data_validations()

        logger.info(
            "Sheet %s parsed: %d cells, %d merges",
            self._sheet_name,
            sheet.cell_count(),
            len(sheet.merged_regions),
        )
        return sheet

    def _extract_cells(
        self,
        sheet: SheetDTO,
        merge_masters: dict[tuple[int, int], tuple[CellCoord, int, int]],
    ) -> None:
        """
        Extract all non-empty cells from the worksheet.

        Uses the worksheet's iter_rows to efficiently scan the used range.
        Applies merge master/slave annotations and computed values.
        """
        cell_count = 0
        for row in self._ws.iter_rows():
            for cell in row:
                if cell_count >= self._max_cells:
                    sheet.errors.append(ParseError(
                        severity=Severity.WARNING,
                        stage="parse",
                        message=f"Cell limit ({self._max_cells}) reached; truncating",
                        sheet_name=self._sheet_name,
                    ))
                    return

                # Skip truly empty cells (no value, no formula, no style worth capturing)
                if cell.value is None and cell.data_type != "f" and not self._has_meaningful_style(cell):
                    # But still capture merged slaves
                    from openpyxl.cell.cell import MergedCell as MergedCellType
                    if not isinstance(cell, MergedCellType):
                        continue

                # Get computed value from data_only pass
                computed_value = None
                if self._computed_ws:
                    try:
                        computed_cell = self._computed_ws.cell(
                            row=cell.row, column=cell.column
                        )
                        computed_value = computed_cell.value
                    except Exception:
                        pass

                cell_dto = self._cell_parser.parse(cell, computed_value)

                # Annotate merge info
                key = (cell.row, cell.column)
                if key in merge_masters:
                    master_coord, row_span, col_span = merge_masters[key]
                    if master_coord.row == cell.row and master_coord.col == cell.column:
                        cell_dto.is_merged_master = True
                        cell_dto.merge_extent = row_span
                        cell_dto.merge_col_extent = col_span
                    else:
                        cell_dto.is_merged_slave = True
                        cell_dto.merge_master = master_coord

                if not cell_dto.is_empty or cell_dto.is_merged_slave or cell_dto.is_merged_master:
                    sheet.set_cell(cell_dto)
                    cell_count += 1

    def _extract_merges(self) -> list[MergedRegion]:
        """Extract all merged cell regions from the worksheet."""
        regions = []
        for merge_range in self._ws.merged_cells.ranges:
            bounds = merge_range.bounds  # (min_col, min_row, max_col, max_row)
            min_col, min_row, max_col, max_row = bounds
            cell_range = CellRange(
                top_left=CellCoord(row=min_row, col=min_col),
                bottom_right=CellCoord(row=max_row, col=max_col),
            )
            master = CellCoord(row=min_row, col=min_col)
            regions.append(MergedRegion(range=cell_range, master=master))
        return regions

    def _build_merge_lookup(
        self, regions: list[MergedRegion]
    ) -> dict[tuple[int, int], tuple[CellCoord, int, int]]:
        """
        Build a lookup dict mapping (row, col) → (master_coord, row_span, col_span)
        for all cells in any merged region.
        """
        lookup: dict[tuple[int, int], tuple[CellCoord, int, int]] = {}
        for region in regions:
            row_span = region.range.row_count()
            col_span = region.range.col_count()
            for r in range(region.range.top_left.row, region.range.bottom_right.row + 1):
                for c in range(region.range.top_left.col, region.range.bottom_right.col + 1):
                    lookup[(r, c)] = (region.master, row_span, col_span)
        return lookup

    def _extract_properties(self) -> SheetProperties:
        """Extract sheet-level properties."""
        ws = self._ws
        freeze_pane = None
        if ws.freeze_panes:
            freeze_pane = str(ws.freeze_panes)

        print_area = None
        if ws.print_area:
            print_area = str(ws.print_area)

        auto_filter = None
        if ws.auto_filter and ws.auto_filter.ref:
            auto_filter = str(ws.auto_filter.ref)

        tab_color = None
        if ws.sheet_properties and ws.sheet_properties.tabColor:
            tab_color = str(ws.sheet_properties.tabColor.rgb) if ws.sheet_properties.tabColor.rgb else None

        is_hidden = False
        if hasattr(ws, "sheet_state"):
            is_hidden = ws.sheet_state == "hidden"

        return SheetProperties(
            is_hidden=is_hidden,
            tab_color=tab_color,
            default_row_height=ws.sheet_format.defaultRowHeight if ws.sheet_format else None,
            default_col_width=ws.sheet_format.defaultColWidth if ws.sheet_format else None,
            freeze_pane=freeze_pane,
            print_area=print_area,
            auto_filter_range=auto_filter,
            sheet_protection=bool(ws.protection and ws.protection.sheet),
        )

    def _extract_dimensions(self, sheet: SheetDTO) -> None:
        """Extract custom row heights and column widths."""
        for row_idx, rd in self._ws.row_dimensions.items():
            if rd.height and rd.height != self._ws.sheet_format.defaultRowHeight:
                sheet.row_heights[row_idx] = rd.height

        for col_letter, cd in self._ws.column_dimensions.items():
            if cd.width:
                from ..models.common import col_letter_to_number
                col_num = col_letter_to_number(col_letter)
                sheet.col_widths[col_num] = cd.width * 7.5  # chars to points approx

    def _extract_hidden(self, sheet: SheetDTO) -> None:
        """Detect hidden rows and columns."""
        for row_idx, rd in self._ws.row_dimensions.items():
            if rd.hidden:
                sheet.hidden_rows.add(row_idx)

        for col_letter, cd in self._ws.column_dimensions.items():
            if cd.hidden:
                from ..models.common import col_letter_to_number
                col_num = col_letter_to_number(col_letter)
                sheet.hidden_cols.add(col_num)

    def _extract_conditional_formats(self) -> list[ConditionalFormatRule]:
        """Extract conditional formatting rules."""
        rules = []
        for cf in self._ws.conditional_formatting:
            for rule in cf.rules:
                formula = None
                if rule.formula and len(rule.formula) > 0:
                    formula = str(rule.formula[0])

                ranges = [str(r) for r in cf.cells.ranges] if hasattr(cf.cells, "ranges") else [str(cf.cells)]

                rules.append(ConditionalFormatRule(
                    ranges=ranges,
                    rule_type=rule.type or "unknown",
                    operator=rule.operator,
                    formula=formula,
                    priority=rule.priority,
                    stop_if_true=bool(rule.stopIfTrue),
                ))
        return rules

    def _extract_data_validations(self) -> list[DataValidationRule]:
        """Extract data validation rules."""
        rules = []
        if not self._ws.data_validations:
            return rules
        for dv in self._ws.data_validations.dataValidation:
            ranges = [str(r) for r in dv.cells.ranges] if hasattr(dv.cells, "ranges") else [str(dv.cells)] if dv.cells else []
            rules.append(DataValidationRule(
                ranges=ranges,
                validation_type=dv.type or "none",
                operator=dv.operator,
                formula1=str(dv.formula1) if dv.formula1 else None,
                formula2=str(dv.formula2) if dv.formula2 else None,
                allow_blank=bool(dv.allow_blank) if dv.allow_blank is not None else True,
                show_error_message=bool(dv.showErrorMessage),
                error_title=dv.errorTitle,
                error_message=dv.error,
                prompt_title=dv.promptTitle,
                prompt_message=dv.prompt,
            ))
        return rules

    @staticmethod
    def _has_meaningful_style(cell) -> bool:
        """Check if a cell has non-default styling worth preserving."""
        try:
            if cell.font and (cell.font.bold or cell.font.italic or cell.font.color):
                return True
            if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
                return True
            if cell.border:
                for side in ("left", "right", "top", "bottom"):
                    s = getattr(cell.border, side, None)
                    if s and s.style:
                        return True
        except Exception:
            pass
        return False
