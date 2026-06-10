"""
Legacy ``.xls`` (BIFF8) → ``.xlsx`` (OOXML) in-memory converter.

The whole parser is built on :mod:`openpyxl`, which only reads the OOXML
``.xlsx``/``.xlsm`` container. Legacy ``.xls`` workbooks use the binary BIFF
format that openpyxl cannot open. Rather than threading a second reader through
every downstream stage (cells, styles, merges, tables, charts, formulas...),
we convert ``.xls`` bytes to an equivalent ``.xlsx`` workbook **once, up front**
and let the existing pipeline run unchanged.

Conversion is **tiered** — it picks the best backend available at runtime:

  1. **LibreOffice (headless)** — the full-fidelity path. LibreOffice has a
     complete BIFF reader and writes real OOXML, so it preserves **formula
     text**, cached values, charts, shapes, pivots, and all styling. Used
     automatically whenever a ``soffice``/``libreoffice`` binary is found,
     unless disabled (see below). This is the recommended path for workbooks
     that carry formulas you need to reason about.

  2. **xlrd (pure-Python fallback)** — used when LibreOffice is unavailable or
     its conversion fails. Built on :mod:`xlrd` (the de-facto legacy ``.xls``
     reader). Preserves, best-effort: every sheet (name + hidden state); cell
     values with correct types (text, number, date/datetime, bool, error);
     number formats; merged ranges; basic font styling and fill color; column
     widths and hidden rows/columns. Two limitations are **inherent** to this
     path and surfaced to the caller as parse warnings:

       * **Formula text is not recoverable.** BIFF stores formulas as parsed
         token streams that ``xlrd`` does not re-serialize; only the *cached
         result* value survives.
       * **Charts, shapes, and pivot caches are dropped.** ``xlrd`` does not
         expose drawing objects.

Controlling the backend (env vars):

  * ``EXCEL_PARSER_SOFFICE_BIN`` — explicit path to the LibreOffice binary.
  * ``EXCEL_PARSER_DISABLE_SOFFICE`` — set truthy to force the xlrd path
    (e.g. in sandboxes where spawning a subprocess is undesirable).
  * ``EXCEL_PARSER_SOFFICE_TIMEOUT`` — per-conversion timeout in seconds
    (default 120).

If neither backend can run, :func:`convert_xls_to_xlsx` raises
:class:`XlsConversionError` with an actionable message.
"""

from __future__ import annotations

import datetime
import io
import logging
import os
import platform
import shutil
import subprocess
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)

# OLE2 / Compound File Binary signature that every BIFF ``.xls`` begins with.
# (Also shared by legacy .doc/.ppt and encrypted OOXML — extension is the
# primary signal; this is only a fallback for content with no useful name.)
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_DEFAULT_SOFFICE_TIMEOUT = 120

try:
    import xlrd  # type: ignore[import-untyped]

    _HAS_XLRD = True
except ImportError:  # pragma: no cover - optional dependency
    xlrd = None  # type: ignore[assignment]
    _HAS_XLRD = False


class XlsConversionError(RuntimeError):
    """Raised when a legacy ``.xls`` workbook cannot be converted."""


def is_legacy_xls(
    *,
    filename: str | None = None,
    path: str | Path | None = None,
    content: bytes | None = None,
) -> bool:
    """Return True when the input looks like a legacy BIFF ``.xls`` workbook.

    Decision order:
      1. A ``.xls`` extension (and *not* ``.xlsx``/``.xlsm``) on the filename or
         path is treated as authoritative.
      2. Otherwise, if raw bytes are available with no informative extension,
         fall back to the OLE2 compound-file magic. Encrypted ``.xlsx`` also
         carries this magic, so we only trust it when the name is uninformative.
    """
    name = (filename or (Path(path).name if path else "")).lower()
    if name.endswith((".xlsx", ".xlsm", ".xlsb")):
        return False
    if name.endswith(".xls"):
        return True

    if content is not None:
        return content[:8] == _OLE2_MAGIC
    return False


def convert_xls_to_xlsx(
    *,
    path: str | Path | None = None,
    content: bytes | None = None,
    prefer_libreoffice: bool = True,
) -> tuple[bytes, list[str]]:
    """Convert a legacy ``.xls`` workbook to ``.xlsx`` bytes.

    Exactly one of ``path`` / ``content`` should be supplied.

    Tries the full-fidelity LibreOffice backend first (preserves formula text,
    charts, shapes, pivots) and transparently falls back to the pure-Python
    xlrd backend (values + styling only) when LibreOffice is unavailable,
    disabled, or fails. See the module docstring for the env-var controls.

    Args:
        path: Path to the ``.xls`` file.
        content: Raw ``.xls`` bytes (alternative to ``path``).
        prefer_libreoffice: When True (default), attempt the LibreOffice path
            first. Set False to force the xlrd path.

    Returns:
        ``(xlsx_bytes, warnings)`` — the converted workbook as OOXML bytes, plus
        a list of human-readable warning strings describing anything dropped in
        translation. Empty on the LibreOffice path (full fidelity); the xlrd
        path appends a formula/charts caveat. Never ``None``.

    Raises:
        XlsConversionError: if no backend can convert the file.
    """
    if prefer_libreoffice and not _soffice_disabled():
        soffice = _find_soffice()
        if soffice is not None:
            try:
                xlsx_bytes = _convert_via_libreoffice(soffice, path=path, content=content)
                logger.info("Converted .xls via LibreOffice (full fidelity)")
                return xlsx_bytes, []
            except Exception as exc:  # noqa: BLE001 - fall back to xlrd
                logger.warning(
                    "LibreOffice .xls conversion failed (%s); falling back to xlrd "
                    "(formula text and charts will be lost)", exc,
                )

    return _convert_via_xlrd(path=path, content=content)


def _convert_via_xlrd(
    *,
    path: str | Path | None = None,
    content: bytes | None = None,
) -> tuple[bytes, list[str]]:
    """Pure-Python ``.xls`` → ``.xlsx`` conversion (values + styling only).

    Raises:
        XlsConversionError: if ``xlrd`` is unavailable or the file cannot be
            opened/converted.
    """
    if not _HAS_XLRD:
        raise XlsConversionError(
            "Reading legacy .xls workbooks requires the 'xlrd' package. "
            "Install it with: pip install xlrd"
        )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    warnings: list[str] = []

    try:
        if content is not None:
            book = xlrd.open_workbook(file_contents=content, formatting_info=True)
        else:
            book = xlrd.open_workbook(str(path), formatting_info=True)
    except NotImplementedError:
        # Some .xls variants (very old BIFF or unusual records) refuse
        # formatting_info; retry for values only rather than failing outright.
        warnings.append("Could not read .xls cell formatting; converted values only.")
        try:
            if content is not None:
                book = xlrd.open_workbook(file_contents=content, formatting_info=False)
            else:
                book = xlrd.open_workbook(str(path), formatting_info=False)
        except Exception as exc:  # noqa: BLE001
            raise XlsConversionError(f"Failed to open .xls workbook: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise XlsConversionError(f"Failed to open .xls workbook: {exc}") from exc

    has_formatting = bool(getattr(book, "formatting_info", False))
    out = Workbook()
    # Remove the default sheet openpyxl creates; we add sheets explicitly.
    out.remove(out.active)

    for sheet_idx in range(book.nsheets):
        xls_sheet = book.sheet_by_index(sheet_idx)
        ws = out.create_sheet(title=_safe_sheet_name(xls_sheet.name, sheet_idx))

        # Sheet visibility (xlrd: 0 visible, 1 hidden, 2 very hidden).
        visibility = getattr(xls_sheet, "visibility", 0)
        if visibility == 1:
            ws.sheet_state = "hidden"
        elif visibility == 2:
            ws.sheet_state = "veryHidden"

        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                cell_type = xls_sheet.cell_type(row, col)
                if cell_type in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    # Still apply styling to blank-but-styled cells below, but
                    # skip writing a value.
                    value: object = None
                else:
                    value = _coerce_value(book, xls_sheet, row, col, cell_type)

                # openpyxl is 1-indexed.
                target = ws.cell(row=row + 1, column=col + 1)
                if value is not None:
                    target.value = value

                if has_formatting:
                    _apply_style(book, xls_sheet, row, col, target, Font, PatternFill)

        # Merged cell ranges. xlrd gives half-open (rlo, rhi, clo, chi).
        for rlo, rhi, clo, chi in getattr(xls_sheet, "merged_cells", []):
            try:
                ws.merge_cells(
                    start_row=rlo + 1,
                    start_column=clo + 1,
                    end_row=rhi,
                    end_column=chi,
                )
            except Exception:  # noqa: BLE001 - never let a bad range abort conversion
                pass

        _apply_dimensions(xls_sheet, ws, has_formatting, get_column_letter)

    # xlrd cannot recover formula text or drawing objects from BIFF. Point the
    # user at the full-fidelity backend, which does.
    warnings.append(
        "Legacy .xls converted via the xlrd fallback: formula text is not "
        "recoverable (cached values preserved); charts and shapes are dropped. "
        "Install LibreOffice for full-fidelity .xls conversion (formulas, charts)."
    )

    buf = io.BytesIO()
    out.save(buf)
    return buf.getvalue(), warnings


# ---------------------------------------------------------------------------
# LibreOffice (full-fidelity) backend
# ---------------------------------------------------------------------------


def libreoffice_available() -> bool:
    """True when a usable LibreOffice binary is found and not disabled."""
    return not _soffice_disabled() and _find_soffice() is not None


def _soffice_disabled() -> bool:
    """Honor ``EXCEL_PARSER_DISABLE_SOFFICE`` for sandboxes that forbid subprocesses."""
    return os.environ.get("EXCEL_PARSER_DISABLE_SOFFICE", "").strip().lower() in {
        "1", "true", "yes", "on",
    }


def _find_soffice() -> str | None:
    """Locate the LibreOffice/OpenOffice ``soffice`` binary.

    Order: explicit ``EXCEL_PARSER_SOFFICE_BIN`` env var, then ``PATH``, then
    a few well-known per-platform install locations. Returns ``None`` if none
    is found.
    """
    override = os.environ.get("EXCEL_PARSER_SOFFICE_BIN", "").strip()
    if override:
        return override if os.path.isfile(override) else None

    for name in ("soffice", "libreoffice", "soffice.bin"):
        found = shutil.which(name)
        if found:
            return found

    system = platform.system()
    candidates: tuple[str, ...]
    if system == "Darwin":
        candidates = ("/Applications/LibreOffice.app/Contents/MacOS/soffice",)
    elif system == "Windows":  # pragma: no cover - platform-specific
        candidates = (
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        )
    else:  # Linux / other unixes
        candidates = (
            "/usr/bin/soffice",
            "/usr/bin/libreoffice",
            "/usr/local/bin/soffice",
            "/opt/libreoffice/program/soffice",
            "/snap/bin/libreoffice",
        )
    for cand in candidates:
        if os.path.isfile(cand):
            return cand
    return None


def _convert_via_libreoffice(
    soffice: str,
    *,
    path: str | Path | None = None,
    content: bytes | None = None,
) -> bytes:
    """Convert ``.xls`` → ``.xlsx`` by shelling out to headless LibreOffice.

    Returns the converted ``.xlsx`` bytes. Raises on any failure so the caller
    can fall back to the xlrd path. Each call uses a throwaway temp directory
    for both the user profile and the output, so concurrent conversions (and a
    LibreOffice the user may already have open) never collide.
    """
    try:
        timeout = int(os.environ.get("EXCEL_PARSER_SOFFICE_TIMEOUT", "").strip())
    except ValueError:
        timeout = _DEFAULT_SOFFICE_TIMEOUT
    if timeout <= 0:
        timeout = _DEFAULT_SOFFICE_TIMEOUT

    with tempfile.TemporaryDirectory(prefix="xls2xlsx_") as workdir:
        # Source file: use the provided path, or spill content to disk —
        # soffice is strictly file-based.
        if path is not None:
            src = str(path)
        else:
            src = os.path.join(workdir, "input.xls")
            with open(src, "wb") as f:
                f.write(content or b"")

        outdir = os.path.join(workdir, "out")
        os.makedirs(outdir, exist_ok=True)
        # A private, per-call user profile avoids the "another instance is
        # running" lock and keeps headless runs hermetic.
        profile = Path(workdir, "profile").as_uri()

        cmd = [
            soffice,
            f"-env:UserInstallation={profile}",
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            outdir,
            src,
        ]
        try:
            proc = subprocess.run(
                cmd,
                capture_output=True,
                timeout=timeout,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            raise XlsConversionError(
                f"LibreOffice conversion timed out after {timeout}s"
            ) from exc
        except OSError as exc:
            raise XlsConversionError(f"Could not launch LibreOffice: {exc}") from exc

        produced = list(Path(outdir).glob("*.xlsx"))
        if proc.returncode != 0 or not produced:
            stderr = proc.stderr.decode("utf-8", "replace").strip()[:500]
            raise XlsConversionError(
                f"LibreOffice exited {proc.returncode} without output. {stderr}"
            )

        return produced[0].read_bytes()


# ---------------------------------------------------------------------------
# Internal helpers (xlrd backend)
# ---------------------------------------------------------------------------


def _safe_sheet_name(name: str, idx: int) -> str:
    """Coerce an .xls sheet name into a legal openpyxl title."""
    name = (name or f"Sheet{idx + 1}").strip()
    # openpyxl forbids these characters and caps length at 31.
    for ch in r"\/*?:[]":
        name = name.replace(ch, "_")
    name = name[:31] or f"Sheet{idx + 1}"
    return name


def _coerce_value(book, sheet, row, col, cell_type):
    """Translate an xlrd cell value into a native Python value openpyxl stores."""
    raw = sheet.cell_value(row, col)

    if cell_type == xlrd.XL_CELL_NUMBER:
        return raw
    if cell_type == xlrd.XL_CELL_TEXT:
        return raw
    if cell_type == xlrd.XL_CELL_BOOLEAN:
        return bool(raw)
    if cell_type == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate_as_datetime(raw, book.datemode)
        except Exception:  # noqa: BLE001 - bad serial date; keep the number
            return raw
        # A pure date (midnight, no fractional day) → date, else datetime.
        if dt.time() == datetime.time(0, 0) and float(raw).is_integer():
            return dt.date()
        return dt
    if cell_type == xlrd.XL_CELL_ERROR:
        # xlrd stores the BIFF error code; map to the Excel display text.
        return xlrd.error_text_from_code.get(raw, "#ERR!")
    return raw


def _apply_style(book, sheet, row, col, target, Font, PatternFill) -> None:
    """Best-effort port of number format + font + fill onto an openpyxl cell.

    Any failure is swallowed — styling must never abort a conversion.
    """
    try:
        xf_index = sheet.cell_xf_index(row, col)
    except Exception:  # noqa: BLE001 - formatting_info missing for this cell
        return

    try:
        xf = book.xf_list[xf_index]
    except (IndexError, AttributeError):
        return

    # Number format string.
    try:
        fmt = book.format_map.get(xf.format_key)
        if fmt is not None and fmt.format_str and fmt.format_str != "General":
            target.number_format = fmt.format_str
    except Exception:  # noqa: BLE001
        pass

    # Font: bold / italic / size / name / color.
    try:
        font = book.font_list[xf.font_index]
        color = _argb_from_index(book, font.colour_index)
        target.font = Font(
            name=font.name or None,
            size=(font.height / 20.0) if font.height else None,
            bold=bool(font.weight and font.weight >= 700) or bool(font.bold),
            italic=bool(font.italic),
            color=color,
        )
    except Exception:  # noqa: BLE001
        pass

    # Fill (solid background pattern only — the common case).
    try:
        bg = xf.background
        # pattern 1 == solid fill in BIFF.
        if bg and bg.fill_pattern == 1:
            fill_color = _argb_from_index(book, bg.pattern_colour_index)
            if fill_color:
                target.fill = PatternFill(
                    fill_type="solid", fgColor=fill_color, bgColor=fill_color
                )
    except Exception:  # noqa: BLE001
        pass


def _argb_from_index(book, colour_index) -> str | None:
    """Resolve an xlrd palette colour index to an ``AARRGGBB`` hex string."""
    if colour_index is None:
        return None
    try:
        rgb = book.colour_map.get(colour_index)
    except Exception:  # noqa: BLE001
        return None
    if not rgb:
        return None
    r, g, b = rgb
    return f"FF{r:02X}{g:02X}{b:02X}"


def _apply_dimensions(sheet, ws, has_formatting, get_column_letter) -> None:
    """Port column widths + hidden row/column flags where available."""
    if not has_formatting:
        return
    # Columns.
    try:
        for col_idx, info in getattr(sheet, "colinfo_map", {}).items():
            letter = get_column_letter(col_idx + 1)
            dim = ws.column_dimensions[letter]
            if info.hidden:
                dim.hidden = True
            # xlrd width is in 1/256 of a character; openpyxl width is chars.
            if info.width:
                dim.width = info.width / 256.0
    except Exception:  # noqa: BLE001
        pass
    # Rows.
    try:
        for row_idx, info in getattr(sheet, "rowinfo_map", {}).items():
            dim = ws.row_dimensions[row_idx + 1]
            if info.hidden:
                dim.hidden = True
            if info.height and getattr(info, "has_default_height", 0) == 0:
                dim.height = info.height / 20.0
    except Exception:  # noqa: BLE001
        pass


__all__ = [
    "convert_xls_to_xlsx",
    "is_legacy_xls",
    "libreoffice_available",
    "XlsConversionError",
]
