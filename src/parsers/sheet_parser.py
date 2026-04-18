"""
Sheet-level parsing: extract all data from a single worksheet.

Orchestrates cell parsing, merge resolution, conditional formatting,
data validation, and sheet property extraction. Designed for
independent execution to support parallel sheet processing.
"""

from __future__ import annotations

import io
import logging
import re
import zipfile
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from models.cell import CellDTO
from models.common import CellCoord, CellRange, ParseError, Severity, col_letter_to_number

# OOXML namespace for spreadsheetML
_OOXML_NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

# Regex to split an A1-style cell reference like "B1" into ("B", "1")
_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$")
from models.sheet import (
    ConditionalFormatRule,
    DataValidationRule,
    MergedRegion,
    SheetDTO,
    SheetProperties,
)
from .cell_parser import CellParser

logger = logging.getLogger(__name__)


class SheetParser:
    """
    Parses a single openpyxl Worksheet into a SheetDTO.

    Handles cell extraction, merge resolution, property extraction,
    conditional formatting, data validations, and used-range detection.
    Designed for single-sheet processing; one instance per sheet.
    """

    def __init__(
        self,
        ws: OpenpyxlWorksheet,
        sheet_index: int,
        computed_ws: OpenpyxlWorksheet | None = None,
        computed_values: dict[tuple[int, int], Any] | None = None,
        max_cells: int = 2_000_000,
        workbook_path: Path | None = None,
        workbook_content: bytes | None = None,
    ):
        """
        Args:
            ws: The openpyxl worksheet (with formulas).
            sheet_index: 0-based sheet index in the workbook.
            computed_ws: The same sheet opened with data_only=True for computed values.
                Used only when `computed_values` is None (fallback path).
            computed_values: Cached formula values from calamine, keyed by
                ``(row_1idx, col_1idx)``. Preferred: avoids a second openpyxl load.
            max_cells: Safety limit on cell count to prevent memory issues.
            workbook_path: Path to the .xlsx file (for raw OOXML fallback).
            workbook_content: Raw bytes of the .xlsx file (for raw OOXML fallback).
        """
        self._ws = ws
        self._sheet_index = sheet_index
        self._computed_ws = computed_ws
        self._computed_values = computed_values
        self._max_cells = max_cells
        self._sheet_name = ws.title
        self._cell_parser = CellParser(self._sheet_name)
        self._workbook_path = workbook_path
        self._workbook_content = workbook_content

    def parse(self) -> SheetDTO:
        """
        Parse the worksheet into a SheetDTO.

        Returns a fully populated SheetDTO with cells, merges,
        properties, conditional formats, and data validations.
        """
        logger.info("Parsing sheet: %s (index=%d)", self._sheet_name, self._sheet_index)

        sheet = SheetDTO(
            sheet_name=self._sheet_name,
            sheet_index=self._sheet_index,
        )

        # Extract properties first
        sheet.properties = self._extract_properties()

        # Extract merged regions
        sheet.merged_regions = self._extract_merges()

        # Build merge lookup for cell parsing
        merge_masters = self._build_merge_lookup(sheet.merged_regions)

        # Extract cells
        self._extract_cells(sheet, merge_masters)

        # Recover values from empty merge masters via raw OOXML
        self._recover_empty_merge_masters(sheet)

        # Extract row heights and column widths
        self._extract_dimensions(sheet)

        # Extract hidden rows/cols
        self._extract_hidden(sheet)

        # Extract conditional formatting
        sheet.conditional_format_rules = self._extract_conditional_formats()

        # Extract data validations
        sheet.data_validations = self._extract_data_validations()

        # Extract autofilter state
        self._extract_autofilter(sheet)

        logger.info(
            "Sheet %s parsed: %d cells, %d merges",
            self._sheet_name,
            sheet.cell_count(),
            len(sheet.merged_regions),
        )
        return sheet

    def _extract_cells(
        self,
        sheet: SheetDTO,
        merge_masters: dict[tuple[int, int], tuple[CellCoord, int, int]],
    ) -> None:
        """
        Extract all non-empty cells from the worksheet.

        Uses the worksheet's iter_rows to efficiently scan the used range.
        Applies merge master/slave annotations and computed values.
        """
        cell_count = 0
        # Use openpyxl's internal cell storage when available so sparse sheets
        # (e.g. A1 + XFD1048576) don't force a walk over ~17B empty cells.
        # Fall back to iter_rows() for read-only mode where _cells is absent.
        stored_cells = getattr(self._ws, "_cells", None)
        if isinstance(stored_cells, dict):
            # Merged cells are not in _cells; materialise them separately.
            merge_keys = set()
            for (mr, mc), (master, _rs, _cs) in merge_masters.items():
                merge_keys.add((mr, mc))
            cell_iter = [
                self._ws.cell(row=r, column=c)
                for (r, c) in sorted(set(stored_cells.keys()) | merge_keys)
            ]
        else:
            cell_iter = (cell for row in self._ws.iter_rows() for cell in row)

        for cell in cell_iter:
            if cell_count >= self._max_cells:
                sheet.errors.append(ParseError(
                    severity=Severity.WARNING,
                    stage="parse",
                    message=f"Cell limit ({self._max_cells}) reached; truncating",
                    sheet_name=self._sheet_name,
                ))
                return

            # Skip truly empty cells (no value, no formula, no style worth capturing)
            if cell.value is None and cell.data_type != "f" and not self._has_meaningful_style(cell):
                # But still capture merged slaves
                from openpyxl.cell.cell import MergedCell as MergedCellType
                if not isinstance(cell, MergedCellType):
                    continue

            # Get computed value. Prefer the calamine-provided dict (populated
            # at workbook load time) over an openpyxl ``data_only=True`` lookup;
            # dict lookups are O(1) and calamine is ~30-167× faster on the
            # underlying XLSX parse.
            computed_value = None
            if self._computed_values is not None:
                computed_value = self._computed_values.get((cell.row, cell.column))
            elif self._computed_ws:
                try:
                    computed_cell = self._computed_ws.cell(
                        row=cell.row, column=cell.column
                    )
                    computed_value = computed_cell.value
                except Exception:
                    pass

            cell_dto = self._cell_parser.parse(cell, computed_value)

            # Annotate merge info
            key = (cell.row, cell.column)
            if key in merge_masters:
                master_coord, row_span, col_span = merge_masters[key]
                if master_coord.row == cell.row and master_coord.col == cell.column:
                    cell_dto.is_merged_master = True
                    cell_dto.merge_extent = row_span
                    cell_dto.merge_col_extent = col_span
                else:
                    cell_dto.is_merged_slave = True
                    cell_dto.merge_master = master_coord

            if not cell_dto.is_empty or cell_dto.is_merged_slave or cell_dto.is_merged_master:
                sheet.set_cell(cell_dto)
                cell_count += 1

    def _extract_merges(self) -> list[MergedRegion]:
        """Extract all merged cell regions from the worksheet."""
        regions = []
        for merge_range in self._ws.merged_cells.ranges:
            bounds = merge_range.bounds  # (min_col, min_row, max_col, max_row)
            min_col, min_row, max_col, max_row = bounds
            cell_range = CellRange(
                top_left=CellCoord(row=min_row, col=min_col),
                bottom_right=CellCoord(row=max_row, col=max_col),
            )
            master = CellCoord(row=min_row, col=min_col)
            regions.append(MergedRegion(range=cell_range, master=master))
        return regions

    def _build_merge_lookup(
        self, regions: list[MergedRegion]
    ) -> dict[tuple[int, int], tuple[CellCoord, int, int]]:
        """
        Build a lookup dict mapping (row, col) → (master_coord, row_span, col_span)
        for all cells in any merged region.
        """
        lookup: dict[tuple[int, int], tuple[CellCoord, int, int]] = {}
        for region in regions:
            row_span = region.range.row_count()
            col_span = region.range.col_count()
            for r in range(region.range.top_left.row, region.range.bottom_right.row + 1):
                for c in range(region.range.top_left.col, region.range.bottom_right.col + 1):
                    lookup[(r, c)] = (region.master, row_span, col_span)
        return lookup

    def _extract_properties(self) -> SheetProperties:
        """Extract sheet-level properties."""
        ws = self._ws
        freeze_pane = None
        if ws.freeze_panes:
            freeze_pane = str(ws.freeze_panes)

        print_area = None
        if ws.print_area:
            print_area = str(ws.print_area)

        auto_filter = None
        if ws.auto_filter and ws.auto_filter.ref:
            auto_filter = str(ws.auto_filter.ref)

        tab_color = None
        if ws.sheet_properties and ws.sheet_properties.tabColor:
            tab_color = str(ws.sheet_properties.tabColor.rgb) if ws.sheet_properties.tabColor.rgb else None

        is_hidden = False
        if hasattr(ws, "sheet_state"):
            is_hidden = ws.sheet_state == "hidden"

        return SheetProperties(
            is_hidden=is_hidden,
            tab_color=tab_color,
            default_row_height=ws.sheet_format.defaultRowHeight if ws.sheet_format else None,
            default_col_width=ws.sheet_format.defaultColWidth if ws.sheet_format else None,
            freeze_pane=freeze_pane,
            print_area=print_area,
            auto_filter_range=auto_filter,
            sheet_protection=bool(ws.protection and ws.protection.sheet),
        )

    def _extract_dimensions(self, sheet: SheetDTO) -> None:
        """Extract custom row heights and column widths."""
        for row_idx, rd in self._ws.row_dimensions.items():
            if rd.height and rd.height != self._ws.sheet_format.defaultRowHeight:
                sheet.row_heights[row_idx] = rd.height

        for col_letter, cd in self._ws.column_dimensions.items():
            if cd.width:
                from models.common import col_letter_to_number
                col_num = col_letter_to_number(col_letter)
                sheet.col_widths[col_num] = cd.width * 7.5  # chars to points approx

    def _extract_hidden(self, sheet: SheetDTO) -> None:
        """Detect hidden rows and columns."""
        for row_idx, rd in self._ws.row_dimensions.items():
            if rd.hidden:
                sheet.hidden_rows.add(row_idx)

        for col_letter, cd in self._ws.column_dimensions.items():
            if cd.hidden:
                from models.common import col_letter_to_number
                col_num = col_letter_to_number(col_letter)
                sheet.hidden_cols.add(col_num)

    def _extract_conditional_formats(self) -> list[ConditionalFormatRule]:
        """Extract conditional formatting rules."""
        rules = []
        for cf in self._ws.conditional_formatting:
            for rule in cf.rules:
                formula = None
                if rule.formula and len(rule.formula) > 0:
                    formula = str(rule.formula[0])

                ranges = [str(r) for r in cf.cells.ranges] if hasattr(cf.cells, "ranges") else [str(cf.cells)]

                rules.append(ConditionalFormatRule(
                    ranges=ranges,
                    rule_type=rule.type or "unknown",
                    operator=rule.operator,
                    formula=formula,
                    priority=rule.priority,
                    stop_if_true=bool(rule.stopIfTrue),
                ))
        return rules

    def _extract_data_validations(self) -> list[DataValidationRule]:
        """Extract data validation rules."""
        rules = []
        if not self._ws.data_validations:
            return rules
        for dv in self._ws.data_validations.dataValidation:
            ranges = [str(r) for r in dv.cells.ranges] if hasattr(dv.cells, "ranges") else [str(dv.cells)] if dv.cells else []
            rules.append(DataValidationRule(
                ranges=ranges,
                validation_type=dv.type or "none",
                operator=dv.operator,
                formula1=str(dv.formula1) if dv.formula1 else None,
                formula2=str(dv.formula2) if dv.formula2 else None,
                allow_blank=bool(dv.allow_blank) if dv.allow_blank is not None else True,
                show_error_message=bool(dv.showErrorMessage),
                error_title=dv.errorTitle,
                error_message=dv.error,
                prompt_title=dv.promptTitle,
                prompt_message=dv.prompt,
            ))
        return rules

    def _extract_autofilter(self, sheet: SheetDTO) -> None:
        """Extract autofilter range and criteria from the worksheet."""
        af = self._ws.auto_filter
        if not af or not af.ref:
            return
        try:
            from models.common import FilterCriteria
            ref = str(af.ref)
            parts = ref.split(":")
            if len(parts) == 2:
                from models.common import col_letter_to_number as c2n
                start_match = _CELL_REF_RE.match(parts[0])
                end_match = _CELL_REF_RE.match(parts[1])
                if start_match and end_match:
                    sheet.autofilter_range = CellRange(
                        top_left=CellCoord(
                            row=int(start_match.group(2)),
                            col=col_letter_to_number(start_match.group(1)),
                        ),
                        bottom_right=CellCoord(
                            row=int(end_match.group(2)),
                            col=col_letter_to_number(end_match.group(1)),
                        ),
                    )

            # Extract filter criteria from individual column filters
            if hasattr(af, "filterColumn") and af.filterColumn:
                for fc in af.filterColumn:
                    col_id = fc.colId if hasattr(fc, "colId") else 0
                    vals: list[str] = []
                    if hasattr(fc, "filters") and fc.filters:
                        for filt in fc.filters.filter:
                            if hasattr(filt, "val") and filt.val:
                                vals.append(str(filt.val))
                    if vals:
                        sheet.autofilter_criteria.append(FilterCriteria(
                            col_index=col_id,
                            filter_type="values",
                            values=vals,
                        ))
        except Exception as e:
            logger.debug("Could not fully parse autofilter: %s", e)

    # ------------------------------------------------------------------
    # Empty-master merge recovery via raw OOXML
    # ------------------------------------------------------------------

    def _recover_empty_merge_masters(self, sheet: SheetDTO) -> None:
        """
        Fix the merge_empty_master issue: when the master cell of a merged
        region has no value, scan the raw OOXML XML for values in any cell
        within the merge range and promote the first found value to the master.

        openpyxl's MergedCell class discards values from non-master cells,
        so this fallback reads the sheet XML directly from the .xlsx ZIP.
        """
        if not sheet.merged_regions:
            return
        if self._workbook_path is None and self._workbook_content is None:
            return

        # Identify empty-master regions
        empty_masters: list[tuple[CellRange, CellCoord]] = []
        for region in sheet.merged_regions:
            master_cell = sheet.get_cell(region.master.row, region.master.col)
            if master_cell is None or master_cell.raw_value is None:
                empty_masters.append((region.range, region.master))

        if not empty_masters:
            return

        logger.debug(
            "Sheet %s: %d empty-master merged regions detected, attempting OOXML recovery",
            self._sheet_name,
            len(empty_masters),
        )

        try:
            self._do_xml_recovery(sheet, empty_masters)
        except Exception as e:
            logger.warning(
                "OOXML merge recovery failed for sheet %s: %s",
                self._sheet_name, e,
            )
            sheet.errors.append(ParseError(
                severity=Severity.WARNING,
                stage="parse",
                message=f"Merge empty-master recovery failed: {e}",
                sheet_name=self._sheet_name,
            ))

    def _do_xml_recovery(
        self,
        sheet: SheetDTO,
        empty_masters: list[tuple[CellRange, CellCoord]],
    ) -> None:
        """Open the .xlsx ZIP and parse the sheet XML to recover values."""
        source: Any
        if self._workbook_path:
            source = str(self._workbook_path)
        else:
            source = io.BytesIO(self._workbook_content)

        with zipfile.ZipFile(source, "r") as zf:
            # Find the correct sheet XML path
            sheet_xml_path = self._find_sheet_xml_path(zf, self._sheet_index)
            if sheet_xml_path is None:
                return

            # Load shared strings (needed for t="s" cell values)
            shared_strings = self._load_shared_strings(zf)

            # Parse the sheet XML
            with zf.open(sheet_xml_path) as f:
                tree = etree.parse(f)  # noqa: S320
            root = tree.getroot()

            # Parse all cell values from XML
            xml_values = self._parse_cell_values_from_xml(root, shared_strings)

            # For each empty-master region, find a value in the range
            for cell_range, master_coord in empty_masters:
                value = self._find_value_in_region(xml_values, cell_range, master_coord)
                if value is not None:
                    self._update_master_cell(sheet, master_coord, value)
                    logger.debug(
                        "Recovered value for %s!%s from merged region %s",
                        self._sheet_name, master_coord.to_a1(), cell_range.to_a1(),
                    )

    @staticmethod
    def _find_sheet_xml_path(zf: zipfile.ZipFile, sheet_index: int) -> str | None:
        """Determine the sheet XML path inside the ZIP."""
        # Try workbook.xml.rels to find the correct mapping
        # Fallback: xl/worksheets/sheet{N}.xml (1-indexed)
        candidate = f"xl/worksheets/sheet{sheet_index + 1}.xml"
        if candidate in zf.namelist():
            return candidate

        # Try all sheet files and pick by index
        sheet_files = sorted(
            n for n in zf.namelist()
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
        )
        if sheet_index < len(sheet_files):
            return sheet_files[sheet_index]

        return None

    @staticmethod
    def _load_shared_strings(zf: zipfile.ZipFile) -> list[str]:
        """Parse xl/sharedStrings.xml to get the shared string table."""
        ss_path = "xl/sharedStrings.xml"
        if ss_path not in zf.namelist():
            return []

        with zf.open(ss_path) as f:
            tree = etree.parse(f)  # noqa: S320
        root = tree.getroot()

        strings: list[str] = []
        for si in root.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"):
            # Simple case: <si><t>text</t></si>
            t_elem = si.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
            if t_elem is not None and t_elem.text is not None:
                strings.append(t_elem.text)
            else:
                # Rich text case: <si><r><t>part1</t></r><r><t>part2</t></r></si>
                parts = []
                for r in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                    if r.text:
                        parts.append(r.text)
                strings.append("".join(parts))

        return strings

    @staticmethod
    def _cell_ref_to_coord(ref: str) -> tuple[int, int] | None:
        """Convert 'B1' to (row=1, col=2). Returns None if unparseable."""
        m = _CELL_REF_RE.match(ref)
        if not m:
            return None
        col = col_letter_to_number(m.group(1))
        row = int(m.group(2))
        return (row, col)

    @classmethod
    def _parse_cell_values_from_xml(
        cls,
        root: etree._Element,
        shared_strings: list[str],
    ) -> dict[tuple[int, int], Any]:
        """Extract all cell values from the sheet XML as {(row, col): value}."""
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        values: dict[tuple[int, int], Any] = {}

        for c_elem in root.iter(f"{{{ns}}}c"):
            ref = c_elem.get("r")
            if not ref:
                continue
            coord = cls._cell_ref_to_coord(ref)
            if coord is None:
                continue

            value = cls._resolve_xml_value(c_elem, shared_strings, ns)
            if value is not None:
                values[coord] = value

        return values

    @staticmethod
    def _resolve_xml_value(
        c_elem: etree._Element,
        shared_strings: list[str],
        ns: str,
    ) -> Any:
        """Resolve the value of a <c> element considering its type attribute."""
        cell_type = c_elem.get("t")
        v_elem = c_elem.find(f"{{{ns}}}v")

        if cell_type == "inlineStr":
            # Inline string: <c t="inlineStr"><is><t>text</t></is></c>
            is_elem = c_elem.find(f"{{{ns}}}is")
            if is_elem is not None:
                t_elem = is_elem.find(f"{{{ns}}}t")
                if t_elem is not None and t_elem.text is not None:
                    return t_elem.text
            return None

        if v_elem is None or v_elem.text is None:
            return None

        raw = v_elem.text

        if cell_type == "s":
            # Shared string index
            try:
                idx = int(raw)
                if 0 <= idx < len(shared_strings):
                    return shared_strings[idx]
            except (ValueError, IndexError):
                pass
            return None

        if cell_type == "str":
            return raw

        if cell_type == "b":
            return raw == "1"

        if cell_type == "e":
            return raw  # Error value like #REF!, #VALUE!, etc.

        # Default: number
        try:
            if "." in raw:
                return float(raw)
            return int(raw)
        except ValueError:
            return raw

    @staticmethod
    def _find_value_in_region(
        xml_values: dict[tuple[int, int], Any],
        cell_range: CellRange,
        master_coord: CellCoord,
    ) -> Any:
        """Find the first non-None value in a cell range, excluding the master."""
        for r in range(cell_range.top_left.row, cell_range.bottom_right.row + 1):
            for c in range(cell_range.top_left.col, cell_range.bottom_right.col + 1):
                if r == master_coord.row and c == master_coord.col:
                    continue
                value = xml_values.get((r, c))
                if value is not None:
                    return value
        # Also check master in XML (might have a value openpyxl missed)
        return xml_values.get((master_coord.row, master_coord.col))

    @staticmethod
    def _update_master_cell(sheet: SheetDTO, master_coord: CellCoord, value: Any) -> None:
        """Update or create the master cell with the recovered value."""
        master_cell = sheet.get_cell(master_coord.row, master_coord.col)
        if master_cell is not None:
            master_cell.raw_value = value
            master_cell.display_value = str(value) if value is not None else None
        else:
            # Create a new cell for the master
            new_cell = CellDTO(
                coord=master_coord,
                sheet_name=sheet.sheet_name,
                raw_value=value,
                display_value=str(value) if value is not None else None,
                is_merged_master=True,
            )
            sheet.set_cell(new_cell)

    @staticmethod
    def _has_meaningful_style(cell) -> bool:
        """Check if a cell has non-default styling worth preserving."""
        try:
            if cell.font and (cell.font.bold or cell.font.italic or cell.font.color):
                return True
            if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
                return True
            if cell.border:
                for side in ("left", "right", "top", "bottom"):
                    s = getattr(cell.border, side, None)
                    if s and s.style:
                        return True
        except Exception:
            pass
        return False
