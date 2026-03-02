#!/usr/bin/env python3
"""
Complex Excel Builder for Parser Stress Testing.

Builds progressively more complicated .xlsx files based on a level index.
Each level adds features that stress different parts of the parser pipeline.
Output includes a reference JSON describing what was built for validation.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# Output directory
STRESS_DIR = Path(__file__).parent


def _add_merge_across_sheet(wb: Workbook, ref: dict) -> None:
    """Add sheet with many horizontal merges (merge-across / merge columns pattern)."""
    ws = wb.create_sheet("MergeAcross")
    # 50 rows, each row has A:C merged — simulates Excel's "Merge Across"
    for r in range(1, 51):
        ws.merge_cells(f"A{r}:C{r}")
        ws.cell(row=r, column=1, value=f"Row{r} merged")
    ref["features"].append("merge_across_many")
    ref["expected"]["merged_regions_across"] = 50


def build_level(level: int, out_path: Path | None = None) -> dict:
    """
    Build an Excel workbook at the given complexity level.
    Returns a reference JSON describing what was built.
    """
    out_path = out_path or (STRESS_DIR / f"stress_level_{level}.xlsx")
    ref: dict = {"level": level, "features": [], "expected": {}}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Level 0: Minimal ---
    ws["A1"] = "Minimal"
    ws["B1"] = 1
    ref["features"].append("minimal_cells")
    ref["expected"]["cells"] = 2

    if level >= 1:
        # Level 1: Formulas, extra sheet
        ws["A2"] = "Sum"
        ws["B2"] = "=B1+10"
        ws2 = wb.create_sheet("Data")
        ws2["A1"] = 100
        ws2["B1"] = "=A1*2"
        ref["features"].extend(["formulas", "multi_sheet"])
        ref["expected"]["formulas"] = 2

    if level >= 2:
        # Level 2: Tables, merges, formatting
        ws.merge_cells("A3:C3")
        ws["A3"] = "Merged Header"
        ws["A3"].font = Font(bold=True, size=12)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws["A3"].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws["A3"].font = Font(bold=True, size=12, color="FFFFFF")
        ref["features"].extend(["merged_cells", "font_fill_alignment"])
        ref["expected"]["merged_regions"] = 1

        # Table
        for col, h in enumerate(["Prod", "Qty", "Price"], 1):
            ws.cell(row=4, column=col, value=h).font = Font(bold=True)
        for r in range(5, 8):
            ws.cell(row=r, column=1, value=f"P{r-4}")
            ws.cell(row=r, column=2, value=(r - 4) * 10)
            ws.cell(row=r, column=3, value=f"=B{r}*5")
        tab = Table(displayName=f"Table_L{level}", ref="A4:C7")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)
        ref["features"].append("listobject_table")
        ref["expected"]["tables"] = 1

    if level >= 3:
        # Level 3: Charts, conditional formatting
        chart = BarChart()
        chart.title = "Quantity Chart"
        data = Reference(ws, min_col=2, min_row=4, max_row=7)
        cats = Reference(ws, min_col=1, min_row=5, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E4")
        ref["features"].append("chart_bar")

        ws.conditional_formatting.add(
            "B5:B7",
            CellIsRule(operator="greaterThan", formula=["20"], fill=PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )),
        )
        ref["features"].append("conditional_formatting")

    if level >= 4:
        # Level 4: Data validation, comments
        dv = DataValidation(type="list", formula1='"Low,Medium,High"')
        ws.add_data_validation(dv)
        dv.add(ws["D5"])
        ws["D5"] = "Medium"
        ws["D5"].comment = Comment("Status level", "Builder")
        ref["features"].extend(["data_validation", "comments"])

        # Named ranges
        wb.defined_names.add(DefinedName("TotalQty", attr_text="Sheet1!$B$7"))
        ref["features"].append("named_ranges")

    if level >= 5:
        # Level 5: Cross-sheet formulas, more complex formulas
        ws3 = wb.create_sheet("Summary")
        ws3["A1"] = "From Data"
        ws3["B1"] = "=Data!B1"
        ws3["A2"] = "SUM from Sheet1"
        ws3["B2"] = "=SUM(Sheet1!B5:B7)"
        ref["features"].extend(["cross_sheet_formulas", "sum_range"])

    if level >= 6:
        # Level 6: Lookup formulas, nested IFs
        ws["A10"] = "Lookup ID"
        ws["B10"] = 2
        ws["A11"] = "VLOOKUP"
        ws["B11"] = '=VLOOKUP(B10,A5:C7,2,FALSE)'
        ws["A12"] = "Nested IF"
        ws["B12"] = '=IF(B11>20,"High",IF(B11>10,"Mid","Low"))'
        ref["features"].extend(["vlookup", "nested_if"])

    if level >= 7:
        # Level 7: Multiple charts, multiple tables, wide sheet
        line_chart = LineChart()
        line_chart.title = "Trend"
        line_chart.add_data(Reference(ws, min_col=2, min_row=4, max_row=7), titles_from_data=True)
        line_chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=7))
        ws.add_chart(line_chart, "E18")
        ref["features"].append("chart_line")

        # Second table
        ws.cell(row=15, column=1, value="Region").font = Font(bold=True)
        ws.cell(row=15, column=2, value="Sales").font = Font(bold=True)
        for r in range(16, 19):
            ws.cell(row=r, column=1, value=f"R{r-15}")
            ws.cell(row=r, column=2, value=(r - 15) * 1000)
        tab2 = Table(displayName=f"Sales_L{level}", ref="A15:B18")
        ws.add_table(tab2)
        ref["expected"]["tables"] = 2

        # Wide: 50 columns
        for c in range(20, 70):
            ws.cell(row=1, column=c, value=f"Col{c}").font = Font(bold=True)
            ws.cell(row=2, column=c, value=c * 10)
        ref["features"].extend(["wide_sheet", "multi_chart", "multi_table"])

    if level >= 8:
        # Level 8: Hidden rows/cols, freeze panes
        ws.row_dimensions[6].hidden = True
        ws.column_dimensions["F"].hidden = True
        ws.freeze_panes = "A5"
        ref["features"].extend(["hidden_rows", "hidden_cols", "freeze_panes"])

    if level >= 9:
        # Level 9: Large sparse data (cells scattered)
        ws.cell(row=100, column=1, value="Sparse 100")
        ws.cell(row=500, column=50, value="Sparse 500,50")
        ws.cell(row=1000, column=100, value="Sparse 1000,100")
        ref["features"].append("sparse_large")
        ref["expected"]["sparse_cells"] = 3

    if level >= 10:
        # Level 10: Special sheet names, unicode, borders
        ws4 = wb.create_sheet("Data'With'Apostrophes")
        ws4["A1"] = "Test"
        ws5 = wb.create_sheet("Unicode 中文 ñ")
        ws5["A1"] = "Unicode: 日本語 €"
        ref["features"].extend(["special_sheet_names", "unicode"])

        thin = Side(style="thin")
        for r in range(1, 5):
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ref["features"].append("borders")

    if level >= 11:
        # Level 11: Complex merges, color scale conditional
        ws.merge_cells("A20:D21")
        ws["A20"] = "Big Merge"
        ws.merge_cells("E20:F20")
        ws["E20"] = "Small"
        ws.conditional_formatting.add(
            "B20:D21",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )
        ref["features"].extend(["complex_merges", "color_scale"])

    if level >= 12:
        # Level 12: Pie chart, date formatting
        ws["A25"] = date(2024, 3, 15)
        ws["A25"].number_format = "YYYY-MM-DD"
        ws["B25"] = "=A25+30"
        ws["B25"].number_format = "YYYY-MM-DD"
        pie = PieChart()
        pie.title = "Pie"
        pie.add_data(Reference(ws, min_col=2, min_row=15, max_row=18), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=16, max_row=18))
        ws.add_chart(pie, "H4")
        ref["features"].extend(["dates", "date_formulas", "chart_pie"])

    if level >= 13:
        # Level 13: Circular ref (parser should detect)
        ws["A30"] = "=B30+1"
        ws["B30"] = "=A30*2"
        ref["features"].append("circular_ref")
        ref["expected"]["circular_detected"] = True

    if level >= 14:
        # Level 14: Hidden sheet
        ws_hidden = wb.create_sheet("HiddenSheet")
        ws_hidden.sheet_state = "hidden"
        ws_hidden["A1"] = "Secret"
        ref["features"].append("hidden_sheet")

    if level >= 15:
        # Level 15: Empty sheet (edge case)
        wb.create_sheet("Empty")
        ref["features"].append("empty_sheet")

    if level >= 16:
        # Level 16: Sheet with only merged cell (value in A1)
        ws_only_merge = wb.create_sheet("OnlyMerge")
        ws_only_merge.merge_cells("A1:C2")
        ws_only_merge["A1"] = "Solo merge"
        ref["features"].append("sheet_only_merged")

    if level >= 17:
        # Level 17: External reference in defined name (if supported)
        wb.defined_names.add(DefinedName("ExternalRef", attr_text="[Other.xlsx]Sheet1!$A$1"))
        ref["features"].append("external_ref_in_name")

    if level >= 18:
        # Level 18: Very long formula
        long_formula = "=IF(A1>0," + "+".join(f"A{i}" for i in range(1, 51)) + ",0)"
        ws["A35"] = long_formula
        ref["features"].append("long_formula")

    # --- Merge stress levels (openpyxl is weak on merge cells / merge columns) ---
    if level >= 19:
        # Level 19: Merge-across pattern — many horizontal 1-row merges (stress merge columns)
        _add_merge_across_sheet(wb, ref)

    if level >= 20:
        # Level 20: Empty master — value in B1, merge A1:B1 (master A1 empty; openpyxl quirk)
        ws_empty = wb.create_sheet("EmptyMaster")
        ws_empty["B1"] = "Value was here"
        ws_empty.merge_cells("A1:B1")  # A1 is top-left master, empty; B1 becomes MergedCell
        ref["features"].append("merge_empty_master")

    if level >= 21:
        # Level 21: Table with merged header — merge cells in range before adding table
        ws_tab = wb.create_sheet("TableMergeHdr")
        ws_tab["A1"] = "Cat"
        ws_tab["B1"] = "A"
        ws_tab["C1"] = "B"
        ws_tab.merge_cells("A1:B1")  # Merged header over first two columns
        for r in range(2, 6):
            ws_tab.cell(row=r, column=1, value=f"C{r-1}")
            ws_tab.cell(row=r, column=2, value=(r - 1) * 10)
            ws_tab.cell(row=r, column=3, value=(r - 1) * 20)
        tab = Table(displayName="TableMergeHdr", ref="A1:C5")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws_tab.add_table(tab)
        ref["features"].append("table_merged_header")

    if level >= 22:
        # Level 22: Dense merge grid — many adjacent 2x2 merges (100+ merge regions)
        ws_dense = wb.create_sheet("DenseMerges")
        for block_r in range(0, 10):  # 10 rows of blocks
            for block_c in range(0, 10):  # 10 cols of blocks
                r1, c1 = block_r * 2 + 1, block_c * 2 + 1
                r2, c2 = r1 + 1, c1 + 1
                rng = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                ws_dense.merge_cells(rng)
                ws_dense.cell(row=r1, column=c1, value=f"{block_r},{block_c}")
        ref["features"].append("merge_dense_grid")
        ref["expected"]["merged_regions_dense"] = 100

    if level >= 23:
        # Level 23: Long vertical merge — A1:A200 (openpyxl + segmenter stress)
        ws_vert = wb.create_sheet("VerticalMerge")
        ws_vert.merge_cells("A1:A200")
        ws_vert["A1"] = "One cell, 200 rows"
        ref["features"].append("merge_vertical_long")

    if level >= 24:
        # Level 24: Alternating merge pattern — complex layout, shared borders
        ws_alt = wb.create_sheet("Alternating")
        for i in range(0, 10):
            r = i * 3 + 1
            ws_alt.merge_cells(f"A{r}:C{r}")      # row 1, 4, 7...
            ws_alt.merge_cells(f"A{r+1}:B{r+1}")   # row 2, 5, 8...
            ws_alt.cell(row=r, column=1, value=f"Row{r}")
            ws_alt.cell(row=r+1, column=1, value=f"Sub{r}")
        ref["features"].append("merge_alternating_pattern")

    if level >= 25:
        # Level 25: Single-cell merge A1:A1 — invalid edge case, openpyxl allows it
        ws_single = wb.create_sheet("SingleMerge")
        ws_single.merge_cells("A1:A1")
        ws_single["A1"] = "Solo"
        ref["features"].append("merge_single_cell")

    wb.save(out_path)
    ref["path"] = str(out_path)
    ref["expected"]["sheets"] = len(wb.sheetnames)
    return ref


def build_all_levels(max_level: int = 25, out_dir: Path | None = None) -> list[dict]:
    """Build workbooks for levels 0 through max_level. Returns list of reference dicts."""
    out_dir = out_dir or STRESS_DIR
    refs = []
    for level in range(max_level + 1):
        path = out_dir / f"stress_level_{level}.xlsx"
        ref = build_level(level, path)
        refs.append(ref)
    spec_path = out_dir / "built_reference.json"
    with open(spec_path, "w") as f:
        json.dump({"levels": refs}, f, indent=2)
    return refs


def build_merge_stress_workbooks(out_dir: Path | None = None) -> list[Path]:
    """
    Build dedicated merge-stress workbooks for parser robustness testing.
    Targets openpyxl's known weaknesses: merge cells, merge columns.
    Returns list of output paths.
    """
    out_dir = out_dir or STRESS_DIR
    paths = []

    # 1. Merge across — 80 rows of horizontal merges
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "MergeAcross"
    for r in range(1, 81):
        ws1.merge_cells(f"A{r}:D{r}")
        ws1.cell(row=r, column=1, value=f"Merged row {r}")
    p1 = out_dir / "merge_stress_across.xlsx"
    wb1.save(p1)
    paths.append(p1)

    # 2. Empty master — value only in right cell, then merge left
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "EmptyMaster"
    ws2["B1"] = "Content"
    ws2.merge_cells("A1:B1")
    p2 = out_dir / "merge_stress_empty_master.xlsx"
    wb2.save(p2)
    paths.append(p2)

    # 3. Table with merged header
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "TableMergeHdr"
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Merged Header"
    ws3["C1"] = "Single"
    for r in range(2, 12):
        ws3.cell(row=r, column=1, value=f"A{r}")
        ws3.cell(row=r, column=2, value=r * 2)
        ws3.cell(row=r, column=3, value=r * 3)
    tab = Table(displayName="MergeTable", ref="A1:C11")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws3.add_table(tab)
    p3 = out_dir / "merge_stress_table_header.xlsx"
    wb3.save(p3)
    paths.append(p3)

    # 4. Dense grid — 15x15 blocks of 2x2 merges = 225 merge regions
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.title = "DenseGrid"
    for br in range(15):
        for bc in range(15):
            r1, c1 = br * 2 + 1, bc * 2 + 1
            r2, c2 = r1 + 1, c1 + 1
            ws4.merge_cells(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")
            ws4.cell(row=r1, column=c1, value=f"{br},{bc}")
    p4 = out_dir / "merge_stress_dense_grid.xlsx"
    wb4.save(p4)
    paths.append(p4)

    # 5. Long vertical merge
    wb5 = Workbook()
    ws5 = wb5.active
    ws5.title = "Vertical"
    ws5.merge_cells("A1:A300")
    ws5["A1"] = "Tall"
    p5 = out_dir / "merge_stress_vertical.xlsx"
    wb5.save(p5)
    paths.append(p5)

    # 6. Single-cell merge (invalid edge case)
    wb6 = Workbook()
    ws6 = wb6.active
    ws6.title = "SingleMerge"
    ws6.merge_cells("A1:A1")
    ws6["A1"] = "One"
    p6 = out_dir / "merge_stress_single_cell.xlsx"
    wb6.save(p6)
    paths.append(p6)

    return paths


if __name__ == "__main__":
    import sys
    if "--merge-stress" in sys.argv:
        sys.argv.remove("--merge-stress")
        paths = build_merge_stress_workbooks()
        print(f"Built {len(paths)} merge-stress workbooks in {STRESS_DIR}:")
        for p in paths:
            print(f"  {p.name}")
    else:
        max_lvl = int(sys.argv[1]) if len(sys.argv) > 1 else 25
        refs = build_all_levels(max_level=max_lvl)
        print(f"Built {len(refs)} workbooks (levels 0..{max_lvl}) in {STRESS_DIR}")
        for r in refs:
            print(f"  Level {r['level']}: {r['path']} - {', '.join(r['features'])}")
