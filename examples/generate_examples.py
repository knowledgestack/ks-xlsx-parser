#!/usr/bin/env python3
"""
Generate example Excel workbooks for demonstrating the xlsx_parser.

Creates several representative workbooks in the examples/ folder
that showcase the parser's capabilities across different Excel features.

Run: python examples/generate_examples.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

EXAMPLES_DIR = Path(__file__).parent.parent / "testBench" / "real_world"
EXAMPLES_DIR.mkdir(parents=True, exist_ok=True)


def create_financial_model():
    """
    Financial model workbook with assumptions, calculations, and results.

    Demonstrates:
    - Multiple sheets with cross-sheet formulas
    - Named ranges
    - Number formatting (currency, percentage)
    - Bold headers and styled output cells
    - Assumptions + Results block segmentation
    """
    wb = Workbook()

    # --- Assumptions Sheet ---
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_assumptions.freeze_panes = "A2"

    headers = ["Parameter", "Value", "Unit", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_assumptions.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    assumptions = [
        ("Revenue Growth Rate", 0.08, "%", "Year-over-year"),
        ("COGS Margin", 0.45, "%", "% of revenue"),
        ("OpEx Growth Rate", 0.05, "%", "Annual increase"),
        ("Tax Rate", 0.21, "%", "Federal + State"),
        ("Discount Rate (WACC)", 0.10, "%", "Risk-adjusted"),
        ("Initial Revenue", 10000000, "$", "Year 0 baseline"),
        ("Initial OpEx", 3000000, "$", "Year 0 baseline"),
        ("CapEx (Annual)", 500000, "$", "Flat assumption"),
        ("Working Capital % Rev", 0.12, "%", "NWC/Revenue"),
        ("Terminal Growth Rate", 0.025, "%", "Long-term GDP"),
    ]
    for i, (param, value, unit, notes) in enumerate(assumptions, 2):
        ws_assumptions.cell(row=i, column=1, value=param)
        cell = ws_assumptions.cell(row=i, column=2, value=value)
        if unit == "%":
            cell.number_format = "0.0%"
        elif unit == "$":
            cell.number_format = "#,##0"
        ws_assumptions.cell(row=i, column=3, value=unit)
        ws_assumptions.cell(row=i, column=4, value=notes)
        ws_assumptions.cell(row=i, column=4).font = Font(italic=True, color="808080")

    ws_assumptions.column_dimensions["A"].width = 25
    ws_assumptions.column_dimensions["B"].width = 15
    ws_assumptions.column_dimensions["D"].width = 20

    # Named ranges
    wb.defined_names.add(DefinedName("GrowthRate", attr_text="Assumptions!$B$2"))
    wb.defined_names.add(DefinedName("COGSMargin", attr_text="Assumptions!$B$3"))
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Assumptions!$B$5"))
    wb.defined_names.add(DefinedName("WACC", attr_text="Assumptions!$B$6"))
    wb.defined_names.add(DefinedName("InitialRevenue", attr_text="Assumptions!$B$7"))

    # --- P&L Sheet ---
    ws_pnl = wb.create_sheet("P&L")
    ws_pnl.freeze_panes = "B2"

    # Year headers
    ws_pnl["A1"] = "Income Statement"
    ws_pnl["A1"].font = Font(bold=True, size=14)
    ws_pnl.merge_cells("A1:F1")

    years = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for col, yr in enumerate(years, 1):
        cell = ws_pnl.cell(row=2, column=col, value=yr)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if col > 1:
            cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Revenue
    ws_pnl["A3"] = "Revenue"
    ws_pnl["A3"].font = Font(bold=True)
    ws_pnl["B3"] = "=Assumptions!B7*(1+Assumptions!B2)"
    ws_pnl["B3"].number_format = "#,##0"
    for col in range(3, 7):
        ref = get_column_letter(col - 1)
        ws_pnl.cell(row=3, column=col, value=f"={ref}3*(1+Assumptions!$B$2)")
        ws_pnl.cell(row=3, column=col).number_format = "#,##0"

    # COGS
    ws_pnl["A4"] = "COGS"
    for col in range(2, 7):
        ref = get_column_letter(col)
        ws_pnl.cell(row=4, column=col, value=f"=-{ref}3*Assumptions!$B$3")
        ws_pnl.cell(row=4, column=col).number_format = "#,##0"

    # Gross Profit
    ws_pnl["A5"] = "Gross Profit"
    ws_pnl["A5"].font = Font(bold=True)
    for col in range(2, 7):
        ref = get_column_letter(col)
        ws_pnl.cell(row=5, column=col, value=f"={ref}3+{ref}4")
        ws_pnl.cell(row=5, column=col).number_format = "#,##0"
        ws_pnl.cell(row=5, column=col).font = Font(bold=True)

    # OpEx
    ws_pnl["A6"] = "Operating Expenses"
    ws_pnl["B6"] = "=-Assumptions!B8*(1+Assumptions!B4)"
    ws_pnl["B6"].number_format = "#,##0"
    for col in range(3, 7):
        ref = get_column_letter(col - 1)
        ws_pnl.cell(row=6, column=col, value=f"={ref}6*(1+Assumptions!$B$4)")
        ws_pnl.cell(row=6, column=col).number_format = "#,##0"

    # EBIT
    ws_pnl["A7"] = "EBIT"
    ws_pnl["A7"].font = Font(bold=True)
    for col in range(2, 7):
        ref = get_column_letter(col)
        ws_pnl.cell(row=7, column=col, value=f"={ref}5+{ref}6")
        ws_pnl.cell(row=7, column=col).number_format = "#,##0"
        ws_pnl.cell(row=7, column=col).font = Font(bold=True)

    # Tax
    ws_pnl["A8"] = "Tax"
    for col in range(2, 7):
        ref = get_column_letter(col)
        ws_pnl.cell(row=8, column=col, value=f"=-MAX({ref}7*Assumptions!$B$5,0)")
        ws_pnl.cell(row=8, column=col).number_format = "#,##0"

    # Net Income
    ws_pnl["A9"] = "Net Income"
    ws_pnl["A9"].font = Font(bold=True, color="006100")
    for col in range(2, 7):
        ref = get_column_letter(col)
        cell = ws_pnl.cell(row=9, column=col, value=f"={ref}7+{ref}8")
        cell.number_format = "#,##0"
        cell.font = Font(bold=True, color="006100")
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws_pnl.column_dimensions["A"].width = 22

    # Add borders
    thin = Side(style="thin")
    for row in range(2, 10):
        for col in range(1, 7):
            ws_pnl.cell(row=row, column=col).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )

    # --- Summary Sheet ---
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "5-Year Financial Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)

    ws_summary["A3"] = "Total Revenue (5yr)"
    ws_summary["B3"] = "=SUM('P&L'!B3:F3)"
    ws_summary["B3"].number_format = "#,##0"

    ws_summary["A4"] = "Total Net Income (5yr)"
    ws_summary["B4"] = "=SUM('P&L'!B9:F9)"
    ws_summary["B4"].number_format = "#,##0"

    ws_summary["A5"] = "Average Net Margin"
    ws_summary["B5"] = "=B4/B3"
    ws_summary["B5"].number_format = "0.0%"

    # Add a line chart for revenue
    chart = LineChart()
    chart.title = "Revenue Forecast"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Year"
    chart.width = 15
    chart.height = 10
    data = Reference(ws_pnl, min_col=2, max_col=6, min_row=3, max_row=3)
    cats = Reference(ws_pnl, min_col=2, max_col=6, min_row=2, max_row=2)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.width = 25000
    ws_summary.add_chart(chart, "A8")

    wb.save(EXAMPLES_DIR / "financial_model.xlsx")
    print("Created: financial_model.xlsx")


def create_sales_dashboard():
    """
    Sales dashboard workbook with tables, charts, and conditional formatting.

    Demonstrates:
    - Excel ListObject tables
    - Bar and pie charts
    - Conditional formatting (color scales, cell rules)
    - Data validation dropdowns
    - Multi-level headers with merges
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Sales Data"

    # Multi-level header with merges
    ws.merge_cells("A1:G1")
    ws["A1"] = "Regional Sales Dashboard — FY2024"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "All figures in thousands ($)"
    ws["A2"].font = Font(italic=True, color="808080", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Table headers
    headers = ["Product", "Region", "Q1", "Q2", "Q3", "Q4", "FY Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Sales data
    data = [
        ("Widget Pro", "North America", 245, 312, 287, 356, None),
        ("Widget Pro", "Europe", 189, 234, 267, 298, None),
        ("Widget Pro", "Asia Pacific", 156, 198, 223, 245, None),
        ("Widget Lite", "North America", 89, 112, 134, 156, None),
        ("Widget Lite", "Europe", 67, 89, 98, 123, None),
        ("Widget Lite", "Asia Pacific", 45, 67, 78, 89, None),
        ("Enterprise", "North America", 890, 1023, 1156, 1234, None),
        ("Enterprise", "Europe", 567, 678, 756, 823, None),
        ("Enterprise", "Asia Pacific", 345, 456, 534, 612, None),
    ]

    for i, row_data in enumerate(data, 4):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if col >= 3 and col <= 6:
                cell.number_format = "#,##0"
        # FY Total formula
        ws.cell(row=i, column=7, value=f"=SUM(C{i}:F{i})")
        ws.cell(row=i, column=7).number_format = "#,##0"
        ws.cell(row=i, column=7).font = Font(bold=True)

    # Create table
    tab = Table(displayName="SalesData", ref="A3:G12")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=True,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Conditional formatting: color scale on Q1-Q4
    ws.conditional_formatting.add(
        "C4:F12",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )

    # Data validation on Region
    dv = DataValidation(
        type="list",
        formula1='"North America,Europe,Asia Pacific,Latin America"',
        allow_blank=False,
    )
    dv.prompt = "Select a region"
    dv.promptTitle = "Region"
    ws.add_data_validation(dv)
    for row in range(4, 13):
        dv.add(ws.cell(row=row, column=2))

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18

    # --- Charts Sheet ---
    ws_charts = wb.create_sheet("Charts")

    # Bar chart
    bar = BarChart()
    bar.type = "col"
    bar.title = "Sales by Region"
    bar.y_axis.title = "Revenue ($K)"
    data_ref = Reference(ws, min_col=7, min_row=3, max_row=12)
    cats_ref = Reference(ws, min_col=2, min_row=4, max_row=12)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.width = 18
    bar.height = 12
    ws_charts.add_chart(bar, "A1")

    # Pie chart
    pie = PieChart()
    pie.title = "Revenue by Product"
    # Create summary data for pie
    ws_charts["M1"] = "Product"
    ws_charts["N1"] = "Total"
    ws_charts["M2"] = "Widget Pro"
    ws_charts["N2"] = f"=SUM('Sales Data'!G4:G6)"
    ws_charts["M3"] = "Widget Lite"
    ws_charts["N3"] = f"=SUM('Sales Data'!G7:G9)"
    ws_charts["M4"] = "Enterprise"
    ws_charts["N4"] = f"=SUM('Sales Data'!G10:G12)"

    pie_data = Reference(ws_charts, min_col=14, min_row=1, max_row=4)
    pie_cats = Reference(ws_charts, min_col=13, min_row=2, max_row=4)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 12
    pie.height = 10
    ws_charts.add_chart(pie, "A18")

    wb.save(EXAMPLES_DIR / "sales_dashboard.xlsx")
    print("Created: sales_dashboard.xlsx")


def create_project_tracker():
    """
    Project tracker workbook with tasks, statuses, and timeline.

    Demonstrates:
    - Data validation dropdowns (status, priority)
    - Conditional formatting (status-based highlighting)
    - Comments on cells
    - Hyperlinks
    - Date formatting
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.freeze_panes = "A2"

    headers = ["Task ID", "Task Name", "Owner", "Status", "Priority", "Start Date", "Due Date", "% Complete", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    from datetime import date, timedelta

    tasks = [
        ("TSK-001", "Requirements gathering", "Alice", "Complete", "High", date(2024, 1, 15), date(2024, 2, 1), 1.0, ""),
        ("TSK-002", "System design", "Bob", "Complete", "High", date(2024, 2, 1), date(2024, 3, 1), 1.0, ""),
        ("TSK-003", "Backend development", "Charlie", "In Progress", "High", date(2024, 3, 1), date(2024, 5, 1), 0.65, "On track"),
        ("TSK-004", "Frontend development", "Diana", "In Progress", "Medium", date(2024, 3, 15), date(2024, 5, 15), 0.40, ""),
        ("TSK-005", "API integration", "Eve", "Not Started", "Medium", date(2024, 4, 1), date(2024, 5, 1), 0.0, "Blocked by TSK-003"),
        ("TSK-006", "Testing", "Frank", "Not Started", "High", date(2024, 5, 1), date(2024, 6, 1), 0.0, ""),
        ("TSK-007", "Documentation", "Grace", "Not Started", "Low", date(2024, 5, 15), date(2024, 6, 15), 0.0, ""),
        ("TSK-008", "Deployment", "Alice", "Not Started", "Critical", date(2024, 6, 1), date(2024, 6, 15), 0.0, ""),
    ]

    for i, (tid, name, owner, status, priority, start, due, pct, notes) in enumerate(tasks, 2):
        ws.cell(row=i, column=1, value=tid)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=owner)
        ws.cell(row=i, column=4, value=status)
        ws.cell(row=i, column=5, value=priority)
        ws.cell(row=i, column=6, value=start).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=7, value=due).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=8, value=pct).number_format = "0%"
        ws.cell(row=i, column=9, value=notes)

    # Status validation
    status_dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,On Hold,Cancelled"',
    )
    ws.add_data_validation(status_dv)
    for r in range(2, 10):
        status_dv.add(ws.cell(row=r, column=4))

    # Priority validation
    priority_dv = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
    )
    ws.add_data_validation(priority_dv)
    for r in range(2, 10):
        priority_dv.add(ws.cell(row=r, column=5))

    # Conditional formatting for status
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add("D2:D10", CellIsRule(operator="equal", formula=['"Complete"'], fill=green))
    ws.conditional_formatting.add("D2:D10", CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow))
    ws.conditional_formatting.add("D2:D10", CellIsRule(operator="equal", formula=['"Not Started"'], fill=red))

    # Comments
    ws["E5"].comment = Comment("Blocked until backend API is ready", "PM")
    ws["H4"].comment = Comment("Updated 2024-03-20", "Diana")

    # Column widths
    widths = [10, 25, 12, 15, 12, 12, 12, 12, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(EXAMPLES_DIR / "project_tracker.xlsx")
    print("Created: project_tracker.xlsx")


def create_engineering_calcs():
    """
    Engineering calculations workbook with units and formulas.

    Demonstrates:
    - Dense calculation blocks with unit annotations
    - Multiple formula dependency chains
    - Hidden helper rows
    - Print area settings
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Beam Design"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Steel Beam Design Calculator"
    ws["A1"].font = Font(bold=True, size=14)

    # Input section
    ws["A3"] = "INPUTS"
    ws["A3"].font = Font(bold=True, size=12, color="2F5496")

    inputs = [
        ("Span Length", "L", 6.0, "m"),
        ("Distributed Load", "w", 25.0, "kN/m"),
        ("Point Load (mid-span)", "P", 50.0, "kN"),
        ("Steel Yield Strength", "fy", 250.0, "MPa"),
        ("Elastic Modulus", "E", 200000.0, "MPa"),
        ("Safety Factor", "SF", 1.5, "—"),
    ]

    for i, (desc, symbol, value, unit) in enumerate(inputs, 4):
        ws.cell(row=i, column=1, value=desc)
        ws.cell(row=i, column=2, value=symbol).font = Font(italic=True)
        ws.cell(row=i, column=3, value=value)
        ws.cell(row=i, column=4, value=unit).font = Font(color="808080")

    # Calculation section
    ws["A11"] = "CALCULATIONS"
    ws["A11"].font = Font(bold=True, size=12, color="2F5496")

    calcs = [
        ("Max Moment (distributed)", "Mw", "=C5*C4^2/8", "kN·m"),
        ("Max Moment (point load)", "Mp", "=C6*C4/4", "kN·m"),
        ("Total Max Moment", "Mtotal", "=C12+C13", "kN·m"),
        ("Design Moment", "Md", "=C14*C9", "kN·m"),
        ("Required Section Modulus", "Sx_req", "=C15*1000000/C7", "mm³"),
        ("Max Deflection (distributed)", "δw", "=5*C5*C4^4/(384*C8*1000)", "mm (approx)"),
        ("Max Deflection (point load)", "δp", "=C6*C4^3/(48*C8*1000)", "mm (approx)"),
    ]

    for i, (desc, symbol, formula, unit) in enumerate(calcs, 12):
        ws.cell(row=i, column=1, value=desc)
        ws.cell(row=i, column=2, value=symbol).font = Font(italic=True)
        ws.cell(row=i, column=3, value=formula)
        ws.cell(row=i, column=3).number_format = "#,##0.00"
        ws.cell(row=i, column=4, value=unit).font = Font(color="808080")

    # Results section
    ws["A20"] = "RESULTS"
    ws["A20"].font = Font(bold=True, size=12, color="006100")

    result_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws["A21"] = "Design Moment"
    ws["B21"] = "Md"
    ws["C21"] = "=C15"
    ws["C21"].number_format = "#,##0.00"
    ws["C21"].font = Font(bold=True)
    ws["C21"].fill = result_fill
    ws["D21"] = "kN·m"

    ws["A22"] = "Required Sx"
    ws["B22"] = "Sx"
    ws["C22"] = "=C16"
    ws["C22"].number_format = "#,##0"
    ws["C22"].font = Font(bold=True)
    ws["C22"].fill = result_fill
    ws["D22"] = "mm³"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 15

    # Named ranges for key outputs
    wb.defined_names.add(DefinedName("DesignMoment", attr_text="'Beam Design'!$C$15"))
    wb.defined_names.add(DefinedName("RequiredSx", attr_text="'Beam Design'!$C$16"))

    ws.print_area = "A1:D22"

    wb.save(EXAMPLES_DIR / "engineering_calcs.xlsx")
    print("Created: engineering_calcs.xlsx")


def create_data_inventory():
    """
    Data inventory with wide columns and sparse data.

    Demonstrates:
    - Wide sheet (many columns)
    - Sparse data patterns
    - Hidden columns
    - Multiple data types
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Headers
    base_headers = ["SKU", "Product Name", "Category", "Supplier", "Unit Cost", "Qty on Hand", "Reorder Level", "Status"]
    # Monthly sales columns
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    all_headers = base_headers + [f"Sales {m}" for m in months] + ["YTD Sales", "YTD Revenue"]

    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        if col <= len(base_headers):
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Sample data
    import random
    random.seed(42)

    products = [
        ("SKU-001", "Bolt M8x30", "Fasteners", "SupplierA", 0.15),
        ("SKU-002", "Nut M8", "Fasteners", "SupplierA", 0.08),
        ("SKU-003", "Washer M8", "Fasteners", "SupplierA", 0.05),
        ("SKU-010", "Pipe 25mm", "Plumbing", "SupplierB", 12.50),
        ("SKU-011", "Elbow 25mm", "Plumbing", "SupplierB", 3.20),
        ("SKU-020", "Cable 2.5mm", "Electrical", "SupplierC", 1.80),
        ("SKU-021", "Switch 10A", "Electrical", "SupplierC", 4.50),
        ("SKU-030", "Timber 2x4", "Lumber", "SupplierD", 8.00),
    ]

    for i, (sku, name, cat, supplier, cost) in enumerate(products, 2):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=cat)
        ws.cell(row=i, column=4, value=supplier)
        ws.cell(row=i, column=5, value=cost).number_format = "#,##0.00"
        ws.cell(row=i, column=6, value=random.randint(50, 5000))
        ws.cell(row=i, column=7, value=random.randint(10, 100))
        ws.cell(row=i, column=8, value="Active")

        # Monthly sales (sparse — some months have no data)
        for m in range(12):
            col = 9 + m
            if random.random() > 0.3:  # 70% chance of having data
                ws.cell(row=i, column=col, value=random.randint(0, 500))

        # YTD formulas
        ws.cell(row=i, column=21, value=f"=SUM(I{i}:T{i})")
        ws.cell(row=i, column=22, value=f"=U{i}*E{i}")
        ws.cell(row=i, column=22).number_format = "#,##0.00"

    # Hide supplier column
    ws.column_dimensions["D"].hidden = True

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18

    # Conditional formatting on status
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.conditional_formatting.add(
        "H2:H9",
        CellIsRule(operator="equal", formula=['"Active"'], fill=green_fill),
    )

    wb.save(EXAMPLES_DIR / "data_inventory.xlsx")
    print("Created: data_inventory.xlsx")


if __name__ == "__main__":
    create_financial_model()
    create_sales_dashboard()
    create_project_tracker()
    create_engineering_calcs()
    create_data_inventory()
    print("\nAll example workbooks generated in examples/")
