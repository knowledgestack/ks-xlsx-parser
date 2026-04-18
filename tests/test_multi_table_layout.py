"""
Tests for multi-table layout detection and classification.

Verifies that the segmenter correctly identifies multiple tables within
a single sheet, detects their boundaries, headers, colors, and block types
across various spatial arrangements.
"""

import pytest

from chunking.segmenter import LayoutSegmenter
from models import BlockType
from parsers import WorkbookParser
from pipeline import parse_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_and_segment(path):
    """Parse a workbook and segment its first sheet into blocks."""
    result = WorkbookParser(path=path).parse()
    sheet = result.sheets[0]
    tables = [t for t in result.tables if t.sheet_name == sheet.sheet_name]
    segmenter = LayoutSegmenter(sheet, tables=tables)
    blocks = segmenter.segment()
    return result, sheet, blocks


# ---------------------------------------------------------------------------
# Test classes
# ---------------------------------------------------------------------------


class TestVerticalLayout:
    """Two tables stacked vertically separated by blank rows."""

    def test_detects_two_blocks(self, two_tables_vertical):
        _, _, blocks = _parse_and_segment(two_tables_vertical)
        assert len(blocks) == 2

    def test_block_types_are_table(self, two_tables_vertical):
        _, _, blocks = _parse_and_segment(two_tables_vertical)
        assert blocks[0].block_type == BlockType.TABLE
        assert blocks[1].block_type == BlockType.TABLE

    def test_first_table_range(self, two_tables_vertical):
        _, _, blocks = _parse_and_segment(two_tables_vertical)
        assert blocks[0].cell_range.to_a1() == "A1:C4"

    def test_second_table_range(self, two_tables_vertical):
        _, _, blocks = _parse_and_segment(two_tables_vertical)
        assert blocks[1].cell_range.to_a1() == "A7:B10"

    def test_first_table_has_bold_header(self, two_tables_vertical):
        _, sheet, _ = _parse_and_segment(two_tables_vertical)
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.style is not None
        assert a1.style.font.bold is True

    def test_blocks_sorted_by_position(self, two_tables_vertical):
        _, _, blocks = _parse_and_segment(two_tables_vertical)
        assert blocks[0].cell_range.top_left.row < blocks[1].cell_range.top_left.row

    def test_cell_counts(self, two_tables_vertical):
        _, _, blocks = _parse_and_segment(two_tables_vertical)
        assert blocks[0].cell_count == 12  # 4 rows x 3 cols
        assert blocks[1].cell_count == 8   # 4 rows x 2 cols


class TestHorizontalLayout:
    """Two tables side by side separated by an empty column."""

    def test_detects_two_blocks(self, two_tables_horizontal):
        _, _, blocks = _parse_and_segment(two_tables_horizontal)
        assert len(blocks) == 2

    def test_left_table_range(self, two_tables_horizontal):
        _, _, blocks = _parse_and_segment(two_tables_horizontal)
        left_block = [b for b in blocks if b.cell_range.top_left.col == 1][0]
        assert left_block.cell_range.to_a1() == "A1:C4"

    def test_right_table_range(self, two_tables_horizontal):
        _, _, blocks = _parse_and_segment(two_tables_horizontal)
        right_block = [b for b in blocks if b.cell_range.top_left.col == 5][0]
        assert right_block.cell_range.to_a1() == "E1:F4"

    def test_both_classified_as_table(self, two_tables_horizontal):
        _, _, blocks = _parse_and_segment(two_tables_horizontal)
        for b in blocks:
            assert b.block_type == BlockType.TABLE

    def test_blocks_sorted_left_to_right(self, two_tables_horizontal):
        _, _, blocks = _parse_and_segment(two_tables_horizontal)
        assert blocks[0].cell_range.top_left.col < blocks[1].cell_range.top_left.col


class TestTableChartTable:
    """Table, then chart anchor area, then another table."""

    def test_detects_two_data_blocks(self, table_chart_table):
        _, _, blocks = _parse_and_segment(table_chart_table)
        table_blocks = [b for b in blocks if b.block_type == BlockType.TABLE]
        assert len(table_blocks) == 2

    def test_chart_extracted_separately(self, table_chart_table):
        result, _, _ = _parse_and_segment(table_chart_table)
        assert len(result.charts) >= 1

    def test_first_table_ends_before_gap(self, table_chart_table):
        _, _, blocks = _parse_and_segment(table_chart_table)
        table_blocks = sorted(
            [b for b in blocks if b.block_type == BlockType.TABLE],
            key=lambda b: b.cell_range.top_left.row,
        )
        assert table_blocks[0].cell_range.bottom_right.row <= 5

    def test_second_table_starts_at_row_10(self, table_chart_table):
        _, _, blocks = _parse_and_segment(table_chart_table)
        table_blocks = sorted(
            [b for b in blocks if b.block_type == BlockType.TABLE],
            key=lambda b: b.cell_range.top_left.row,
        )
        assert table_blocks[1].cell_range.top_left.row == 10

    def test_pipeline_produces_chunks(self, table_chart_table):
        result = parse_workbook(path=table_chart_table)
        # At least 2 table chunks + chart
        assert result.total_chunks >= 2


class TestTableInMiddle:
    """Single table centered in the sheet with empty space around."""

    def test_detects_one_block(self, table_in_middle):
        _, _, blocks = _parse_and_segment(table_in_middle)
        assert len(blocks) == 1

    def test_block_range_matches_data(self, table_in_middle):
        _, _, blocks = _parse_and_segment(table_in_middle)
        assert blocks[0].cell_range.to_a1() == "D5:F9"

    def test_classified_as_table(self, table_in_middle):
        _, _, blocks = _parse_and_segment(table_in_middle)
        assert blocks[0].block_type == BlockType.TABLE

    def test_bounding_box_offset(self, table_in_middle):
        _, _, blocks = _parse_and_segment(table_in_middle)
        assert blocks[0].bounding_box.x > 0
        assert blocks[0].bounding_box.y > 0


class TestMixedContentLayout:
    """Mixed content: header, table, text block, assumptions block."""

    def test_detects_multiple_blocks(self, mixed_content_layout):
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        assert len(blocks) >= 2

    def test_title_merged_into_table(self, mixed_content_layout):
        """Title rows are absorbed into the following data block for RAG context."""
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        # The first block should start at row 1 (title) and include the table
        first = blocks[0]
        assert first.cell_range.top_left.row == 1
        assert first.block_type == BlockType.TABLE

    def test_table_block_detected(self, mixed_content_layout):
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        table_blocks = [b for b in blocks if b.block_type == BlockType.TABLE]
        assert len(table_blocks) >= 1

    def test_text_block_detected(self, mixed_content_layout):
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        text_blocks = [b for b in blocks if b.block_type == BlockType.TEXT_BLOCK]
        assert len(text_blocks) >= 1

    def test_assumptions_block_detected(self, mixed_content_layout):
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        # The assumptions block contains keyword "Assumption" + "Input"
        assumption_blocks = [
            b for b in blocks if b.block_type == BlockType.ASSUMPTIONS_TABLE
        ]
        assert len(assumption_blocks) >= 1

    def test_blocks_ordered_by_position(self, mixed_content_layout):
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        rows = [b.cell_range.top_left.row for b in blocks]
        assert rows == sorted(rows)

    def test_no_blocks_overlap(self, mixed_content_layout):
        _, _, blocks = _parse_and_segment(mixed_content_layout)
        for i, a in enumerate(blocks):
            for b in blocks[i + 1 :]:
                # Two ranges overlap if they share any row AND any column
                a_r = a.cell_range
                b_r = b.cell_range
                rows_overlap = (
                    a_r.top_left.row <= b_r.bottom_right.row
                    and b_r.top_left.row <= a_r.bottom_right.row
                )
                cols_overlap = (
                    a_r.top_left.col <= b_r.bottom_right.col
                    and b_r.top_left.col <= a_r.bottom_right.col
                )
                assert not (rows_overlap and cols_overlap), (
                    f"Blocks overlap: {a_r.to_a1()} and {b_r.to_a1()}"
                )


class TestTitleMerging:
    """Title/header rows should be merged into the following data block."""

    @pytest.fixture
    def title_above_table(self, tmp_dir):
        """Title row followed by a data table with no gap."""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.worksheet.table import Table, TableStyleInfo

        path = tmp_dir / "title_above_table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        # Title row
        ws["A1"] = "Regional Sales Dashboard — FY2024"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")
        # Subtitle
        ws["A2"] = "All figures in thousands ($)"

        # Data table starting at row 3
        for ci, h in enumerate(["Product", "Region", "Q1", "Q2"], 1):
            ws.cell(row=3, column=ci, value=h).font = Font(bold=True)
        for ri, row_data in enumerate(
            [("Widget", "NA", 100, 200), ("Gadget", "EU", 150, 250)], 4
        ):
            for ci, v in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=v)

        tab = Table(displayName="SalesData", ref="A3:D5")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
        ws.add_table(tab)

        wb.save(path)
        return path

    @pytest.fixture
    def title_with_gap(self, tmp_dir):
        """Title row followed by 1 blank row then a data table."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        path = tmp_dir / "title_with_gap.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        ws["A1"] = "QUARTERLY REPORT"
        ws["A1"].font = Font(bold=True, size=14)
        # Row 2 blank
        for ci, h in enumerate(["Category", "Value"], 1):
            ws.cell(row=3, column=ci, value=h).font = Font(bold=True)
        for ri, row_data in enumerate([("Alpha", 10), ("Beta", 20), ("Gamma", 30)], 4):
            for ci, v in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=v)

        wb.save(path)
        return path

    @pytest.fixture
    def distant_title(self, tmp_dir):
        """Title row with 3 blank rows before data — should NOT merge."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        path = tmp_dir / "distant_title.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Distant"

        ws["A1"] = "FAR AWAY TITLE"
        ws["A1"].font = Font(bold=True, size=14)
        # Rows 2-4 blank
        for ci, h in enumerate(["X", "Y"], 1):
            ws.cell(row=5, column=ci, value=h).font = Font(bold=True)
        for ri, row_data in enumerate([(1, 2), (3, 4), (5, 6), (7, 8)], 6):
            for ci, v in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=v)

        wb.save(path)
        return path

    def test_title_merged_into_excel_table(self, title_above_table):
        """Title rows above an Excel table should merge into the table chunk."""
        _, _, blocks = _parse_and_segment(title_above_table)
        # Title (rows 1-2) + table (rows 3-5) → single block starting at row 1
        table_block = next(b for b in blocks if b.table_name == "SalesData")
        assert table_block.cell_range.top_left.row == 1

    def test_title_merged_preserves_table_type(self, title_above_table):
        result = parse_workbook(path=title_above_table)
        chunk = result.chunks[0]
        assert "Regional Sales Dashboard" in chunk.render_text
        assert "Product" in chunk.render_text

    def test_title_with_one_blank_row_merges(self, title_with_gap):
        """HEADER block with 1 blank row gap should merge into data below."""
        _, _, blocks = _parse_and_segment(title_with_gap)
        first = blocks[0]
        assert first.cell_range.top_left.row == 1
        assert first.cell_range.bottom_right.row >= 4

    def test_distant_title_stays_separate(self, distant_title):
        """Title with 3+ blank rows gap should NOT merge."""
        _, _, blocks = _parse_and_segment(distant_title)
        assert len(blocks) >= 2
        assert blocks[0].cell_range.bottom_right.row < 5

    def test_merged_chunk_text_includes_title(self, title_above_table):
        """The merged chunk's render_text should contain both title and data."""
        result = parse_workbook(path=title_above_table)
        texts = [ch.render_text for ch in result.chunks if ch.sheet_name == "Sales"]
        combined = "\n".join(texts)
        assert "Regional Sales Dashboard" in combined
        assert "Widget" in combined

    def test_merged_chunk_lineage_spans_full_range(self, title_above_table):
        """Merged chunk's source_uri should cover row 1 through the table end."""
        result = parse_workbook(path=title_above_table)
        sales_chunks = [ch for ch in result.chunks if ch.sheet_name == "Sales"]
        assert len(sales_chunks) == 1
        assert sales_chunks[0].top_left_cell == "A1"


class TestColorCodedTables:
    """Two tables with distinct color schemes."""

    def test_detects_two_blocks(self, color_coded_tables):
        _, _, blocks = _parse_and_segment(color_coded_tables)
        assert len(blocks) == 2

    def test_first_table_header_fill_blue(self, color_coded_tables):
        _, sheet, _ = _parse_and_segment(color_coded_tables)
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.style is not None
        assert a1.style.fill is not None
        assert a1.style.fill.fg_color is not None
        assert "4472C4" in a1.style.fill.fg_color.upper()

    def test_first_table_header_font_white(self, color_coded_tables):
        _, sheet, _ = _parse_and_segment(color_coded_tables)
        a1 = sheet.get_cell(1, 1)
        assert a1.style.font.color is not None
        assert "FFFFFF" in a1.style.font.color.upper()

    def test_first_table_data_fill(self, color_coded_tables):
        _, sheet, _ = _parse_and_segment(color_coded_tables)
        a2 = sheet.get_cell(2, 1)
        assert a2 is not None
        assert a2.style is not None
        assert a2.style.fill is not None
        assert a2.style.fill.fg_color is not None
        assert "D9E2F3" in a2.style.fill.fg_color.upper()

    def test_second_table_header_fill_green(self, color_coded_tables):
        _, sheet, _ = _parse_and_segment(color_coded_tables)
        a8 = sheet.get_cell(8, 1)
        assert a8 is not None
        assert a8.style is not None
        assert a8.style.fill is not None
        assert a8.style.fill.fg_color is not None
        assert "70AD47" in a8.style.fill.fg_color.upper()

    def test_second_table_data_fill(self, color_coded_tables):
        _, sheet, _ = _parse_and_segment(color_coded_tables)
        a9 = sheet.get_cell(9, 1)
        assert a9 is not None
        assert a9.style is not None
        assert a9.style.fill is not None
        assert a9.style.fill.fg_color is not None
        assert "E2EFDA" in a9.style.fill.fg_color.upper()

    def test_key_cells_include_colored_headers(self, color_coded_tables):
        _, _, blocks = _parse_and_segment(color_coded_tables)
        for block in blocks:
            assert len(block.key_cells) > 0

    def test_both_have_formatting(self, color_coded_tables):
        _, _, blocks = _parse_and_segment(color_coded_tables)
        for block in blocks:
            assert block.has_formatting is True


class TestComplexHeaders:
    """Tables with merged header rows and multi-style headers."""

    def test_detects_two_blocks(self, complex_headers_layout):
        _, _, blocks = _parse_and_segment(complex_headers_layout)
        assert len(blocks) == 2

    def test_first_block_has_merges(self, complex_headers_layout):
        _, _, blocks = _parse_and_segment(complex_headers_layout)
        assert blocks[0].has_merges is True

    def test_second_block_has_merges(self, complex_headers_layout):
        _, _, blocks = _parse_and_segment(complex_headers_layout)
        assert blocks[1].has_merges is True

    def test_first_block_range(self, complex_headers_layout):
        _, _, blocks = _parse_and_segment(complex_headers_layout)
        assert blocks[0].cell_range.top_left.row == 1
        assert blocks[0].cell_range.bottom_right.row == 6

    def test_second_block_range(self, complex_headers_layout):
        _, _, blocks = _parse_and_segment(complex_headers_layout)
        assert blocks[1].cell_range.top_left.row == 8
        assert blocks[1].cell_range.bottom_right.row == 12

    def test_merged_header_cell_properties(self, complex_headers_layout):
        _, sheet, _ = _parse_and_segment(complex_headers_layout)
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.is_merged_master is True
        assert a1.style.font.bold is True
        assert a1.style.fill is not None

    def test_column_headers_have_different_fills(self, complex_headers_layout):
        _, sheet, _ = _parse_and_segment(complex_headers_layout)
        fills = set()
        for col in range(1, 4):
            cell = sheet.get_cell(2, col)
            assert cell is not None
            assert cell.style is not None
            assert cell.style.fill is not None
            fills.add(cell.style.fill.fg_color)
        assert len(fills) == 3  # Three distinct colors

    def test_first_block_has_formulas(self, complex_headers_layout):
        _, _, blocks = _parse_and_segment(complex_headers_layout)
        assert blocks[0].formula_count > 0


class TestMultiTablePipeline:
    """End-to-end pipeline tests for multi-table layouts."""

    def test_vertical_produces_correct_chunk_count(self, two_tables_vertical):
        result = parse_workbook(path=two_tables_vertical)
        assert result.total_chunks == 2

    def test_horizontal_chunks_have_distinct_ranges(self, two_tables_horizontal):
        result = parse_workbook(path=two_tables_horizontal)
        assert result.total_chunks == 2
        c1, c2 = result.chunks[0], result.chunks[1]
        assert c1.cell_range != c2.cell_range

    def test_color_coded_chunks_have_html_with_styles(self, color_coded_tables):
        result = parse_workbook(path=color_coded_tables)
        for chunk in result.chunks:
            # Rendered HTML should include color or background style
            html_lower = chunk.render_html.lower()
            assert "background" in html_lower or "color" in html_lower

    def test_deterministic_across_runs(self, two_tables_vertical):
        r1 = parse_workbook(path=two_tables_vertical)
        r2 = parse_workbook(path=two_tables_vertical)
        assert r1.total_chunks == r2.total_chunks
        for c1, c2 in zip(r1.chunks, r2.chunks):
            assert c1.chunk_id == c2.chunk_id
            assert c1.content_hash == c2.content_hash
