"""Unit tests for cluster-05 in-scope filter.

Covers ``classify_execution_required`` (parser-independent: reads
input.xlsx[answer_position] and returns True iff every cell is empty)
plus the dual-recall surface in ``aggregate``.

Reference: docs/planning/recall-90/05-out-of-scope-execution-instances.md
"""
from __future__ import annotations

import openpyxl
import pytest

from scripts.eval_retrieval import (
    InstanceResult,
    aggregate,
    classify_execution_required,
)


@pytest.fixture
def empty_target_xlsx(tmp_path):
    """A workbook where the answer range has no values — execution-required."""
    p = tmp_path / "empty_target.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Source"
    ws["A2"] = 10
    ws["A3"] = 20
    # B2:B3 deliberately left blank — the question would be "fill column B".
    wb.save(p)
    return p


@pytest.fixture
def populated_target_xlsx(tmp_path):
    """A workbook where the answer range already has the answer values."""
    p = tmp_path / "populated_target.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Source"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B2"] = 100   # answer cells already filled
    ws["B3"] = 200
    wb.save(p)
    return p


def test_empty_answer_range_is_out_of_scope(empty_target_xlsx):
    regions = [("Sheet1", (2, 2, 3, 2))]  # B2:B3
    assert classify_execution_required(empty_target_xlsx, regions) is True


def test_populated_answer_range_is_in_scope(populated_target_xlsx):
    regions = [("Sheet1", (2, 2, 3, 2))]
    assert classify_execution_required(populated_target_xlsx, regions) is False


def test_header_only_answer_range_is_in_scope(tmp_path):
    """Range with only string headers (no numeric data) is still in-scope.

    Per cluster-05 doc pitfalls: 'don't over-exclude' — a single non-empty
    header cell means the question is asking the system to find/explain
    something already in the input, not to compute and write.
    """
    p = tmp_path / "header_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"   # one non-empty cell in the answer range
    wb.save(p)
    regions = [("Sheet1", (1, 1, 5, 3))]  # A1:C5
    assert classify_execution_required(p, regions) is False


def test_uncached_formula_cell_is_out_of_scope(tmp_path):
    """Per cluster-05 doc: a formula cell with no cached value reads as
    None under data_only — the parser has nothing to retrieve, so the
    instance is execution-required.

    This is the boundary with cluster 03: cluster-03 makes the parser
    surface ``=A1+B1`` as a substring fallback; cluster-05 still excludes
    these instances from the headline recall denominator because the
    formula source is not the *value* the question is asking for.
    """
    p = tmp_path / "uncached_formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["B1"] = "=A1+A2"  # openpyxl saves without computing
    wb.save(p)

    regions = [("Sheet1", (1, 2, 1, 2))]  # B1 — only the uncached formula cell
    assert classify_execution_required(p, regions) is True


def test_formula_cell_with_populated_cache_is_in_scope(tmp_path):
    """If we forge a cached value into the workbook (simulating a calc
    engine having run), the cell IS retrievable — in-scope."""
    p = tmp_path / "cached_formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"] = "=10+20"
    wb.save(p)
    # Patch the cached value by re-opening and writing a numeric literal in
    # place. Equivalent to "the calc engine cached '30'".
    wb2 = openpyxl.load_workbook(p)
    wb2["Sheet1"]["B1"] = 30
    wb2.save(p)

    regions = [("Sheet1", (1, 2, 1, 2))]
    assert classify_execution_required(p, regions) is False


def test_empty_regions_input_returns_false(tmp_path):
    """No regions to classify against → in-scope by default (don't drop)."""
    p = tmp_path / "any.xlsx"
    wb = openpyxl.Workbook()
    wb.save(p)
    assert classify_execution_required(p, []) is False


def test_aggregate_emits_both_all_and_in_scope_recall():
    """When `execution_required` is supplied, both metrics are computed
    and the in-scope denominator excludes the flagged instances."""
    # 4 instances: 1, 2 are in-scope; 3, 4 are out-of-scope.
    # On text@5: 1 hits (rank 1), 2 misses (rank None), 3 hits (rank 2),
    # 4 misses (rank None).
    recs = [
        InstanceResult(
            instance_id="1", parser="p", n_chunks=5, parse_ms=10.0,
            data_position="", answer_position="", data_regions=1,
            chunks_overlapping_data=1, rank_of_first_overlap=1,
            rank_of_text_match=1,
        ),
        InstanceResult(
            instance_id="2", parser="p", n_chunks=5, parse_ms=10.0,
            data_position="", answer_position="", data_regions=1,
            chunks_overlapping_data=0, rank_of_first_overlap=None,
            rank_of_text_match=None,
        ),
        InstanceResult(
            instance_id="3", parser="p", n_chunks=5, parse_ms=10.0,
            data_position="", answer_position="", data_regions=1,
            chunks_overlapping_data=1, rank_of_first_overlap=2,
            rank_of_text_match=2,
        ),
        InstanceResult(
            instance_id="4", parser="p", n_chunks=5, parse_ms=10.0,
            data_position="", answer_position="", data_regions=1,
            chunks_overlapping_data=0, rank_of_first_overlap=None,
            rank_of_text_match=None,
        ),
    ]
    exec_map = {"1": False, "2": False, "3": True, "4": True}
    summary = aggregate(recs, execution_required=exec_map)
    m = summary["p"]

    assert m["instances"] == 4
    assert m["in_scope_instances"] == 2
    assert m["out_of_scope_instances"] == 2

    # All-instance recall: 2 hits / 4 = 0.50
    assert m["recall_text@5"] == 0.5
    # In-scope recall: 1 hit / 2 = 0.50 too (different denominator, same ratio here)
    assert m["recall_text@5_in_scope"] == 0.5

    # Make the in-scope subset a strict miss to confirm the denominator
    # changes the ratio: drop instance 1's rank.
    recs[0].rank_of_text_match = None
    summary2 = aggregate(recs, execution_required=exec_map)
    m2 = summary2["p"]
    # All-instance: 1 hit / 4 = 0.25
    assert m2["recall_text@5"] == 0.25
    # In-scope: 0 hits / 2 = 0.0
    assert m2["recall_text@5_in_scope"] == 0.0


def test_aggregate_without_exec_map_in_scope_equals_all():
    """When no execution map is passed, in-scope == all."""
    recs = [
        InstanceResult(
            instance_id="1", parser="p", n_chunks=1, parse_ms=1.0,
            data_position="", answer_position="", data_regions=1,
            chunks_overlapping_data=1, rank_of_first_overlap=1,
            rank_of_text_match=1,
        ),
    ]
    summary = aggregate(recs)
    m = summary["p"]
    assert m["recall_text@5"] == m["recall_text@5_in_scope"] == 1.0
    assert m["out_of_scope_instances"] == 0
    assert m["in_scope_instances"] == 1


# ── harness accuracy fixes (geometric sheet-None + text unscoreable) ────────


def test_overlaps_unspecified_sheet_matches_geometry_only():
    """When the ground-truth sheet is unspecified, geometric overlap must match
    on geometry alone — a real chunk's sheet name can never equal "" and the
    dataset omits the sheet for ~62% of (single-sheet) instances."""
    from scripts.eval_retrieval import Chunk

    c = Chunk(parser="ks", sheet="Sheet2", top_left=(1, 1), bottom_right=(11, 3),
              text="")
    # Unspecified GT sheet ("" or None) → match purely on geometry.
    assert c.overlaps("", (2, 3, 11, 3))      # C2:C11 ⊂ A1:C11
    assert c.overlaps(None, (2, 3, 11, 3))
    # Specified-and-different sheet → no match.
    assert not c.overlaps("Other", (2, 3, 11, 3))
    # Wildcard sheet but disjoint geometry → no match.
    assert not c.overlaps("", (20, 20, 21, 21))


def test_text_recall_excludes_unscoreable_instances():
    """Instances whose answer region has no scoreable values must not count as
    text-match misses (they're excluded from the bucket histogram already)."""
    def mk(iid, rank_text, had_values):
        return InstanceResult(
            instance_id=iid, parser="p", n_chunks=5, parse_ms=1.0,
            data_position="", answer_position="", data_regions=1,
            chunks_overlapping_data=1, rank_of_first_overlap=1,
            rank_of_text_match=rank_text,
            extra={"had_answer_values": had_values},
        )
    recs = [
        mk("a", 1, True),       # scoreable hit
        mk("b", None, True),    # scoreable miss
        mk("c", None, False),   # unscoreable — must be excluded, not a miss
    ]
    m = aggregate(recs)["p"]
    # Text recall excludes "c": 1 hit / 2 scoreable = 0.5 (not 1/3 = 0.33).
    assert m["recall_text@5"] == 0.5
    # Geometric still counts all three (geometry needs no answer values).
    assert m["recall_geometric@5"] == 1.0
