"""Tier-1 chunker-quality improvements in the text renderer.

Three changes, all in `text_renderer.render_block`:

1. Row-number anchors — each data row of the markdown grid carries
   its sheet row number, so an LLM consuming the chunk can compute
   `(row, col)` cell coordinates deterministically. The block header
   already prints the A1 range; per-row anchors close the gap.

2. Number-format-aware value rendering — Excel cells store a raw
   value (`0.06`, `46022`) but DISPLAY a formatted form (`6%`,
   `2025-12-31`) per the cell's `number_format`. When the two differ
   meaningfully, render BOTH so substring-match retrieval hits either
   form (the question may quote either, and `answer.xlsx` may use
   the display form even though `input.xlsx` keeps the raw).

3. Merged-cell value propagation — slave cells in a merged region
   currently render as blank because openpyxl returns `None` for
   them. That kills text-match retrieval whenever a question
   references the cell by a slave coordinate. Render the master's
   value at each slave with a `←` marker indicating propagation.

Each test uses an openpyxl-built fixture and asserts on the chunk's
`render_text` end-to-end (via `parse_workbook`) so the change is
visible to the downstream consumer that matters.
"""
from __future__ import annotations

import openpyxl

from ks_xlsx_parser.api import parse_workbook


def _all_text(workbook_path) -> str:
    result = parse_workbook(path=str(workbook_path))
    return "\n".join(c.render_text or "" for c in result.chunks)


# ────────────────────────────────────────────────── #1 row anchors


def test_row_anchor_appears_for_each_data_row(tmp_path):
    """Every non-hidden data row gets a `r<N>` prefix where N is the
    sheet row number (1-indexed), not the 0-indexed position within
    the block."""
    p = tmp_path / "anchors.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "name"
    ws["B1"] = "qty"
    ws["A2"] = "Widget"
    ws["B2"] = 100
    ws["A3"] = "Hub"
    ws["B3"] = 200
    wb.save(p)

    text = _all_text(p)
    # Every row from the block should be addressable.
    assert "r1" in text, f"missing row-1 anchor; chunk text was:\n{text}"
    assert "r2" in text, f"missing row-2 anchor; chunk text was:\n{text}"
    assert "r3" in text, f"missing row-3 anchor; chunk text was:\n{text}"
    # The anchors should be visibly bound to the right values — `r2`
    # should appear ahead of `Widget` on the same line.
    widget_line = next(
        (line for line in text.splitlines() if "Widget" in line), None,
    )
    assert widget_line is not None and "r2" in widget_line, (
        f"row anchor not on the same line as its data: "
        f"line={widget_line!r}"
    )


def test_row_anchor_uses_sheet_row_not_chunk_offset(tmp_path):
    """A block that starts at row 5 must anchor its first data row as
    `r5`, not `r1`. Citation depends on this."""
    p = tmp_path / "offset.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    # Leave rows 1-4 empty (no header lines, no data); block starts at 5.
    ws["A5"] = "name"
    ws["B5"] = "qty"
    ws["A6"] = "Widget"
    ws["B6"] = 100
    wb.save(p)

    text = _all_text(p)
    assert "r5" in text, f"missing r5 anchor:\n{text}"
    assert "r6" in text, f"missing r6 anchor:\n{text}"


# ────────────────────────────────────────────────── #2 number-format expansion


def test_percent_format_renders_both_forms(tmp_path):
    """A cell storing 0.06 with number_format '0%' must render both
    `0.06` (raw) AND `6%` (displayed). Substring match hits either."""
    p = tmp_path / "percent.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "rate"
    ws["A2"] = 0.06
    ws["A2"].number_format = "0%"
    wb.save(p)

    text = _all_text(p)
    assert "0.06" in text, f"raw form missing:\n{text}"
    assert "6%" in text, f"displayed form missing:\n{text}"


def test_decimal_format_renders_both_forms(tmp_path):
    """Format `#,##0.00` should add the formatted form (`1,272.00`)
    alongside the raw `1272`."""
    p = tmp_path / "decimal.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "amount"
    ws["A2"] = 1272
    ws["A2"].number_format = "#,##0.00"
    wb.save(p)

    text = _all_text(p)
    assert "1272" in text, f"raw form missing:\n{text}"
    assert "1,272.00" in text, f"comma-format missing:\n{text}"


def test_no_format_expansion_when_format_is_general(tmp_path):
    """When number_format is 'General' (the default), do NOT add a
    redundant '(value)' clause — the raw is the display."""
    p = tmp_path / "general.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "amount"
    ws["A2"] = 1272  # default General format
    wb.save(p)

    text = _all_text(p)
    assert "1272" in text
    # No bracketed duplicate of the raw form.
    assert "1272 [1272]" not in text


# ────────────────────────────────────────────────── #3 merged-cell propagation


def test_merged_cell_master_value_propagates_to_slaves(tmp_path):
    """`ws.merge_cells('A1:C1')` with A1='Total' should render the
    string 'Total' at A1 AND a propagation marker at B1 and C1 so the
    full grid carries the visible value at every position it appears
    in Excel."""
    p = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "Total"
    ws.merge_cells("A1:C1")
    ws["A2"] = 1
    ws["B2"] = 2
    ws["C2"] = 3
    wb.save(p)

    text = _all_text(p)
    # The master value must be present at least once.
    assert "Total" in text, f"master value missing:\n{text}"
    # The propagated form contains the master's value with a marker.
    # We accept either "← Total" or any text containing both 'Total'
    # and a propagation indicator on the slave column's cell.
    header_row = next(
        (l for l in text.splitlines() if "Total" in l), None,
    )
    assert header_row is not None, "no row contains 'Total'"
    # Count occurrences of the string `Total` on that row — should be
    # ≥ 2 (master + at least one slave) for the propagation to be
    # observable.
    assert header_row.count("Total") >= 2, (
        f"expected 'Total' to repeat across the merged region; "
        f"row was:\n{header_row}"
    )


def test_unmerged_cells_render_normally(tmp_path):
    """Sanity check: a workbook with no merged regions must produce
    the same shape as before this change (no spurious markers)."""
    p = tmp_path / "no_merge.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "name"
    ws["B1"] = "qty"
    ws["A2"] = "Widget"
    ws["B2"] = 100
    wb.save(p)

    text = _all_text(p)
    # No propagation marker should appear because nothing is merged.
    assert "←" not in text, f"unexpected propagation marker:\n{text}"
