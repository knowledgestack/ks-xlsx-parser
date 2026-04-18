"""
Pytest fixtures for xlsx_parser tests.

Provides programmatically generated .xlsx workbooks covering various
edge cases: merged cells, formulas, tables, large sheets, sparse data,
cross-sheet references, conditional formatting, data validation, etc.
"""



import os
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# All-xlsx-files collection for cross-validation and invariant tests
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).parent.parent
_TESTBENCH_DIR = _PROJECT_ROOT / "testBench"
_EXAMPLES_DIR = _TESTBENCH_DIR / "real_world"
_DATASETS_DIR = _TESTBENCH_DIR / "github_datasets"

# Names of conftest fixtures that produce .xlsx files
PROGRAMMATIC_FIXTURE_NAMES = [
    "simple_workbook",
    "merged_cells_workbook",
    # empty_master_workbook excluded: it's a raw OOXML ZIP (not openpyxl-generated)
    # that intentionally diverges from calamine results (our recovery is better)
    "formula_workbook",
    "table_workbook",
    "chart_workbook",
    "large_sparse_workbook",
    "conditional_format_workbook",
    "data_validation_workbook",
    "multi_sheet_workbook",
    "hidden_rows_cols_workbook",
    "comment_workbook",
    "freeze_panes_workbook",
    "wide_workbook",
    "styled_workbook",
    "assumptions_workbook",
    "hyperlink_workbook",
    "two_tables_vertical",
    "two_tables_horizontal",
    "table_chart_table",
    "table_in_middle",
    "mixed_content_layout",
    "color_coded_tables",
    "complex_headers_layout",
    "simple_formulas",
    "nested_formulas",
    "cross_sheet_formulas",
    "text_formulas",
    "circular_ref_formulas",
    "lookup_formulas",
    "mixed_formula_types",
    "stress_comprehensive",
    "stress_tough",
]


def collect_static_xlsx_files() -> list[Path]:
    """Collect all static .xlsx files from examples and github_datasets."""
    files = []
    for d in [_EXAMPLES_DIR, _DATASETS_DIR]:
        if d.exists():
            files.extend(sorted(d.glob("*.xlsx")))
    return files


STATIC_XLSX_FILES = collect_static_xlsx_files()


@pytest.fixture(params=PROGRAMMATIC_FIXTURE_NAMES)
def programmatic_xlsx(request, tmp_dir) -> Path:
    """Yields each programmatic fixture as a Path (re-uses other fixtures)."""
    return request.getfixturevalue(request.param)


@pytest.fixture(
    params=STATIC_XLSX_FILES,
    ids=[f.stem for f in STATIC_XLSX_FILES],
)
def static_xlsx(request) -> Path:
    """Yields each static .xlsx file path."""
    return request.param


@pytest.fixture
def tmp_dir():
    """Provide a temporary directory for test workbooks."""
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


@pytest.fixture
def simple_workbook(tmp_dir) -> Path:
    """A minimal workbook with a few cells and basic formatting."""
    path = tmp_dir / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    ws["A2"] = "Revenue"
    ws["B2"] = 1000
    ws["B2"].number_format = "#,##0"

    ws["A3"] = "Cost"
    ws["B3"] = 600
    ws["B3"].number_format = "#,##0"

    ws["A4"] = "Profit"
    ws["B4"] = "=B2-B3"
    ws["B4"].font = Font(bold=True, color="008000")
    ws["B4"].number_format = "#,##0"

    wb.save(path)
    return path


@pytest.fixture
def merged_cells_workbook(tmp_dir) -> Path:
    """Workbook with various merged cell configurations."""
    path = tmp_dir / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Merges"

    # Horizontal merge (header)
    ws.merge_cells("A1:D1")
    ws["A1"] = "Quarterly Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Column headers
    ws["A2"] = "Q1"
    ws["B2"] = "Q2"
    ws["C2"] = "Q3"
    ws["D2"] = "Q4"

    # Data with a vertical merge
    ws.merge_cells("A3:A5")
    ws["A3"] = "Region A"
    ws["B3"] = 100
    ws["B4"] = 110
    ws["B5"] = 120

    # Multi-row multi-col merge
    ws.merge_cells("C3:D4")
    ws["C3"] = "Combined"

    wb.save(path)
    return path


@pytest.fixture
def formula_workbook(tmp_dir) -> Path:
    """Workbook with various formula types and cross-sheet references."""
    path = tmp_dir / "formulas.xlsx"
    wb = Workbook()

    # Sheet1: Inputs
    ws1 = wb.active
    ws1.title = "Inputs"
    ws1["A1"] = "Price"
    ws1["B1"] = 50
    ws1["A2"] = "Quantity"
    ws1["B2"] = 200
    ws1["A3"] = "Tax Rate"
    ws1["B3"] = 0.08
    ws1["B3"].number_format = "0%"

    # Sheet2: Calculations
    ws2 = wb.create_sheet("Calculations")
    ws2["A1"] = "Subtotal"
    ws2["B1"] = "=Inputs!B1*Inputs!B2"
    ws2["A2"] = "Tax"
    ws2["B2"] = "=B1*Inputs!B3"
    ws2["A3"] = "Total"
    ws2["B3"] = "=B1+B2"
    ws2["B3"].font = Font(bold=True)

    # Sheet3: Summary with SUM
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "All Totals"
    ws3["B1"] = "=SUM(Calculations!B1:B3)"
    ws3["A2"] = "Average"
    ws3["B2"] = "=AVERAGE(Calculations!B1:B3)"

    # Named range
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName("Price", attr_text="Inputs!$B$1"))
    wb.defined_names.add(DefinedName("Quantity", attr_text="Inputs!$B$2"))

    wb.save(path)
    return path


@pytest.fixture
def array_formula_workbook(tmp_dir) -> Path:
    """Workbook with an array formula (openpyxl ArrayFormula object)."""
    from openpyxl.worksheet.formula import ArrayFormula

    path = tmp_dir / "array_formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"], ws["A2"], ws["A3"] = 1, 2, 3
    ws["B1"] = ArrayFormula("B1:B3", "=A1:A3*2")
    wb.save(path)
    return path


@pytest.fixture
def table_workbook(tmp_dir) -> Path:
    """Workbook with an Excel ListObject table with filters and totals."""
    path = tmp_dir / "tables.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Table headers
    headers = ["Product", "Region", "Q1", "Q2", "Q3", "Q4", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Table data
    data = [
        ["Widget A", "North", 100, 150, 200, 180, None],
        ["Widget A", "South", 80, 120, 160, 140, None],
        ["Widget B", "North", 200, 250, 300, 280, None],
        ["Widget B", "South", 150, 180, 220, 200, None],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        # Total formula
        ws.cell(row=row_idx, column=7, value=f"=SUM(C{row_idx}:F{row_idx})")
        ws.cell(row=row_idx, column=7).number_format = "#,##0"

    # Create table
    tab = Table(displayName="SalesData", ref="A1:G5")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(path)
    return path


@pytest.fixture
def chart_workbook(tmp_dir) -> Path:
    """Workbook with a bar chart."""
    path = tmp_dir / "charts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ChartData"

    # Data for chart
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    values = [1000, 1200, 1100, 1400, 1300, 1500]
    for i, (month, val) in enumerate(zip(months, values), 2):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = val

    # Create chart
    chart = BarChart()
    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

    wb.save(path)
    return path


@pytest.fixture
def large_sparse_workbook(tmp_dir) -> Path:
    """Workbook with data scattered across a large range."""
    path = tmp_dir / "large_sparse.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"

    # Data in top-left
    ws["A1"] = "Start"
    ws["B1"] = 1

    # Data far away
    ws["Z100"] = "Far away"
    ws["Z100"].font = Font(italic=True)

    # Data very far away
    ws.cell(row=1000, column=100, value="Very far")

    wb.save(path)
    return path


@pytest.fixture
def conditional_format_workbook(tmp_dir) -> Path:
    """Workbook with conditional formatting rules."""
    path = tmp_dir / "conditional.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ConditionalFmt"

    # Data
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=f"Item {i}")
        ws.cell(row=i, column=2, value=i * 10)

    # Conditional formatting: highlight cells > 50
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.conditional_formatting.add(
        "B1:B10",
        CellIsRule(operator="greaterThan", formula=["50"], fill=red_fill),
    )

    wb.save(path)
    return path


@pytest.fixture
def data_validation_workbook(tmp_dir) -> Path:
    """Workbook with data validation dropdowns."""
    path = tmp_dir / "validation.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    ws["A1"] = "Status"
    ws["B1"] = "Priority"

    # Dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"Active,Inactive,Pending"',
        allow_blank=True,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(ws["A2"])
    dv.add(ws["A3"])

    ws["A2"] = "Active"
    ws["A3"] = "Pending"

    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_workbook(tmp_dir) -> Path:
    """Workbook with multiple sheets including hidden ones."""
    path = tmp_dir / "multi_sheet.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Visible1"
    ws1["A1"] = "Sheet 1 data"

    ws2 = wb.create_sheet("Hidden")
    ws2.sheet_state = "hidden"
    ws2["A1"] = "Hidden data"

    ws3 = wb.create_sheet("Visible2")
    ws3["A1"] = "Sheet 3 data"
    ws3["A2"] = "=Visible1!A1"

    wb.save(path)
    return path


@pytest.fixture
def hidden_rows_cols_workbook(tmp_dir) -> Path:
    """Workbook with hidden rows and columns."""
    path = tmp_dir / "hidden.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Hidden"

    for r in range(1, 11):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")

    # Hide row 3 and column B
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["B"].hidden = True

    wb.save(path)
    return path


@pytest.fixture
def comment_workbook(tmp_dir) -> Path:
    """Workbook with cell comments."""
    path = tmp_dir / "comments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    from openpyxl.comments import Comment

    ws["A1"] = "Revenue"
    ws["A1"].comment = Comment("Total annual revenue", "Analyst")

    ws["B1"] = 50000
    ws["B1"].comment = Comment("As of Q4 2024", "Manager")

    wb.save(path)
    return path


@pytest.fixture
def freeze_panes_workbook(tmp_dir) -> Path:
    """Workbook with freeze panes."""
    path = tmp_dir / "freeze.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "FrozenHeaders"

    # Headers
    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Value"
    ws.freeze_panes = "A2"

    for i in range(2, 20):
        ws[f"A{i}"] = i - 1
        ws[f"B{i}"] = f"Item {i-1}"
        ws[f"C{i}"] = (i - 1) * 100

    wb.save(path)
    return path


@pytest.fixture
def wide_workbook(tmp_dir) -> Path:
    """Workbook with many columns (simulating wide sheets)."""
    path = tmp_dir / "wide.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "WideSheet"

    # 100 columns of data
    for col in range(1, 101):
        ws.cell(row=1, column=col, value=f"Col{col}")
        ws.cell(row=1, column=col).font = Font(bold=True)
        for row in range(2, 6):
            ws.cell(row=row, column=col, value=col * row)

    wb.save(path)
    return path


@pytest.fixture
def styled_workbook(tmp_dir) -> Path:
    """Workbook with rich formatting: borders, fills, fonts, alignment."""
    path = tmp_dir / "styled.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Styled"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header with fill
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.value = f"Header {col}"
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data with borders
    for row in range(2, 5):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col, value=row * col)
            cell.border = thin_border
            if col == 3:
                cell.number_format = "#,##0.00"

    wb.save(path)
    return path


@pytest.fixture
def assumptions_workbook(tmp_dir) -> Path:
    """Workbook with an assumptions section and results section."""
    path = tmp_dir / "assumptions.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"

    # Assumptions block
    ws["A1"] = "ASSUMPTIONS"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Growth Rate"
    ws["B2"] = 0.05
    ws["B2"].number_format = "0.0%"
    ws["A3"] = "Discount Rate"
    ws["B3"] = 0.10
    ws["B3"].number_format = "0.0%"
    ws["A4"] = "Tax Rate"
    ws["B4"] = 0.21
    ws["B4"].number_format = "0.0%"
    ws["A5"] = "Initial Investment"
    ws["B5"] = 1000000
    ws["B5"].number_format = "#,##0"

    # Blank row separator
    # Row 6 is blank

    # Results block
    ws["A7"] = "RESULTS"
    ws["A7"].font = Font(bold=True, size=12)
    ws["A8"] = "Year 1 Revenue"
    ws["B8"] = "=B5*(1+B2)"
    ws["B8"].number_format = "#,##0"
    ws["A9"] = "Year 1 Tax"
    ws["B9"] = "=B8*B4"
    ws["B9"].number_format = "#,##0"
    ws["A10"] = "Year 1 Net"
    ws["B10"] = "=B8-B9"
    ws["B10"].number_format = "#,##0"
    ws["B10"].font = Font(bold=True, color="008000")

    wb.save(path)
    return path


@pytest.fixture
def hyperlink_workbook(tmp_dir) -> Path:
    """Workbook with hyperlinks."""
    path = tmp_dir / "hyperlinks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Links"

    ws["A1"] = "Google"
    ws["A1"].hyperlink = "https://www.google.com"
    ws["A1"].font = Font(color="0563C1", underline="single")

    ws["A2"] = "Internal ref"

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Multi-table layout fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def two_tables_vertical(tmp_dir) -> Path:
    """Two tables stacked vertically, separated by blank rows."""
    path = tmp_dir / "two_tables_vertical.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "VertLayout"

    # Table A: rows 1-4, columns A-C
    for col_idx, name in enumerate(["Name", "Amount", "Date"], 1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = Font(bold=True)
    for r in range(2, 5):
        ws.cell(row=r, column=1, value=f"Item {r - 1}")
        ws.cell(row=r, column=2, value=r * 100)
        ws.cell(row=r, column=3, value=f"2024-0{r}")

    # Rows 5-6 blank (gap)

    # Table B: rows 7-10, columns A-B
    for col_idx, name in enumerate(["Category", "Score"], 1):
        c = ws.cell(row=7, column=col_idx, value=name)
        c.font = Font(bold=True)
    for r in range(8, 11):
        ws.cell(row=r, column=1, value=f"Cat {r - 7}")
        ws.cell(row=r, column=2, value=(r - 7) * 25)

    wb.save(path)
    return path


@pytest.fixture
def two_tables_horizontal(tmp_dir) -> Path:
    """Two tables side by side, separated by an empty column."""
    path = tmp_dir / "two_tables_horizontal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "HorizLayout"

    # Table A: cols A-C (1-3), rows 1-4
    for col_idx, name in enumerate(["Product", "Price", "Stock"], 1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = Font(bold=True)
    for r in range(2, 5):
        ws.cell(row=r, column=1, value=f"Prod {r - 1}")
        ws.cell(row=r, column=2, value=r * 10.5)
        ws.cell(row=r, column=3, value=r * 50)

    # Column D (col=4) is empty gap

    # Table B: cols E-F (5-6), rows 1-4
    for col_idx, name in [(5, "Region"), (6, "Revenue")]:
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = Font(bold=True)
    for r in range(2, 5):
        ws.cell(row=r, column=5, value=f"Region {r - 1}")
        ws.cell(row=r, column=6, value=r * 1000)

    wb.save(path)
    return path


@pytest.fixture
def table_chart_table(tmp_dir) -> Path:
    """Table, chart in the middle, then another table."""
    path = tmp_dir / "table_chart_table.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ChartBetween"

    # Table 1: rows 1-5, columns A-B
    ws.cell(row=1, column=1, value="Month").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Sales").font = Font(bold=True)
    months = ["Jan", "Feb", "Mar", "Apr"]
    values = [100, 200, 150, 250]
    for i, (m, v) in enumerate(zip(months, values), 2):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)

    # Chart placed at D7 (does not create cell data)
    chart = BarChart()
    chart.title = "Sales Chart"
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D7")

    # Table 2: rows 10-14, columns A-C
    for col_idx, name in enumerate(["Quarter", "Target", "Actual"], 1):
        ws.cell(row=10, column=col_idx, value=name).font = Font(bold=True)
    for r in range(11, 15):
        ws.cell(row=r, column=1, value=f"Q{r - 10}")
        ws.cell(row=r, column=2, value=(r - 10) * 500)
        ws.cell(row=r, column=3, value=(r - 10) * 480)

    wb.save(path)
    return path


@pytest.fixture
def table_in_middle(tmp_dir) -> Path:
    """Single table centered in the sheet with empty space around it."""
    path = tmp_dir / "table_in_middle.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Centered"

    # Table at rows 5-9, columns D-F (4-6)
    for col_idx, name in [(4, "X"), (5, "Y"), (6, "Z")]:
        ws.cell(row=5, column=col_idx, value=name).font = Font(bold=True)
    for r in range(6, 10):
        ws.cell(row=r, column=4, value=f"x{r - 5}")
        ws.cell(row=r, column=5, value=(r - 5) * 3.14)
        ws.cell(row=r, column=6, value=(r - 5) ** 2)

    wb.save(path)
    return path


@pytest.fixture
def mixed_content_layout(tmp_dir) -> Path:
    """Mixed content: header, table, text block, assumptions block."""
    path = tmp_dir / "mixed_content.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"

    # Header block (row 1)
    ws["A1"] = "REPORT TITLE"
    ws["A1"].font = Font(bold=True, size=14)

    # Row 2 blank

    # Table block (rows 3-6, cols A-C)
    for col_idx, name in enumerate(["Metric", "Value", "Unit"], 1):
        ws.cell(row=3, column=col_idx, value=name).font = Font(bold=True)
    ws.cell(row=4, column=1, value="Revenue")
    ws.cell(row=4, column=2, value=50000)
    ws.cell(row=4, column=3, value="USD")
    ws.cell(row=5, column=1, value="Costs")
    ws.cell(row=5, column=2, value=30000)
    ws.cell(row=5, column=3, value="USD")
    ws.cell(row=6, column=1, value="Profit")
    ws.cell(row=6, column=2, value=20000)
    ws.cell(row=6, column=3, value="USD")

    # Row 7 blank

    # Text block (rows 8-9)
    ws["A8"] = "Notes:"
    ws["A9"] = "Refer to appendix B for methodology."

    # Rows 10-11 blank

    # Assumptions block (rows 12-13)
    ws["A12"] = "Assumption"
    ws["A12"].font = Font(bold=True)
    ws["B12"] = "Value"
    ws["B12"].font = Font(bold=True)
    ws["A13"] = "Growth Rate Input"
    ws["B13"] = 0.05
    ws["B13"].number_format = "0.0%"

    wb.save(path)
    return path


@pytest.fixture
def color_coded_tables(tmp_dir) -> Path:
    """Two tables with distinct color schemes (blue and green)."""
    path = tmp_dir / "color_coded.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ColorTables"

    blue_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    blue_data_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")

    green_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    green_data_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Table A: rows 1-5, blue scheme
    for col_idx, name in enumerate(["Sales Rep", "Region", "Total"], 1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = white_font
        c.fill = blue_header_fill
    for r in range(2, 6):
        for col_idx in range(1, 4):
            ws.cell(row=r, column=col_idx).fill = blue_data_fill
        ws.cell(row=r, column=1, value=f"Rep {r - 1}")
        ws.cell(row=r, column=2, value=f"Region {r - 1}")
        ws.cell(row=r, column=3, value=(r - 1) * 1000)

    # Rows 6-7 blank

    # Table B: rows 8-12, green scheme
    for col_idx, name in enumerate(["Product", "Category", "Qty"], 1):
        c = ws.cell(row=8, column=col_idx, value=name)
        c.font = white_font
        c.fill = green_header_fill
    for r in range(9, 13):
        for col_idx in range(1, 4):
            ws.cell(row=r, column=col_idx).fill = green_data_fill
        ws.cell(row=r, column=1, value=f"Product {r - 8}")
        ws.cell(row=r, column=2, value=f"Cat {r - 8}")
        ws.cell(row=r, column=3, value=(r - 8) * 50)

    wb.save(path)
    return path


@pytest.fixture
def complex_headers_layout(tmp_dir) -> Path:
    """Tables with merged header rows and multi-color column headers."""
    path = tmp_dir / "complex_headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ComplexHdr"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    col_fills = [
        PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
        PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    ]

    # Block 1: Merged title + column headers + data (rows 1-6)
    ws.merge_cells("A1:C1")
    ws["A1"] = "Financial Summary"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    col_names = ["Revenue", "Cost", "Profit"]
    for col_idx, (name, fill) in enumerate(zip(col_names, col_fills), 1):
        c = ws.cell(row=2, column=col_idx, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill

    for r in range(3, 7):
        ws.cell(row=r, column=1, value=r * 1000)
        ws.cell(row=r, column=2, value=r * 600)
        ws.cell(row=r, column=3, value=f"=A{r}-B{r}")

    # Row 7 blank

    # Block 2: Merged title + headers + data (rows 8-12)
    ws.merge_cells("A8:B8")
    ws["A8"] = "Operational Metrics"
    ws["A8"].font = Font(bold=True, size=12)
    ws["A8"].alignment = Alignment(horizontal="center")

    ws.cell(row=9, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=9, column=2, value="Value").font = Font(bold=True)

    ws.cell(row=10, column=1, value="Efficiency")
    ws.cell(row=10, column=2, value=0.92)
    ws.cell(row=11, column=1, value="Utilization")
    ws.cell(row=11, column=2, value=0.85)
    ws.cell(row=12, column=1, value="Throughput")
    ws.cell(row=12, column=2, value=1500)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Formula Test Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def simple_formulas(tmp_dir) -> Path:
    """Workbook with basic arithmetic and function formulas."""
    path = tmp_dir / "simple_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"

    # Header row
    ws["A1"] = "Value_A"
    ws["B1"] = "Value_B"
    ws["C1"] = "Sum"
    ws["D1"] = "Diff"
    ws["E1"] = "Product"
    ws["F1"] = "Division"
    for col in range(1, 7):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Data rows
    data = [(10, 5), (20, 8), (30, 12), (100, 25), (50, 10)]
    for i, (a, b) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=f"=A{i}+B{i}")
        ws.cell(row=i, column=4, value=f"=A{i}-B{i}")
        ws.cell(row=i, column=5, value=f"=A{i}*B{i}")
        ws.cell(row=i, column=6, value=f"=A{i}/B{i}")

    # Summary row with aggregate functions
    ws.cell(row=7, column=1, value="=SUM(A2:A6)")
    ws.cell(row=7, column=2, value="=SUM(B2:B6)")
    ws.cell(row=7, column=3, value="=SUM(C2:C6)")
    ws.cell(row=7, column=4, value="=AVERAGE(D2:D6)")
    ws.cell(row=7, column=5, value="=MAX(E2:E6)")
    ws.cell(row=7, column=6, value="=MIN(F2:F6)")

    # Additional functions
    ws.cell(row=8, column=1, value="=COUNT(A2:A6)")
    ws.cell(row=8, column=2, value="=COUNTA(B2:B6)")

    wb.save(path)
    return path


@pytest.fixture
def nested_formulas(tmp_dir) -> Path:
    """Workbook with nested and conditional formulas."""
    path = tmp_dir / "nested_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Nested"

    ws["A1"] = "Score"
    ws["B1"] = "Grade"
    ws["C1"] = "Status"
    ws["D1"] = "Adjusted"
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = Font(bold=True)

    scores = [95, 82, 67, 45, 78, 91, 55, 88]
    for i, score in enumerate(scores, start=2):
        ws.cell(row=i, column=1, value=score)
        # Nested IF for grade
        ws.cell(row=i, column=2, value=f'=IF(A{i}>=90,"A",IF(A{i}>=80,"B",IF(A{i}>=70,"C",IF(A{i}>=60,"D","F"))))')
        # IF with AND
        ws.cell(row=i, column=3, value=f'=IF(AND(A{i}>=60,A{i}<90),"Pass","Review")')
        # IFERROR with division
        ws.cell(row=i, column=4, value=f"=IFERROR(A{i}/(A{i}-50),0)")

    # SUMIF / COUNTIF
    ws.cell(row=11, column=1, value="Passing:")
    ws.cell(row=11, column=2, value='=COUNTIF(A2:A9,">=60")')
    ws.cell(row=12, column=1, value="Sum Pass:")
    ws.cell(row=12, column=2, value='=SUMIF(A2:A9,">=60")')

    wb.save(path)
    return path


@pytest.fixture
def cross_sheet_formulas(tmp_dir) -> Path:
    """Workbook with formulas referencing cells across sheets."""
    path = tmp_dir / "cross_sheet_formulas.xlsx"
    wb = Workbook()

    # Sheet 1: Source data
    ws1 = wb.active
    ws1.title = "Revenue"
    ws1["A1"] = "Q1"
    ws1["B1"] = "Q2"
    ws1["C1"] = "Q3"
    ws1["D1"] = "Q4"
    for col in range(1, 5):
        ws1.cell(row=1, column=col).font = Font(bold=True)
    ws1["A2"] = 10000
    ws1["B2"] = 15000
    ws1["C2"] = 12000
    ws1["D2"] = 18000

    # Sheet 2: Costs
    ws2 = wb.create_sheet("Costs")
    ws2["A1"] = "Q1"
    ws2["B1"] = "Q2"
    ws2["C1"] = "Q3"
    ws2["D1"] = "Q4"
    for col in range(1, 5):
        ws2.cell(row=1, column=col).font = Font(bold=True)
    ws2["A2"] = 6000
    ws2["B2"] = 8000
    ws2["C2"] = 7000
    ws2["D2"] = 9000

    # Sheet 3: Summary with cross-sheet formulas
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Quarter"
    ws3["B1"] = "Revenue"
    ws3["C1"] = "Cost"
    ws3["D1"] = "Profit"
    ws3["E1"] = "Margin %"
    for col in range(1, 6):
        ws3.cell(row=1, column=col).font = Font(bold=True)

    quarters = ["Q1", "Q2", "Q3", "Q4"]
    rev_cols = ["A", "B", "C", "D"]
    for i, (q, col) in enumerate(zip(quarters, rev_cols), start=2):
        ws3.cell(row=i, column=1, value=q)
        ws3.cell(row=i, column=2, value=f"=Revenue!{col}2")
        ws3.cell(row=i, column=3, value=f"=Costs!{col}2")
        ws3.cell(row=i, column=4, value=f"=B{i}-C{i}")
        ws3.cell(row=i, column=5, value=f"=D{i}/B{i}")

    # Total row
    ws3.cell(row=6, column=1, value="Total")
    ws3.cell(row=6, column=2, value="=SUM(B2:B5)")
    ws3.cell(row=6, column=3, value="=SUM(C2:C5)")
    ws3.cell(row=6, column=4, value="=SUM(D2:D5)")
    ws3.cell(row=6, column=5, value="=D6/B6")

    wb.save(path)
    return path


@pytest.fixture
def text_formulas(tmp_dir) -> Path:
    """Workbook with text manipulation formulas."""
    path = tmp_dir / "text_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TextFormulas"

    ws["A1"] = "First"
    ws["B1"] = "Last"
    ws["C1"] = "Full Name"
    ws["D1"] = "Upper"
    ws["E1"] = "Initials"
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = Font(bold=True)

    names = [("John", "Doe"), ("Jane", "Smith"), ("Bob", "Wilson")]
    for i, (first, last) in enumerate(names, start=2):
        ws.cell(row=i, column=1, value=first)
        ws.cell(row=i, column=2, value=last)
        ws.cell(row=i, column=3, value=f'=CONCATENATE(A{i}," ",B{i})')
        ws.cell(row=i, column=4, value=f"=UPPER(C{i})")
        ws.cell(row=i, column=5, value=f'=CONCATENATE(LEFT(A{i},1),LEFT(B{i},1))')

    # LEN and TRIM
    ws.cell(row=5, column=1, value="  padded  ")
    ws.cell(row=5, column=2, value="=LEN(A5)")
    ws.cell(row=5, column=3, value="=TRIM(A5)")
    ws.cell(row=5, column=4, value="=LEN(C5)")

    wb.save(path)
    return path


@pytest.fixture
def circular_ref_formulas(tmp_dir) -> Path:
    """Workbook with circular formula references."""
    path = tmp_dir / "circular_refs.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Circular"

    # Direct circular: A1 -> B1 -> A1
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"

    # Indirect circular: A3 -> B3 -> C3 -> A3
    ws["A3"] = "=C3*2"
    ws["B3"] = "=A3+10"
    ws["C3"] = "=B3-5"

    # Non-circular reference chain for comparison
    ws["A5"] = 100
    ws["B5"] = "=A5*2"
    ws["C5"] = "=B5+A5"

    wb.save(path)
    return path


@pytest.fixture
def lookup_formulas(tmp_dir) -> Path:
    """Workbook with VLOOKUP and INDEX/MATCH formulas."""
    path = tmp_dir / "lookup_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Lookups"

    # Lookup table
    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Price"
    for col in range(1, 4):
        ws.cell(row=1, column=col).font = Font(bold=True)

    items = [(101, "Widget", 9.99), (102, "Gadget", 19.99),
             (103, "Gizmo", 14.99), (104, "Doohickey", 4.99)]
    for i, (id_, name, price) in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=id_)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=price)

    # Lookup section
    ws["E1"] = "Search ID"
    ws["F1"] = "VLOOKUP Name"
    ws["G1"] = "INDEX/MATCH Price"
    for col in range(5, 8):
        ws.cell(row=1, column=col).font = Font(bold=True)

    ws["E2"] = 102
    ws["F2"] = "=VLOOKUP(E2,A2:C5,2,FALSE)"
    ws["G2"] = "=INDEX(C2:C5,MATCH(E2,A2:A5,0))"

    ws["E3"] = 104
    ws["F3"] = "=VLOOKUP(E3,A2:C5,2,FALSE)"
    ws["G3"] = "=INDEX(C2:C5,MATCH(E3,A2:A5,0))"

    wb.save(path)
    return path


@pytest.fixture
def mixed_formula_types(tmp_dir) -> Path:
    """Workbook combining many formula types on one sheet for dependency testing."""
    path = tmp_dir / "mixed_formula_types.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = "Month"
    ws1["B1"] = "Sales"
    ws1["C1"] = "Target"
    for col in range(1, 4):
        ws1.cell(row=1, column=col).font = Font(bold=True)

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    sales = [1000, 1200, 900, 1500, 1100, 1800]
    for i, (m, s) in enumerate(zip(months, sales), start=2):
        ws1.cell(row=i, column=1, value=m)
        ws1.cell(row=i, column=2, value=s)
        ws1.cell(row=i, column=3, value=1100)  # Target

    # Analysis sheet with formulas referencing Data
    ws2 = wb.create_sheet("Analysis")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    for col in range(1, 3):
        ws2.cell(row=1, column=col).font = Font(bold=True)

    ws2["A2"] = "Total Sales"
    ws2["B2"] = "=SUM(Data!B2:B7)"
    ws2["A3"] = "Avg Sales"
    ws2["B3"] = "=AVERAGE(Data!B2:B7)"
    ws2["A4"] = "Max Month"
    ws2["B4"] = "=MAX(Data!B2:B7)"
    ws2["A5"] = "Min Month"
    ws2["B5"] = "=MIN(Data!B2:B7)"
    ws2["A6"] = "Above Target"
    ws2["B6"] = '=COUNTIF(Data!B2:B7,">"&Data!C2)'
    ws2["A7"] = "Hit Rate"
    ws2["B7"] = "=B6/COUNT(Data!B2:B7)"
    ws2["A8"] = "Variance"
    ws2["B8"] = "=B4-B5"

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Empty-master merge fixture (merge_empty_master issue)
# ---------------------------------------------------------------------------


@pytest.fixture
def empty_master_workbook(tmp_dir) -> Path:
    """Workbook simulating the merge_empty_master issue via raw OOXML.

    When Excel merges cells, it may keep values in non-master <c> elements.
    openpyxl's MergedCell ignores these, so the parser's OOXML recovery
    must read the raw XML to find them. openpyxl's own save() clears
    non-master values, so we construct the ZIP directly to match real
    Excel behavior.
    """
    import zipfile

    path = tmp_dir / "empty_master.xlsx"

    # Minimal OOXML structure with values in non-master merged cells
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '</Types>'
    )

    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>'
    )

    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '</Relationships>'
    )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="EmptyMaster" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )

    # Sheet with:
    #   B1 = "Recovered Text" (shared string index 0), A1 empty — merged A1:B1
    #   B2 = 42 (number), A2 empty — merged A2:B2
    #   A3 = "Master Has Value" (shared string index 1) — merged A3:B3
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>'
        '<row r="1"><c r="B1" t="s"><v>0</v></c></row>'
        '<row r="2"><c r="B2"><v>42</v></c></row>'
        '<row r="3"><c r="A3" t="s"><v>1</v></c></row>'
        '</sheetData>'
        '<mergeCells count="3">'
        '<mergeCell ref="A1:B1"/>'
        '<mergeCell ref="A2:B2"/>'
        '<mergeCell ref="A3:B3"/>'
        '</mergeCells>'
        '</worksheet>'
    )

    shared_strings_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="2" uniqueCount="2">'
        '<si><t>Recovered Text</t></si>'
        '<si><t>Master Has Value</t></si>'
        '</sst>'
    )

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        zf.writestr("xl/sharedStrings.xml", shared_strings_xml)

    return path


# ---------------------------------------------------------------------------
# Comprehensive Stress Fixture (tough coverage)
# ---------------------------------------------------------------------------


@pytest.fixture
def stress_comprehensive(tmp_dir) -> Path:
    """Tough workbook: tab colors, sheet-scoped names, number formats, ColorScale, IconSet, print area."""
    path = tmp_dir / "stress_comprehensive.xlsx"
    wb = Workbook()

    # Sheet 1: Tab color + sheet-scoped defined name + print area
    ws1 = wb.active
    ws1.title = "StressMain"
    ws1.sheet_properties.tabColor = "4472C4"  # Blue tab

    ws1["A1"] = "KPI"
    ws1["B1"] = "Value"
    ws1["C1"] = "Status"
    for col in range(1, 4):
        ws1.cell(row=1, column=col).font = Font(bold=True)

    ws1["A2"] = "Revenue"
    ws1["B2"] = 1250000
    ws1["B2"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Accounting
    ws1["A3"] = "Growth %"
    ws1["B3"] = 0.0923
    ws1["B3"].number_format = "0.00%"
    ws1["A4"] = "Fraction"
    ws1["B4"] = 0.375
    ws1["B4"].number_format = "# ?/?"
    ws1["A5"] = "Date"
    ws1["B5"] = 45300  # Excel date serial
    ws1["B5"].number_format = "yyyy-mm-dd"

    for row in range(2, 6):
        ws1.cell(row=row, column=3, value=row * 20)

    # ColorScale conditional formatting
    ws1.conditional_formatting.add(
        "B2:B5",
        ColorScaleRule(
            start_type="percentile", start_value=0, start_color="FF638EC6",
            end_type="percentile", end_value=100, end_color="FF9BC2E6",
        ),
    )
    # IconSet on column C
    ws1.conditional_formatting.add(
        "C2:C5",
        IconSetRule("3Arrows", "percent", [0, 33, 67]),
    )

    ws1.print_area = "A1:C5"
    ws1.page_margins.left = 0.5
    ws1.page_margins.right = 0.5

    # Workbook-level defined name
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName("MainRevenue", attr_text="StressMain!$B$2"))

    # Sheet 2: Hidden + very different content
    ws2 = wb.create_sheet("StressAux")
    ws2.sheet_state = "hidden"
    ws2.sheet_properties.tabColor = "ED7D31"
    ws2["A1"] = "Lookup"
    ws2["B1"] = "=StressMain!B2*1.1"
    ws2["A2"] = "Reference"
    ws2["B2"] = "=StressMain!B3"

    wb.save(path)
    return path


@pytest.fixture
def stress_tough(tmp_dir) -> Path:
    """Tougher: table with formulas, chart, external ref placeholder, print titles."""
    path = tmp_dir / "stress_tough.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "ToughData"
    ws1.sheet_properties.tabColor = "70AD47"

    # Table with formulas
    headers = ["Id", "Name", "Qty", "Price", "Total"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h).font = Font(bold=True)
    for r in range(2, 6):
        ws1.cell(row=r, column=1, value=r - 1)
        ws1.cell(row=r, column=2, value=f"Item {r - 1}")
        ws1.cell(row=r, column=3, value=(r - 1) * 10)
        ws1.cell(row=r, column=4, value=9.99)
        ws1.cell(row=r, column=5, value=f"=C{r}*D{r}")
    tab = Table(displayName="ToughTable", ref="A1:E5")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws1.add_table(tab)

    # Chart
    chart = BarChart()
    chart.title = "Quantities"
    data = Reference(ws1, min_col=3, min_row=1, max_row=5)
    cats = Reference(ws1, min_col=2, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws1.add_chart(chart, "G2")

    # External ref placeholder
    ws1["A8"] = "External"
    ws1["B8"] = "=[OtherWorkbook.xlsx]Sheet1!$A$1"
    ws1["A9"] = "Named"
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName("ToughTotal", attr_text="ToughData!$E$5"))
    ws1["B9"] = "=ToughTotal*2"

    ws1.print_title_rows = "1:1"

    wb.save(path)
    return path
