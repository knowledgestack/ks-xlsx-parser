"""
Tests for legacy ``.xls`` (BIFF) support.

The parser is openpyxl-based and only reads OOXML; ``.xls`` workbooks are
converted to ``.xlsx`` in-memory before the pipeline runs. Conversion is tiered:

  * **LibreOffice** (full fidelity) — preserves formula text, charts, shapes.
  * **xlrd** (pure-Python fallback) — values + styling only; formulas/charts lost.

These tests cover detection, both backends, and end-to-end parsing. Tests that
assert formula-text recovery are skipped when LibreOffice is not installed.
"""

import datetime
import io

import pytest
from openpyxl import load_workbook

from excel_parser.parsers import WorkbookParser
from excel_parser.parsers.xls_converter import (
    XlsConversionError,
    convert_xls_to_xlsx,
    is_legacy_xls,
    libreoffice_available,
)
from excel_parser.pipeline import parse_workbook

requires_libreoffice = pytest.mark.skipif(
    not libreoffice_available(),
    reason="LibreOffice not installed; full-fidelity .xls path unavailable",
)


class TestIsLegacyXls:
    """Detection of legacy .xls inputs."""

    def test_xls_extension(self):
        assert is_legacy_xls(filename="book.xls") is True
        assert is_legacy_xls(filename="BOOK.XLS") is True

    def test_xlsx_and_xlsm_are_not_legacy(self):
        assert is_legacy_xls(filename="book.xlsx") is False
        assert is_legacy_xls(filename="book.xlsm") is False
        assert is_legacy_xls(filename="book.xlsb") is False

    def test_ole2_magic_fallback_when_name_uninformative(self):
        ole2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 16
        assert is_legacy_xls(filename="data.bin", content=ole2) is True
        # A zip-based (OOXML) blob with no useful name is not legacy.
        assert is_legacy_xls(filename="data.bin", content=b"PK\x03\x04rest") is False

    def test_extension_beats_magic(self):
        # Even with OLE2 magic, an explicit .xlsx name wins (encrypted xlsx).
        ole2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        assert is_legacy_xls(filename="book.xlsx", content=ole2) is False


class TestConvertXlrdFallback:
    """The pure-Python xlrd backend (forced via prefer_libreoffice=False)."""

    def test_returns_xlsx_bytes_and_formula_caveat(self, simple_xls_workbook):
        xlsx_bytes, warnings = convert_xls_to_xlsx(
            path=simple_xls_workbook, prefer_libreoffice=False
        )
        assert xlsx_bytes[:2] == b"PK"  # OOXML is a zip container
        # The fallback path always surfaces the formula/charts caveat.
        assert any("formula" in w.lower() for w in warnings)

    def test_converted_workbook_opens_in_openpyxl(self, simple_xls_workbook):
        xlsx_bytes, _ = convert_xls_to_xlsx(
            path=simple_xls_workbook, prefer_libreoffice=False
        )
        ws = load_workbook(io.BytesIO(xlsx_bytes))["Sheet1"]
        assert ws["A1"].value == "Name"
        assert ws["B2"].value == 1000
        assert ws["B4"].value == 400
        assert isinstance(ws["B5"].value, (datetime.date, datetime.datetime))
        assert ws["B6"].value is True

    def test_convert_from_content_bytes(self, simple_xls_workbook):
        raw = simple_xls_workbook.read_bytes()
        xlsx_bytes, _ = convert_xls_to_xlsx(content=raw, prefer_libreoffice=False)
        assert xlsx_bytes[:2] == b"PK"

    def test_invalid_bytes_raise(self):
        with pytest.raises(XlsConversionError):
            convert_xls_to_xlsx(content=b"not an xls file", prefer_libreoffice=False)


class TestConvertLibreOffice:
    """The full-fidelity LibreOffice backend."""

    @requires_libreoffice
    def test_preserves_formula_text(self, formula_xls_workbook):
        xlsx_bytes, warnings = convert_xls_to_xlsx(path=formula_xls_workbook)
        # Full fidelity → no lossy-conversion warnings.
        assert warnings == []
        ws = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)["Sheet1"]
        assert ws["B3"].value == "=B1+B2"
        assert ws["B4"].value == "=AVERAGE(B1:B2)"

    @requires_libreoffice
    def test_preserves_cached_values(self, formula_xls_workbook):
        xlsx_bytes, _ = convert_xls_to_xlsx(path=formula_xls_workbook)
        ws = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)["Sheet1"]
        assert ws["B3"].value == 30
        assert ws["B4"].value == 15

    def test_falls_back_when_disabled(self, formula_xls_workbook, monkeypatch):
        # With LibreOffice disabled the xlrd path runs and emits the caveat.
        monkeypatch.setenv("EXCEL_PARSER_DISABLE_SOFFICE", "1")
        _, warnings = convert_xls_to_xlsx(path=formula_xls_workbook)
        assert any("formula" in w.lower() for w in warnings)


class TestWorkbookParserXls:
    """End-to-end parsing of .xls through WorkbookParser (either backend)."""

    def test_parse_values_and_types(self, simple_xls_workbook):
        result = WorkbookParser(path=simple_xls_workbook).parse()

        assert result.total_sheets == 1
        sheet = result.sheets[0]
        assert sheet.sheet_name == "Sheet1"

        assert sheet.get_cell(1, 1).raw_value == "Name"
        assert sheet.get_cell(2, 2).raw_value == 1000
        assert sheet.get_cell(4, 2).raw_value == 400

    def test_hash_is_from_original_xls_bytes(self, simple_xls_workbook):
        """The workbook hash should track the source .xls, not the conversion."""
        import xxhash

        result = WorkbookParser(path=simple_xls_workbook).parse()
        expected = xxhash.xxh64(simple_xls_workbook.read_bytes()).hexdigest()
        assert result.workbook_hash == expected

    def test_filename_preserved(self, simple_xls_workbook):
        result = WorkbookParser(path=simple_xls_workbook).parse()
        assert result.filename == "simple.xls"

    def test_parse_from_content(self, simple_xls_workbook):
        raw = simple_xls_workbook.read_bytes()
        result = WorkbookParser(content=raw, filename="simple.xls").parse()
        assert result.total_sheets == 1
        assert result.sheets[0].get_cell(2, 2).raw_value == 1000

    def test_multi_sheet_and_merges(self, multi_sheet_xls_workbook):
        result = WorkbookParser(path=multi_sheet_xls_workbook).parse()
        assert result.total_sheets == 2
        names = [s.sheet_name for s in result.sheets]
        assert names == ["Summary", "Detail"]

        summary = result.sheets[0]
        # Merged A1:C1 title survives the conversion.
        assert any(
            mr.range.top_left.row == 1 and mr.range.top_left.col == 1
            for mr in summary.merged_regions
        )

    @requires_libreoffice
    def test_formulas_recovered_end_to_end(self, formula_xls_workbook):
        """Via LibreOffice, formula text reaches the parsed sheet DTO."""
        result = WorkbookParser(path=formula_xls_workbook).parse()
        sheet = result.sheets[0]
        b3 = sheet.get_cell(3, 2)
        assert b3.formula is not None
        assert "B1" in b3.formula and "B2" in b3.formula
        # Cached value still renders.
        assert str(b3.display_value) in {"30", "30.0"}
        # Full fidelity → no formula-loss warning recorded.
        msgs = " ".join(e.message.lower() for e in result.errors)
        assert "formula text is not recoverable" not in msgs

    def test_fallback_warning_recorded_when_soffice_disabled(
        self, simple_xls_workbook, monkeypatch
    ):
        monkeypatch.setenv("EXCEL_PARSER_DISABLE_SOFFICE", "1")
        result = WorkbookParser(path=simple_xls_workbook).parse()
        msgs = " ".join(e.message.lower() for e in result.errors)
        assert "formula" in msgs


class TestPipelineXls:
    """The public parse_workbook entry point handles .xls."""

    def test_parse_workbook_xls(self, simple_xls_workbook):
        result = parse_workbook(path=simple_xls_workbook)
        assert result.workbook.total_sheets == 1
        assert result.total_chunks > 0
        # Rendered chunk text should contain a value from the sheet.
        text = "\n".join(c.render_text for c in result.chunks)
        assert "Revenue" in text
