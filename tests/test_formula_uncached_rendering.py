"""Regression tests for cluster 03: formula cells without a cached value.

A workbook saved by LibreOffice or generated programmatically often has
formula cells whose computed value never got cached (`data_only` returns
None). The renderer must NOT emit `None` or an empty grid cell — it
must surface the formula source string so retrieval has something to
match against.

Reference: docs/planning/recall-90/03-cell-drop-or-uncached-formula.md
"""
from __future__ import annotations

import openpyxl


def test_uncached_formula_renders_formula_source(tmp_path):
    """openpyxl-saved formula → no calc engine ran → cached value is None.

    The chunk's render_text must contain the formula source verbatim so
    an LLM / embedding-search hit can find the cell at all.
    """
    from excel_parser.api import parse_workbook

    p = tmp_path / "uncached.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = 30
    ws["B1"] = "=SUM(A1:A3)"
    wb.save(p)

    result = parse_workbook(path=str(p))
    assert result.chunks, "no chunks produced"
    blob = "\n".join(c.render_text or "" for c in result.chunks)

    assert "SUM(A1:A3)" in blob, (
        f"uncached formula cell dropped from render_text:\n{blob}"
    )
    # And explicitly NOT the Python None repr.
    for c in result.chunks:
        text = c.render_text or ""
        # Look at cell-grid rows (lines starting with '|') for None leaks.
        for line in text.splitlines():
            if line.startswith("|") and "None" in line:
                # Sometimes string literal "None" is legit; only fail when
                # the cell IS None-rendered, which would show as a bare
                # ` None ` cell, not part of "Nondisclosure" etc.
                assert " None " not in line and not line.endswith("None |"), (
                    f"None leaked into a rendered cell: {line!r}"
                )


def test_multiple_uncached_formulas_all_surface(tmp_path):
    """Cluster-03's named instances were dropping rows of formula cells.
    Verify a column of formulas all show up in render_text."""
    from excel_parser.api import parse_workbook

    p = tmp_path / "multi_uncached.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Inputs"
    ws["B1"] = "Doubled"
    for i in range(2, 7):
        ws.cell(row=i, column=1, value=i * 10)
        ws.cell(row=i, column=2, value=f"=A{i}*2")
    wb.save(p)

    result = parse_workbook(path=str(p))
    blob = "\n".join(c.render_text or "" for c in result.chunks)

    # Every formula source must appear somewhere in render_text.
    for i in range(2, 7):
        assert f"A{i}*2" in blob, (
            f"formula =A{i}*2 dropped from render_text:\n{blob}"
        )
