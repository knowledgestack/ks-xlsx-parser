"""
Tests for column-header span detection, including multi-row header bands.

The detector must (a) keep single-row headers exactly one row, (b) extend over
contiguous *multi-column* label rows that sit above the data, and (c) stop at
the first data row or a single-cell section divider — never swallowing data or
dividers into the header band.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from excel_parser.analysis.header_detector import find_header_span
from excel_parser.parsers.workbook_parser import WorkbookParser


def _sheet(build):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    build(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return WorkbookParser(content=buf.getvalue(), filename="x.xlsx").parse().sheets[0]


def _span(sheet):
    return find_header_span(sheet, sheet.compute_used_range())


def test_single_row_header_stays_one_row():
    def build(ws):
        for ci, h in enumerate(["Item", "Q1", "Q2"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=1, value="B")
        ws.cell(row=3, column=2, value=2)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_multi_row_header_band():
    # Two stacked label rows (group label + sub-label) above numeric data.
    def build(ws):
        for ci, h in enumerate(["", "Revenue", "Revenue", "Cost"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        for ci, h in enumerate(["Region", "2019", "2020", "2020"], 1):
            ws.cell(row=2, column=ci, value=h).font = Font(bold=True)
        ws.cell(row=3, column=1, value="North")
        for ci in range(2, 5):
            ws.cell(row=3, column=ci, value=ci * 10)
        ws.cell(row=4, column=1, value="South")
        for ci in range(2, 5):
            ws.cell(row=4, column=ci, value=ci * 20)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 2)


def test_band_stops_at_data_row():
    def build(ws):
        for ci, h in enumerate(["Item", "V1", "V2"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        # row 2 is already data — band must not extend past row 1
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=2, column=3, value=2)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_band_stops_at_single_cell_divider():
    # A single-cell "NORTH" divider directly under the header must NOT be
    # absorbed into the header band (it's a section divider, not a header row).
    def build(ws):
        for ci, h in enumerate(["Item", "V1", "V2"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.cell(row=2, column=1, value="NORTH")  # single-cell divider
        ws.cell(row=3, column=1, value="Widget")
        ws.cell(row=3, column=2, value=1)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_single_row_header_above_text_data_stays_one_row():
    # Regression guard: a bold one-row header over a *text* first data row must
    # stay one row. The data row is unstyled, so the styled-band extension stops.
    def build(ws):
        for ci, h in enumerate(["Name", "City", "Role"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.cell(row=2, column=1, value="Alice")  # all-text data row, not bold
        ws.cell(row=2, column=2, value="Berlin")
        ws.cell(row=2, column=3, value="Eng")

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_styled_multi_row_header_band():
    # Two bold label rows over data → both rows are the header band.
    def build(ws):
        for ci, h in enumerate(["", "2019", "2020"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        for ci, h in enumerate(["Region", "Rev", "Rev"], 1):
            ws.cell(row=2, column=ci, value=h).font = Font(bold=True)
        ws.cell(row=3, column=1, value="North")
        ws.cell(row=3, column=2, value=10)
        ws.cell(row=3, column=3, value=20)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 2)


def test_no_header_for_pure_data_block():
    def build(ws):
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=r * c)

    assert _span(_sheet(build)) is None


# ── behaviours added with the DECO-tuned detector ───────────────────────────


def test_styled_title_rows_are_skipped_not_anchored():
    # A bold sparse title block above the real header (the classic "EOL Deals /
    # From: ..." preamble) must not become the header or be glued onto it.
    def build(ws):
        ws.cell(row=1, column=1, value="EOL Deals").font = Font(bold=True)
        ws.merge_cells("B1:C1")
        c = ws.cell(row=1, column=2, value="From: 5/1/2001 To: 5/8/2001")
        c.font = Font(bold=True)
        for ci, h in enumerate(["Desk", "Total MWH", "Total Deals"], 1):
            ws.cell(row=2, column=ci, value=h).font = Font(bold=True)
        for r in range(3, 6):
            ws.cell(row=r, column=1, value=f"Desk {r}")
            ws.cell(row=r, column=2, value=r * 100)
            ws.cell(row=r, column=3, value=r)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (2, 2)


def test_key_value_preamble_skipped():
    # "Label: value" preamble rows above the header are not anchors.
    def build(ws):
        ws.cell(row=1, column=1, value="Index").font = Font(bold=True)
        ws.cell(row=1, column=2, value="ANR/ML7-GDM").font = Font(bold=True)
        for ci, h in enumerate(
            ["Month", "Price", "Volume", "Total", "Net", "Fee", "Adj", "Gross", "Tax", "Final"], 1
        ):
            ws.cell(row=3, column=ci, value=h)
        for r in range(4, 9):
            ws.cell(row=r, column=1, value=f"2001-0{r - 3}-01")
            for ci in range(2, 11):
                ws.cell(row=r, column=ci, value=r * ci * 1.5)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (3, 3)


def test_unstyled_multi_row_header_extends():
    # An unstyled header found via contrast extends over a second label row
    # (the dominant multi-row pattern in DECO: deal-summary stacked labels).
    def build(ws):
        for ci, h in enumerate(["Trade", "Start", "End", "Price"], 1):
            ws.cell(row=1, column=ci, value=h)
        for ci, h in enumerate(["Number", "Date", "Date", "USD"], 1):
            ws.cell(row=2, column=ci, value=h)
        for r in range(3, 8):
            ws.cell(row=r, column=1, value=r * 7)
            ws.cell(row=r, column=2, value=f"2001-0{r - 2}-01")
            ws.cell(row=r, column=3, value=f"2001-0{r - 1}-01")
            ws.cell(row=r, column=4, value=r * 1.25)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 2)


def test_all_bold_table_header_stays_one_row():
    # Every cell in the table is bold: styling carries no signal, so the
    # body-signature brake must stop the band after the genuine header row.
    def build(ws):
        for ci, h in enumerate(["Book", "Desk", "Flag"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        rows = [
            ("FB-CENT1", "Central", "X"),
            ("FB-EAST2", "East", "X"),
            ("FB-WEST3", "West", "X"),
            ("FB-SOUTH4", "South", "X"),
        ]
        for ri, vals in enumerate(rows, 2):
            for ci, v in enumerate(vals, 1):
                ws.cell(row=ri, column=ci, value=v).font = Font(bold=True)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_vertical_merge_header_band():
    # Stacked header built from vertical merges: the merged rows belong to the
    # band even though their cells are empty merge slaves.
    def build(ws):
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.cell(row=1, column=1, value="Description").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Strip").font = Font(bold=True)
        ws.cell(row=1, column=3, value="Weighted").font = Font(bold=True)
        ws.cell(row=2, column=3, value="Average Price").font = Font(bold=True)
        for r in range(3, 7):
            ws.cell(row=r, column=1, value=f"Deal {r}")
            ws.cell(row=r, column=2, value=f"Apr0{r - 2}")
            ws.cell(row=r, column=3, value=r * 10.5)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 2)


def test_year_axis_row_is_header_not_data():
    # A row of year numbers over numeric data is a header (axis labels).
    def build(ws):
        ws.cell(row=1, column=1, value="Region").font = Font(bold=True)
        for ci, y in enumerate([2019, 2020, 2021], 2):
            ws.cell(row=1, column=ci, value=y).font = Font(bold=True)
        for r in range(2, 6):
            ws.cell(row=r, column=1, value=f"R{r}")
            for ci in range(2, 5):
                ws.cell(row=r, column=ci, value=r * ci * 3.7)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_date_axis_row_is_header_not_data():
    # Month-column headers (dates serialised as strings) are axis labels.
    import datetime

    def build(ws):
        ws.cell(row=1, column=1, value="Product").font = Font(bold=True)
        for ci, m in enumerate(range(1, 4), 2):
            ws.cell(row=1, column=ci, value=datetime.datetime(2001, m, 1)).font = Font(bold=True)
        for r in range(2, 6):
            ws.cell(row=r, column=1, value=f"P{r}")
            for ci in range(2, 5):
                ws.cell(row=r, column=ci, value=r * ci * 2.5)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_bold_totals_row_is_not_anchor():
    # A bold majority-numeric aggregate row at region top must not anchor the
    # band; the real header below it is used instead.
    def build(ws):
        ws.cell(row=1, column=1, value="Totals").font = Font(bold=True)
        for ci in range(2, 5):
            ws.cell(row=1, column=ci, value=ci * 1000).font = Font(bold=True)
        for ci, h in enumerate(["Book", "Jan", "Feb", "Mar"], 1):
            ws.cell(row=2, column=ci, value=h).font = Font(bold=True)
        for r in range(3, 7):
            ws.cell(row=r, column=1, value=f"B{r}")
            for ci in range(2, 5):
                ws.cell(row=r, column=ci, value=r * ci)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (2, 2)


def test_hidden_helper_row_not_in_band():
    # A hidden styled helper row above the real header is transparent.
    def build(ws):
        for ci, h in enumerate(["HelperA", "HelperB", "HelperC"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.row_dimensions[1].hidden = True
        for ci, h in enumerate(["Name", "Qty", "Price"], 1):
            ws.cell(row=2, column=ci, value=h).font = Font(bold=True)
        for r in range(3, 6):
            ws.cell(row=r, column=1, value=f"N{r}")
            ws.cell(row=r, column=2, value=r)
            ws.cell(row=r, column=3, value=r * 2.5)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (2, 2)


def test_dash_divider_breaks_band():
    # A punctuation-only row under the header is a rule-off, not a header row.
    def build(ws):
        for ci, h in enumerate(["Loc", "Pos", "Limit"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        for ci in range(1, 4):
            ws.cell(row=2, column=ci, value="---------").font = Font(bold=True)
        for r in range(3, 6):
            ws.cell(row=r, column=1, value=f"L{r}")
            ws.cell(row=r, column=2, value=r * 11)
            ws.cell(row=r, column=3, value=r * 100)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_label_value_pair_row_can_anchor():
    # GT convention: a 50/50 "label | value" top row with data beneath is a
    # header (strict-majority rejection keeps it eligible; styling decides).
    def build(ws):
        ws.cell(row=1, column=1, value="ESOP").font = Font(bold=True)
        ws.cell(row=1, column=2, value=50).font = Font(bold=True)
        for r in range(2, 6):
            ws.cell(row=r, column=1, value=r * 10.5)
            ws.cell(row=r, column=2, value=r * 20.5)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_merged_body_label_columns_stay_out_of_band():
    # Grouped tables often merge body label columns (A2:A6 "North"). Those
    # vertical merges hang from DATA rows, not the header — the band must
    # stay one row and never cascade down the merged stripe.
    def build(ws):
        for ci, h in enumerate(["Region", "Manager", "Q1"], 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
        ws.merge_cells("A2:A6")
        ws.merge_cells("B2:B6")
        ws.cell(row=2, column=1, value="North")
        ws.cell(row=2, column=2, value="Alice")
        for r in range(2, 7):
            ws.cell(row=r, column=3, value=r * 10)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (1, 1)


def test_date_valued_key_value_preamble_not_anchor():
    # "Date: | 2024-01-01"-style metadata pairs above a table: the lone date
    # value must not read as an axis label and anchor the header.
    import datetime

    def build(ws):
        ws.cell(row=1, column=1, value="Report")
        ws.cell(row=2, column=1, value="Date:")
        ws.cell(row=2, column=2, value=datetime.datetime(2024, 1, 1))
        ws.cell(row=3, column=1, value="Region:")
        ws.cell(row=3, column=2, value="EMEA")
        ws.cell(row=4, column=1, value="Owner:")
        ws.cell(row=4, column=2, value="JS")
        for ci, h in enumerate(["Product", "Units", "Price", "Total"], 1):
            ws.cell(row=6, column=ci, value=h).font = Font(bold=True)
        for r in range(7, 11):
            ws.cell(row=r, column=1, value=f"P{r}")
            for ci in range(2, 5):
                ws.cell(row=r, column=ci, value=r * ci * 1.1)

    span = _span(_sheet(build))
    assert span is not None and (span.top, span.bottom) == (6, 6)


def test_hidden_rows_cannot_stretch_band_past_cap():
    # Interleaved hidden rows are transparent but must not let the returned
    # span exceed the MAX_HEADER_ROWS bound from the anchor.
    def build(ws):
        ws.merge_cells("A1:A12")
        ws.merge_cells("B1:B12")
        ws.cell(row=1, column=1, value="Group").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Sub").font = Font(bold=True)
        ws.cell(row=1, column=3, value="Val").font = Font(bold=True)
        for r in [3, 5, 7, 9]:
            ws.row_dimensions[r].hidden = True
        for r in range(2, 14):
            ws.cell(row=r, column=3, value=r)

    span = _span(_sheet(build))
    assert span is not None
    from excel_parser.analysis.header_detector import MAX_HEADER_ROWS

    assert span.bottom - span.top + 1 <= MAX_HEADER_ROWS
