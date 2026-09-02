"""
Tests for workbook, sheet, and cell parsing.

Uses programmatically generated fixture workbooks to verify
correct extraction of values, formulas, styles, merges,
tables, comments, data validations, and sheet properties.
"""


import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from excel_parser.models.common import Severity
from excel_parser.parsers import WorkbookParser
from excel_parser.parsers.sheet_parser import SheetParser
from tests.helpers.invariant_checker import check_invariants


class TestSimpleWorkbook:
    """Test basic cell value and formula extraction."""

    def test_parse_cell_values(self, simple_workbook):
        parser = WorkbookParser(path=simple_workbook)
        result = parser.parse()

        assert result.total_sheets == 1
        assert result.total_cells > 0

        sheet = result.sheets[0]
        assert sheet.sheet_name == "Sheet1"

        # Check header cell
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.raw_value == "Name"
        assert a1.style is not None
        assert a1.style.font.bold is True

    def test_parse_formula(self, simple_workbook):
        parser = WorkbookParser(path=simple_workbook)
        result = parser.parse()

        sheet = result.sheets[0]
        b4 = sheet.get_cell(4, 2)  # B4 = =B2-B3
        assert b4 is not None
        assert b4.formula is not None
        assert "B2" in b4.formula and "B3" in b4.formula

    def test_parse_number_format(self, simple_workbook):
        parser = WorkbookParser(path=simple_workbook)
        result = parser.parse()

        sheet = result.sheets[0]
        b2 = sheet.get_cell(2, 2)  # Revenue = 1000
        assert b2 is not None
        assert b2.raw_value == 1000

    def test_workbook_hash_deterministic(self, simple_workbook):
        r1 = WorkbookParser(path=simple_workbook).parse()
        r2 = WorkbookParser(path=simple_workbook).parse()
        assert r1.workbook_hash == r2.workbook_hash

    def test_cell_ids_populated(self, simple_workbook):
        parser = WorkbookParser(path=simple_workbook)
        result = parser.parse()
        sheet = result.sheets[0]
        for cell in sheet.cells.values():
            assert cell.cell_id != ""
            assert cell.cell_hash != ""


class TestMergedCells:
    """Test merged cell handling."""

    def test_merge_regions_detected(self, merged_cells_workbook):
        result = WorkbookParser(path=merged_cells_workbook).parse()
        sheet = result.sheets[0]
        assert len(sheet.merged_regions) >= 2

    def test_merge_master_annotated(self, merged_cells_workbook):
        result = WorkbookParser(path=merged_cells_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)  # Master of A1:D1 merge
        assert a1 is not None
        assert a1.is_merged_master is True
        assert a1.merge_col_extent == 4

    def test_merge_slave_annotated(self, merged_cells_workbook):
        result = WorkbookParser(path=merged_cells_workbook).parse()
        sheet = result.sheets[0]
        b1 = sheet.get_cell(1, 2)  # Slave in A1:D1 merge
        assert b1 is not None
        assert b1.is_merged_slave is True


class TestStyledEmptyCells:
    """
    Formatting on a valueless cell must survive the parse.

    Two bugs met here. ``_has_meaningful_style`` enumerated three font
    attributes and so missed strikethrough, underline, size and name, while
    reporting *untouched* cells as styled (a partial ``Font`` leaves ``color``
    None; the inherited default font has one). And the gate that stores cells
    dropped anything empty regardless, which made the predicate inert — the
    styling was parsed and then discarded.
    """

    @pytest.mark.parametrize(
        ("coord", "attr", "expected"),
        [
            ((1, 2), "strikethrough", True),
            ((2, 2), "underline", "single"),
            ((3, 2), "bold", True),
            ((4, 2), "size", 20.0),
            ((5, 2), "name", "Courier New"),
        ],
    )
    def test_font_only_empty_cell_kept(
        self, styled_empty_cells_workbook, coord, attr, expected
    ):
        result = WorkbookParser(path=styled_empty_cells_workbook).parse()
        cell = result.sheets[0].get_cell(*coord)
        assert cell is not None, f"empty cell with {attr} was discarded"
        assert cell.style is not None and cell.style.font is not None
        assert getattr(cell.style.font, attr) == expected

    @pytest.mark.parametrize(
        ("coord", "attr"),
        [
            ((1, 3), "number_format"),
            ((2, 3), "alignment"),
            ((3, 3), "fill"),
            ((4, 3), "border"),
        ],
    )
    def test_non_font_only_empty_cell_kept(self, styled_empty_cells_workbook, coord, attr):
        result = WorkbookParser(path=styled_empty_cells_workbook).parse()
        cell = result.sheets[0].get_cell(*coord)
        assert cell is not None, f"empty cell with {attr} was discarded"
        assert cell.style is not None
        assert getattr(cell.style, attr) is not None

    def test_untouched_empty_cell_still_dropped(self, styled_empty_cells_workbook):
        """The fix must not start hoarding genuinely blank cells."""
        result = WorkbookParser(path=styled_empty_cells_workbook).parse()
        sheet = result.sheets[0]
        assert sheet.get_cell(6, 1) is None
        assert sheet.get_cell(20, 20) is None

    def test_styled_empty_cells_are_still_empty(self, styled_empty_cells_workbook):
        """Kept for their style, not misreported as carrying data."""
        result = WorkbookParser(path=styled_empty_cells_workbook).parse()
        cell = result.sheets[0].get_cell(1, 2)
        assert cell.is_empty is True
        assert cell.raw_value is None

    def test_unstyled_empty_merge_master_is_kept(self, tmp_dir):
        """
        A merge master that is empty *and* unstyled must survive the skip.

        It arrives as an ordinary Cell, not a MergedCell, so the empty-cell
        skip only spares it via the merge lookup. Dropping it strands every
        slave in the region with a merge_master that does not exist — which
        the old predicate hid by calling every untouched cell styled.
        """
        path = tmp_dir / "empty_unstyled_master.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "anchor"
        ws.merge_cells("B2:D2")  # master B2 left empty and unformatted
        wb.save(path)

        result = WorkbookParser(path=path).parse()
        sheet = result.sheets[0]
        master = sheet.get_cell(2, 2)
        assert master is not None, "empty unstyled merge master was discarded"
        assert master.is_merged_master is True
        assert check_invariants(result) == []

    def test_predicate_rejects_unstyled_cell(self, tmp_dir):
        """_has_meaningful_style must not call every default cell styled."""
        path = tmp_dir / "one_styled.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "anchor"
        ws["B1"].font = Font(strike=True)
        wb.save(path)

        loaded = openpyxl.load_workbook(path).active
        assert SheetParser._has_meaningful_style(loaded["B1"]) is True
        # A1 carries a value but no styling; C1 was never touched at all.
        assert SheetParser._has_meaningful_style(loaded["A1"]) is False
        assert SheetParser._has_meaningful_style(loaded["C1"]) is False


class TestOverlappingMerges:
    """
    Overlapping merged regions must not corrupt the merge invariants.

    A flat (row, col) -> master lookup let a later region overwrite an
    earlier one, so a cell in the intersection came out flagged both master
    and slave, or a slave pointing at a master that was never flagged.
    """

    def test_invariants_hold(self, overlapping_merges_workbook):
        result = WorkbookParser(path=overlapping_merges_workbook).parse()
        assert check_invariants(result) == []

    def test_no_cell_is_both_master_and_slave(self, overlapping_merges_workbook):
        result = WorkbookParser(path=overlapping_merges_workbook).parse()
        for sheet in result.sheets:
            for cell in sheet.cells.values():
                assert not (cell.is_merged_master and cell.is_merged_slave), (
                    f"{cell.a1_ref} is both master and slave"
                )

    def test_every_slave_points_at_a_flagged_master(self, overlapping_merges_workbook):
        result = WorkbookParser(path=overlapping_merges_workbook).parse()
        for sheet in result.sheets:
            for cell in sheet.cells.values():
                if not cell.is_merged_slave:
                    continue
                assert cell.merge_master is not None, f"{cell.a1_ref} slave with no master"
                master = sheet.get_cell(cell.merge_master.row, cell.merge_master.col)
                assert master is not None and master.is_merged_master

    def test_overlaps_are_dropped_not_silently_kept(self, overlapping_merges_workbook):
        result = WorkbookParser(path=overlapping_merges_workbook).parse()
        sheet = result.sheets[0]
        # Four regions declared, two of them overlapping an earlier one.
        assert len(sheet.merged_regions) == 2
        kept = {r.range.to_a1() for r in sheet.merged_regions}
        assert kept == {"B2:C5", "H2:J5"}

    def test_dropped_overlaps_are_reported(self, overlapping_merges_workbook):
        result = WorkbookParser(path=overlapping_merges_workbook).parse()
        sheet = result.sheets[0]
        warnings = [e for e in sheet.errors if "Overlapping merged regions" in e.message]
        assert len(warnings) == 2
        assert all(w.severity == Severity.WARNING for w in warnings)

    def test_resolution_is_reading_order_not_file_order(self, tmp_dir):
        """The surviving region is the upper-left one regardless of declaration order."""
        path = tmp_dir / "reversed_overlap.xlsx"
        wb = Workbook()
        ws = wb.active
        # Declared bottom-right first; B2:C5 must still be the survivor.
        for rng in ("C5:E8", "B2:C5"):
            ws.merge_cells(rng)
        wb.save(path)

        result = WorkbookParser(path=path).parse()
        kept = {r.range.to_a1() for r in result.sheets[0].merged_regions}
        assert kept == {"B2:C5"}

    def test_non_overlapping_merges_are_untouched(self, merged_cells_workbook):
        """The overlap pass must not perturb well-formed workbooks."""
        result = WorkbookParser(path=merged_cells_workbook).parse()
        sheet = result.sheets[0]
        assert not [e for e in sheet.errors if "Overlapping" in e.message]
        assert len(sheet.merged_regions) >= 2


class TestEmptyMasterRecovery:
    """Test OOXML recovery of values from empty merge masters."""

    def test_string_value_recovered(self, empty_master_workbook):
        result = WorkbookParser(path=empty_master_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None, "Master cell A1 should exist after recovery"
        assert a1.raw_value == "Recovered Text"
        assert a1.is_merged_master is True

    def test_number_value_recovered(self, empty_master_workbook):
        result = WorkbookParser(path=empty_master_workbook).parse()
        sheet = result.sheets[0]
        a2 = sheet.get_cell(2, 1)
        assert a2 is not None, "Master cell A2 should exist after recovery"
        assert a2.raw_value == 42
        assert a2.is_merged_master is True

    def test_existing_master_value_untouched(self, empty_master_workbook):
        result = WorkbookParser(path=empty_master_workbook).parse()
        sheet = result.sheets[0]
        a3 = sheet.get_cell(3, 1)
        assert a3 is not None
        assert a3.raw_value == "Master Has Value"
        assert a3.is_merged_master is True

    def test_recovery_from_bytes(self, empty_master_workbook):
        content = empty_master_workbook.read_bytes()
        result = WorkbookParser(content=content, filename="test.xlsx").parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.raw_value == "Recovered Text"


class TestFormulas:
    """Test formula extraction and cross-sheet references."""

    def test_cross_sheet_formulas(self, formula_workbook):
        result = WorkbookParser(path=formula_workbook).parse()
        assert result.total_sheets == 3

        calcs = result.sheets[1]
        b1 = calcs.get_cell(1, 2)
        assert b1 is not None
        assert b1.formula is not None
        assert "Inputs" in b1.formula

    def test_dependency_graph_built(self, formula_workbook):
        result = WorkbookParser(path=formula_workbook).parse()
        assert len(result.dependency_graph.edges) > 0

    def test_named_ranges_extracted(self, formula_workbook):
        result = WorkbookParser(path=formula_workbook).parse()
        names = {nr.name for nr in result.named_ranges}
        assert "Price" in names
        assert "Quantity" in names

    def test_array_formula_extracts_text_not_object_repr(self, array_formula_workbook):
        """ArrayFormula cells should yield formula string, not <openpyxl...object>."""
        result = WorkbookParser(path=array_formula_workbook).parse()
        sheet = result.sheets[0]
        b1 = sheet.get_cell(1, 2)
        assert b1 is not None
        assert b1.formula is not None
        assert "openpyxl" not in b1.formula
        assert "A1:A3" in b1.formula


class TestTables:
    """Test Excel ListObject table extraction."""

    def test_table_detected(self, table_workbook):
        result = WorkbookParser(path=table_workbook).parse()
        assert len(result.tables) == 1

    def test_table_properties(self, table_workbook):
        result = WorkbookParser(path=table_workbook).parse()
        table = result.tables[0]
        assert table.table_name == "SalesData"
        assert len(table.columns) == 7
        assert table.columns[0].name == "Product"

    def test_table_range(self, table_workbook):
        result = WorkbookParser(path=table_workbook).parse()
        table = result.tables[0]
        assert table.ref_range.to_a1() == "A1:G5"


class TestConditionalFormatting:
    """Test conditional formatting rule extraction."""

    def test_rules_extracted(self, conditional_format_workbook):
        result = WorkbookParser(path=conditional_format_workbook).parse()
        sheet = result.sheets[0]
        assert len(sheet.conditional_format_rules) > 0
        rule = sheet.conditional_format_rules[0]
        assert rule.rule_type == "cellIs"
        assert rule.operator == "greaterThan"


class TestDataValidation:
    """Test data validation extraction."""

    def test_validation_extracted(self, data_validation_workbook):
        result = WorkbookParser(path=data_validation_workbook).parse()
        sheet = result.sheets[0]
        assert len(sheet.data_validations) > 0
        dv = sheet.data_validations[0]
        assert dv.validation_type == "list"


class TestMultiSheet:
    """Test multi-sheet workbooks including hidden sheets."""

    def test_all_sheets_parsed(self, multi_sheet_workbook):
        result = WorkbookParser(path=multi_sheet_workbook).parse()
        assert result.total_sheets == 3

    def test_hidden_sheet_detected(self, multi_sheet_workbook):
        result = WorkbookParser(path=multi_sheet_workbook).parse()
        hidden = [s for s in result.sheets if s.properties.is_hidden]
        assert len(hidden) == 1
        assert hidden[0].sheet_name == "Hidden"


class TestHiddenRowsCols:
    """Test hidden row and column detection."""

    def test_hidden_row_detected(self, hidden_rows_cols_workbook):
        result = WorkbookParser(path=hidden_rows_cols_workbook).parse()
        sheet = result.sheets[0]
        assert 3 in sheet.hidden_rows

    def test_hidden_col_detected(self, hidden_rows_cols_workbook):
        result = WorkbookParser(path=hidden_rows_cols_workbook).parse()
        sheet = result.sheets[0]
        assert 2 in sheet.hidden_cols  # Column B = col 2


class TestComments:
    """Test cell comment extraction."""

    def test_comments_extracted(self, comment_workbook):
        result = WorkbookParser(path=comment_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.comment_text == "Total annual revenue"
        assert a1.comment_author == "Analyst"


class TestSparseSheet:
    """Test handling of large sparse sheets."""

    def test_sparse_cells_extracted(self, large_sparse_workbook):
        result = WorkbookParser(path=large_sparse_workbook).parse()
        sheet = result.sheets[0]
        assert sheet.cell_count() == 4  # A1, B1, Z100, CV1000
        assert sheet.get_cell(1000, 100) is not None

    def test_used_range_spans_sparse(self, large_sparse_workbook):
        result = WorkbookParser(path=large_sparse_workbook).parse()
        sheet = result.sheets[0]
        assert sheet.used_range is not None
        assert sheet.used_range.bottom_right.row == 1000


class TestFreezePane:
    """Test freeze pane detection."""

    def test_freeze_pane_extracted(self, freeze_panes_workbook):
        result = WorkbookParser(path=freeze_panes_workbook).parse()
        sheet = result.sheets[0]
        assert sheet.properties.freeze_pane == "A2"


class TestStyledWorkbook:
    """Test rich formatting extraction."""

    def test_font_color_extracted(self, styled_workbook):
        result = WorkbookParser(path=styled_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.style is not None
        assert a1.style.font.bold is True

    def test_fill_extracted(self, styled_workbook):
        result = WorkbookParser(path=styled_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1.style.fill is not None

    def test_border_extracted(self, styled_workbook):
        result = WorkbookParser(path=styled_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1.style.border is not None


class TestStrikethroughOnlyStyles:
    """Cells whose only non-default font attribute must keep their style."""

    def test_strikethrough_only_cell_keeps_style(self, strikethrough_workbook):
        result = WorkbookParser(path=strikethrough_workbook).parse()
        b1 = result.sheets[0].get_cell(1, 2)
        assert b1.style is not None
        assert b1.style.font is not None
        assert b1.style.font.strikethrough is True

    def test_underline_only_cell_keeps_style(self, strikethrough_workbook):
        result = WorkbookParser(path=strikethrough_workbook).parse()
        d1 = result.sheets[0].get_cell(1, 4)
        assert d1.style is not None
        assert d1.style.font is not None
        assert d1.style.font.underline == "single"

    def test_strikethrough_with_other_attributes(self, strikethrough_workbook):
        result = WorkbookParser(path=strikethrough_workbook).parse()
        c1 = result.sheets[0].get_cell(1, 3)
        assert c1.style.font.strikethrough is True
        assert c1.style.font.bold is True

    def test_unstruck_cell_reports_false(self, strikethrough_workbook):
        result = WorkbookParser(path=strikethrough_workbook).parse()
        a1 = result.sheets[0].get_cell(1, 1)
        assert a1.style.font.strikethrough is False


class TestWideSheet:
    """Test wide sheets with many columns."""

    def test_100_columns_parsed(self, wide_workbook):
        result = WorkbookParser(path=wide_workbook).parse()
        sheet = result.sheets[0]
        # Should have 100 header cells + 100*4 data cells = 500
        assert sheet.cell_count() == 500
        cell_100 = sheet.get_cell(1, 100)
        assert cell_100 is not None
        assert cell_100.raw_value == "Col100"


class TestHyperlinks:
    """Test hyperlink extraction."""

    def test_hyperlink_extracted(self, hyperlink_workbook):
        result = WorkbookParser(path=hyperlink_workbook).parse()
        sheet = result.sheets[0]
        a1 = sheet.get_cell(1, 1)
        assert a1 is not None
        assert a1.hyperlink == "https://www.google.com"
