"""Regression tests for cluster 01: array-formula cells.

Guards against `<openpyxl.worksheet.formula.ArrayFormula object at 0x…>`
leaking into chunk render_text. Two leak paths to cover:

1. `cell.value` is an `ArrayFormula` instance with a populated `.text`
   formula AND a cached computed value (the normal case).
2. The cached value is None (uncached array formula — cluster-03
   adjacent; we still must not emit the object repr).

Reference: docs/planning/recall-90/01-array-formula-rendering.md
"""
from __future__ import annotations

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

from ks_xlsx_parser.parsers.cell_parser import CellParser


def test_cell_parser_array_formula_with_cached_value_renders_value(tmp_path):
    p = tmp_path / "af_cached.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["A2"], ws["A3"] = 10, 20, 30
    af = ArrayFormula("B1:B3", "=A1:A3*2")
    ws["B1"] = af
    wb.save(p)

    # Re-read with data_only=False so the cell carries the ArrayFormula obj.
    wb_f = openpyxl.load_workbook(p, data_only=False)
    cell = wb_f.active["B1"]
    parser = CellParser("Sheet")
    dto = parser.parse(cell, computed_value=20)  # pretend the calc engine ran

    assert "ArrayFormula" not in str(dto.raw_value), (
        f"raw_value leaked object repr: {dto.raw_value!r}"
    )
    assert "ArrayFormula" not in (dto.display_value or ""), (
        f"display_value leaked object repr: {dto.display_value!r}"
    )
    assert dto.formula == "A1:A3*2", (
        "ArrayFormula source should be extracted from .text"
    )
    assert dto.raw_value == 20
    assert dto.data_type == "f"


def test_cell_parser_array_formula_without_cached_value_emits_formula_source(tmp_path):
    """Uncached array formula — the renderer must surface SOMETHING readable
    (the formula source) instead of the object repr."""
    p = tmp_path / "af_uncached.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B1"] = ArrayFormula("B1:B3", "=SUM(A1:A3)")
    wb.save(p)

    wb_f = openpyxl.load_workbook(p, data_only=False)
    cell = wb_f.active["B1"]
    dto = CellParser("Sheet").parse(cell, computed_value=None)

    text_blob = " ".join(filter(None, [
        str(dto.raw_value) if dto.raw_value is not None else "",
        dto.display_value or "",
        dto.formula or "",
    ]))
    assert "ArrayFormula" not in text_blob, (
        f"object repr leaked through CellDTO: {text_blob!r}"
    )
    # The formula source string is the acceptable retrieval surface.
    assert dto.formula and "SUM(A1:A3)" in dto.formula


def test_pipeline_array_formula_does_not_leak_object_repr(tmp_path):
    """End-to-end: chunk render_text from a workbook with array formulas
    must not contain 'ArrayFormula' anywhere."""
    from ks_xlsx_parser.api import parse_workbook

    p = tmp_path / "af_pipeline.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, val in enumerate([10, 20, 30, 40, 50], start=1):
        ws.cell(row=i, column=1, value=val)
    ws["B1"] = ArrayFormula("B1:B5", "=A1:A5*2")
    wb.save(p)

    result = parse_workbook(path=str(p))
    for chunk in result.chunks:
        assert "ArrayFormula" not in (chunk.render_text or ""), (
            f"ArrayFormula repr leaked in chunk render_text:\n"
            f"{chunk.render_text}"
        )
