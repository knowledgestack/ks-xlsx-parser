"""
Tests for deterministic section detection (outline levels + merged dividers)
and section-context propagation into windowed chunks.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from excel_parser.analysis.header_detector import find_header_span
from excel_parser.analysis.section_detector import find_sections, section_path_at
from excel_parser.models.common import CellCoord, CellRange
from excel_parser.parsers.workbook_parser import WorkbookParser
from excel_parser.pipeline import parse_workbook


def _sheet(build):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    build(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return WorkbookParser(content=buf.getvalue(), filename="x.xlsx").parse().sheets[0]


def _full_range(sheet):
    used = sheet.compute_used_range()
    return used


# ── outline-level (nested) sections ────────────────────────────────────────


def _nested(ws):
    for ci, h in enumerate(["Label", "Q1", "Q2"], 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ws.cell(row=2, column=1, value="NORTH")
    ws.cell(row=3, column=1, value="Coastal")
    ws.cell(row=4, column=1, value="Widget")
    ws.cell(row=4, column=2, value=1)
    ws.cell(row=5, column=1, value="SOUTH")
    ws.cell(row=6, column=1, value="Bolt")
    ws.cell(row=6, column=2, value=2)
    ws.row_dimensions[3].outline_level = 1
    ws.row_dimensions[4].outline_level = 2
    ws.row_dimensions[6].outline_level = 1


def test_outline_levels_captured():
    sheet = _sheet(_nested)
    assert sheet.row_outline_levels == {3: 1, 4: 2, 6: 1}


def test_nested_sections_and_path():
    sheet = _sheet(_nested)
    rng = _full_range(sheet)
    secs = find_sections(sheet, rng, find_header_span(sheet, rng))
    labels = {(s.label, s.level) for s in secs}
    assert ("NORTH", 0) in labels
    assert ("Coastal", 1) in labels
    # Deepest data row resolves to its full ancestor chain.
    path = [s.label for s in section_path_at(secs, 4)]
    assert path == ["NORTH", "Coastal"]


# ── merged-row dividers (flat) ─────────────────────────────────────────────


def _merged(ws):
    for ci, h in enumerate(["Region", "Q1", "Q2", "Q3"], 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ws.merge_cells("A2:D2")
    ws["A2"] = "NORTH REGION"
    ws.cell(row=3, column=1, value="Widget")
    ws.cell(row=3, column=2, value=1)
    ws.merge_cells("A4:D4")
    ws["A4"] = "SOUTH REGION"
    ws.cell(row=5, column=1, value="Gadget")
    ws.cell(row=5, column=2, value=2)


def test_merged_divider_sections():
    sheet = _sheet(_merged)
    rng = _full_range(sheet)
    secs = find_sections(sheet, rng, find_header_span(sheet, rng))
    assert [(s.label, s.top_row, s.bottom_row) for s in secs] == [
        ("NORTH REGION", 2, 3),
        ("SOUTH REGION", 4, 5),
    ]


# ── negative: a blank data row is NOT a section (locks user requirement #3) ──


def _blank_data_row(ws):
    for ci, h in enumerate(["Item", "V1", "V2"], 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ws.cell(row=2, column=1, value="A")
    ws.cell(row=2, column=2, value=1)
    ws.cell(row=3, column=1, value="B")  # all value columns legitimately blank
    ws.cell(row=4, column=1, value="C")
    ws.cell(row=4, column=2, value=3)


def test_blank_data_row_is_not_a_section():
    sheet = _sheet(_blank_data_row)
    rng = _full_range(sheet)
    assert find_sections(sheet, rng, find_header_span(sheet, rng)) == []


# ── end-to-end: metadata + windowed section context ────────────────────────


def test_sections_in_chunk_metadata():
    wb = Workbook()
    ws = wb.active
    _merged(ws)
    buf = io.BytesIO()
    wb.save(buf)
    chunk = parse_workbook(content=buf.getvalue(), filename="x.xlsx").chunks[0]
    sections = chunk.metadata.get("sections")
    assert sections and {s["label"] for s in sections} == {
        "NORTH REGION",
        "SOUTH REGION",
    }


def test_windowed_part_carries_section_path():
    """A large nested table: a part starting mid-section names its ancestors in
    both render text and metadata['section_path']."""
    wb = Workbook()
    ws = wb.active
    for ci, h in enumerate(["Label", "V1", "V2", "V3"], 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    r = 2
    for region in ["NORTH", "SOUTH", "EAST"]:
        ws.cell(row=r, column=1, value=region)
        r += 1
        for sub in ["Coastal", "Inland"]:
            ws.cell(row=r, column=1, value=sub)
            ws.row_dimensions[r].outline_level = 1
            r += 1
            for k in range(40):
                ws.cell(row=r, column=1, value=f"{region[0]}{sub[0]}{k}")
                for col in (2, 3, 4):
                    ws.cell(row=r, column=col, value=k * col)
                ws.row_dimensions[r].outline_level = 2
                r += 1
    buf = io.BytesIO()
    wb.save(buf)
    res = parse_workbook(content=buf.getvalue(), filename="y.xlsx")
    parts = [c for c in res.chunks if c.metadata.get("table_part")]
    assert len(parts) >= 2
    # At least one part starts mid-section and carries an injected ancestor path.
    mid = [c for c in parts if c.metadata.get("section_path")]
    assert mid, "no part carried a section_path"
    c = mid[0]
    labels = [s["label"] for s in c.metadata["section_path"]]
    assert labels  # non-empty ancestor chain
    assert any(ln.startswith("section:") for ln in c.render_text.splitlines())
