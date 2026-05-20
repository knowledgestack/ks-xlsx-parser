"""Range-tightening invariant for chunk emission.

Cluster-02 closes when every chunk's claimed (top_left, bottom_right)
range is the bounding box of cells with non-empty content within
``block.cell_range``. Today the chunker copies ``block.cell_range``
verbatim — when the segmenter hands it a block whose range is wider
than the data (empty rows/cols at the edges), the chunk overclaims
geometric coverage.

These tests assert the invariant on real fixtures + a hand-built
"data in the upper-left corner of a wider block" workbook.
"""
from __future__ import annotations

import openpyxl

from ks_xlsx_parser.api import parse_workbook
from ks_xlsx_parser.models.common import addr_to_a1


def _a1_to_coord(a1: str) -> tuple[int, int]:
    """Inverse of addr_to_a1 — small helper for tests only."""
    letters = "".join(c for c in a1 if c.isalpha())
    digits = "".join(c for c in a1 if c.isdigit())
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(digits), col


def test_chunk_range_tight_around_actual_content(tmp_path):
    """A block whose range covers A1:E20 but only has data in A1:C5
    must emit a chunk that claims A1:C5 (or tighter), not A1:E20."""
    p = tmp_path / "sparse.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    # Fill only the upper-left 5×3 region; rest of the block is empty.
    ws["A1"] = "Name"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    for i, name in enumerate(["Alpha", "Bravo", "Charlie", "Delta"], start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=i * 10)
        ws.cell(row=i, column=3, value=i * 20)
    wb.save(p)

    result = parse_workbook(path=str(p))
    chunks_on_S = [c for c in result.chunks if c.sheet_name == "S"]
    assert chunks_on_S, "no chunks emitted on S"
    for c in chunks_on_S:
        r1, c1 = _a1_to_coord(c.bottom_right_cell)
        # Bottom-right of the claimed range must not exceed the data bbox.
        assert r1 <= 5, (
            f"chunk over-claims rows: {c.top_left_cell}:{c.bottom_right_cell} "
            f"but data ends at row 5"
        )
        assert c1 <= 3, (
            f"chunk over-claims cols: {c.top_left_cell}:{c.bottom_right_cell} "
            f"but data ends at col C"
        )


def test_chunk_range_unchanged_when_block_is_already_tight(tmp_path):
    """If block.cell_range already matches the data bbox, the chunk's
    claimed range must be identical — no spurious narrowing."""
    p = tmp_path / "dense.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "h1"
    ws["B1"] = "h2"
    ws["A2"] = 1
    ws["B2"] = 2
    wb.save(p)

    result = parse_workbook(path=str(p))
    chunks_on_S = [c for c in result.chunks if c.sheet_name == "S"]
    assert chunks_on_S
    # The single dense block should claim A1:B2 exactly.
    found_tight = any(
        c.top_left_cell == "A1" and c.bottom_right_cell == "B2"
        for c in chunks_on_S
    )
    assert found_tight, [
        (c.top_left_cell, c.bottom_right_cell) for c in chunks_on_S
    ]


def test_chunk_range_invariant_holds_on_corpus_fixture():
    """End-to-end on a real SpreadsheetBench input — every emitted
    chunk's range must be a tight bbox over cells with content.
    Uses one of the 50-sample sheets; skipped if corpus absent."""
    import os
    p = (
        "data/corpora/spreadsheetbench/all_data_912_v0.1"
        "/spreadsheet/54105/1_54105_input.xlsx"
    )
    if not os.path.exists(p):
        import pytest
        pytest.skip(f"corpus fixture not present: {p}")

    result = parse_workbook(path=p)
    import openpyxl
    wb = openpyxl.load_workbook(p, data_only=True)

    for c in result.chunks:
        if not c.top_left_cell or not c.bottom_right_cell:
            continue
        ws = wb[c.sheet_name]
        r0, c0 = _a1_to_coord(c.top_left_cell)
        r1, c1 = _a1_to_coord(c.bottom_right_cell)
        # The CORNER rows/cols of the claimed range must each contain
        # some non-empty cell — otherwise the range is over-claiming.
        def _row_has_content(r: int) -> bool:
            for col in range(c0, c1 + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None and str(v).strip():
                    return True
            return False

        def _col_has_content(col: int) -> bool:
            for r in range(r0, r1 + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None and str(v).strip():
                    return True
            return False

        # Allow chart anchors (single-cell, no real content) to pass.
        if r0 == r1 and c0 == c1:
            continue
        assert _row_has_content(r0), (
            f"top row of chunk {c.top_left_cell}:{c.bottom_right_cell} is empty"
        )
        assert _row_has_content(r1), (
            f"bottom row of chunk {c.top_left_cell}:{c.bottom_right_cell} is empty"
        )
        assert _col_has_content(c0), (
            f"left col of chunk {c.top_left_cell}:{c.bottom_right_cell} is empty"
        )
        assert _col_has_content(c1), (
            f"right col of chunk {c.top_left_cell}:{c.bottom_right_cell} is empty"
        )
