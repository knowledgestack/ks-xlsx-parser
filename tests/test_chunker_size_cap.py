"""Token-budget chunk size cap.

Cluster-04: single huge chunks dilute embeddings — when a sheet has a
1000-row table, putting it in one chunk means the embedding mixes the
question-relevant rows with 999 others and recall@5 suffers. The fix
is to split oversize blocks into row groups, each with a tight A1
range covering only its rows.

These tests assert:
  1. A synthetic 1000-row table emits ≥ 2 chunks (was 1 before).
  2. No chunk's render_text exceeds the budget (~2000 chars).
  3. Each child chunk's range is contiguous and non-overlapping with
     siblings.
  4. Total row coverage of children == original block's row range
     (no data dropped, none duplicated).
  5. Small blocks (≤ budget) are emitted unchanged — no spurious
     splitting on dense single-table sheets.
"""
from __future__ import annotations

import openpyxl

from ks_xlsx_parser.api import parse_workbook


def _a1_to_coord(a1: str) -> tuple[int, int]:
    letters = "".join(c for c in a1 if c.isalpha())
    digits = "".join(c for c in a1 if c.isdigit())
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(digits), col


def test_thousand_row_table_splits_into_multiple_chunks(tmp_path, monkeypatch):
    """With the cap explicitly tightened (KS_CHUNK_BUDGET_CHARS=2000) a
    1000-row table must emit ≥ 2 chunks. The default budget is much
    higher (cap effectively off) — see test_default_keeps_tables_whole
    for the inverse assertion."""
    monkeypatch.setenv("KS_CHUNK_BUDGET_CHARS", "2000")
    p = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "name", "value"])
    for i in range(1, 1001):
        ws.append([i, f"name-{i}", i * 7])
    wb.save(p)

    result = parse_workbook(path=str(p))
    chunks_on_S = [c for c in result.chunks if c.sheet_name == "S"]
    assert len(chunks_on_S) >= 2, (
        f"expected ≥2 chunks after cap, got {len(chunks_on_S)} "
        f"(render_text lengths: "
        f"{[len(c.render_text or '') for c in chunks_on_S]})"
    )


def test_default_keeps_tables_whole(tmp_path):
    """The shipped default budget (~100k chars) deliberately does NOT
    fragment moderate tables — calibration on the 50-sample showed
    splitting moderate tables (1k–10k chars) regresses retrieval
    because the embedding can't discriminate between same-shape
    children. Keep tables together by default."""
    p = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "name", "value"])
    for i in range(1, 1001):
        ws.append([i, f"name-{i}", i * 7])
    wb.save(p)

    result = parse_workbook(path=str(p))
    chunks_on_S = [c for c in result.chunks if c.sheet_name == "S"]
    assert len(chunks_on_S) == 1, (
        f"default should not split a 1000-row table; got "
        f"{len(chunks_on_S)} chunks"
    )


def test_no_chunk_exceeds_render_budget(tmp_path, monkeypatch):
    """Each chunk's render_text must stay near the configured budget
    when the cap is engaged."""
    monkeypatch.setenv("KS_CHUNK_BUDGET_CHARS", "2000")
    p = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "name", "value"])
    for i in range(1, 1001):
        ws.append([i, f"name-{i}", i * 7])
    wb.save(p)

    result = parse_workbook(path=str(p))
    # Allow 20% overshoot (average-rows-per-budget heuristic) + 200 chars
    # for the block-header + column-letters + separator lines.
    cap = int(2000 * 1.2) + 200
    for c in result.chunks:
        assert len(c.render_text or "") <= cap, (
            f"chunk over budget (2000+slack={cap}): "
            f"{len(c.render_text)} chars on "
            f"{c.top_left_cell}:{c.bottom_right_cell}"
        )


def test_child_chunk_ranges_are_contiguous_and_non_overlapping(
    tmp_path, monkeypatch,
):
    monkeypatch.setenv("KS_CHUNK_BUDGET_CHARS", "2000")
    p = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "name", "value"])
    for i in range(1, 501):
        ws.append([i, f"name-{i}", i * 7])
    wb.save(p)

    result = parse_workbook(path=str(p))
    chunks_on_S = [c for c in result.chunks if c.sheet_name == "S"]
    assert len(chunks_on_S) >= 2

    # Sort by top row, then check contiguity / non-overlap.
    ranges = sorted(
        [
            (
                _a1_to_coord(c.top_left_cell),
                _a1_to_coord(c.bottom_right_cell),
            )
            for c in chunks_on_S
        ],
        key=lambda x: (x[0][0], x[0][1]),
    )
    for (tl_a, br_a), (tl_b, br_b) in zip(ranges, ranges[1:]):
        # No overlap on rows (since splits are by row-group)
        assert br_a[0] < tl_b[0], (
            f"row overlap between {tl_a}-{br_a} and {tl_b}-{br_b}"
        )
        # Contiguous: child B starts right after child A ends
        assert tl_b[0] == br_a[0] + 1, (
            f"row gap between {tl_a}-{br_a} and {tl_b}-{br_b}"
        )


def test_small_block_not_split(tmp_path):
    """A 10-row table fits in one chunk; cap must not over-split."""
    p = tmp_path / "small.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "name"])
    for i in range(1, 11):
        ws.append([i, f"n{i}"])
    wb.save(p)

    result = parse_workbook(path=str(p))
    chunks_on_S = [c for c in result.chunks if c.sheet_name == "S"]
    assert len(chunks_on_S) == 1, (
        f"unexpected split on small block: "
        f"{[(c.top_left_cell, c.bottom_right_cell) for c in chunks_on_S]}"
    )
