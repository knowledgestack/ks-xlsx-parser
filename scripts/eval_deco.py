"""
DECO structural benchmark — table-boundary + header-row detection, ks vs docling.

DECO (Dresden Enron COrpus) ships 852 real spreadsheets where annotators marked,
for every sheet, the *table regions* and, inside each table, the *header* / *data*
/ *derived* row ranges. Those annotations live in a hidden worksheet named
``Range_Annotations_Data`` with columns:

    Sheet.Name, Sheet.Index, Annotation.Label, Annotation.Name,
    Annotation.Range, Annotation.Parent, ...

A ``Table`` row's parent is the worksheet; a ``Header``/``Data``/... row's parent
is the ``Annotation.Name`` of its table. That gives us ground truth for two things
SpreadsheetBench cannot score:

  1. Table-boundary detection  — where does each table start/end on the sheet.
  2. Header-row detection       — which row(s) are the column header of a table.

We score each parser against that GT:

  * ks      — full: table-region IoU, fragmentation, plus header-row P/R/F1 using
              the SHIPPED ``find_header_span`` (the single source of truth the
              renderer + segmenter use).
  * docling — xlsx → markdown collapses to ~one table per sheet and exposes no A1
              coordinates, so it cannot be scored on localisation. We score the one
              axis it *can* be measured on: tables-detected-per-sheet vs GT count
              (its multi-table "collapse" rate).

Run one parser per process (memory isolation), then a report pass:

    PYTHONPATH=src uv run python scripts/eval_deco.py --parser ks      --out RUN
    PYTHONPATH=src uv run python scripts/eval_deco.py --parser docling --out RUN
    PYTHONPATH=src uv run python scripts/eval_deco.py --report         --out RUN
"""
from __future__ import annotations

import argparse
import gc
import glob
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ───────────────────────────────────────────────────────────── A1 helpers
_A1 = re.compile(r"\$?([A-Za-z]+)\$?(\d+)")


def _col_to_num(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def parse_a1_range(spec: str) -> tuple[int, int, int, int] | None:
    """'$A$5:$S$44' or 'A5' → (r0, c0, r1, c1), 1-based inclusive. None if bad."""
    cells = _A1.findall(spec or "")
    if not cells:
        return None
    (c_lo, r_lo) = cells[0]
    (c_hi, r_hi) = cells[-1]
    r0, c0 = int(r_lo), _col_to_num(c_lo)
    r1, c1 = int(r_hi), _col_to_num(c_hi)
    return (min(r0, r1), min(c0, c1), max(r0, r1), max(c0, c1))


def _area(b: tuple[int, int, int, int]) -> int:
    return (b[2] - b[0] + 1) * (b[3] - b[1] + 1)


def _inter(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> int:
    r0, c0 = max(a[0], b[0]), max(a[1], b[1])
    r1, c1 = min(a[2], b[2]), min(a[3], b[3])
    if r1 < r0 or c1 < c0:
        return 0
    return (r1 - r0 + 1) * (c1 - c0 + 1)


def iou(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> float:
    i = _inter(a, b)
    if i == 0:
        return 0.0
    return i / (_area(a) + _area(b) - i)


def rows_of(span: tuple[int, int, int, int]) -> set[int]:
    return set(range(span[0], span[2] + 1))


# ───────────────────────────────────────────────────────────── GT loader
ANNOTATION_SHEETS = {"Range_Annotations_Data", "Annotation_Status_Data"}


def load_gt(path: str) -> dict[str, list[dict]]:
    """Return {sheet_name: [ {range, headers:[ranges], name}, ... ]}."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    if "Range_Annotations_Data" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Range_Annotations_Data"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    # tables keyed by their Annotation.Name; children attach by parent name
    tables: dict[str, dict] = {}
    children: list[tuple[str, str, tuple]] = []  # (label, parent, range)
    for r in rows[1:]:
        if not r or len(r) < 6 or not r[2]:
            continue
        sheet_name, _idx, label, name, rng, parent = (
            r[0], r[1], r[2], r[3], r[4], r[5],
        )
        box = parse_a1_range(rng or "")
        if box is None:
            continue
        if label == "Table":
            tables[name] = {"sheet": sheet_name, "range": box, "headers": [], "name": name}
        else:
            children.append((label, parent, box))
    for label, parent, box in children:
        if label == "Header" and parent in tables:
            tables[parent]["headers"].append(box)
    out: dict[str, list[dict]] = defaultdict(list)
    for t in tables.values():
        out[t["sheet"]].append(t)
    return out


# ───────────────────────────────────────────────────────────── ks scorer
def score_ks(path: str, gt: dict[str, list[dict]]) -> dict:
    from excel_parser.analysis.header_detector import find_header_span
    from excel_parser.models.common import CellCoord, CellRange
    from excel_parser.pipeline import parse_workbook

    result = parse_workbook(path)
    sheets = {s.sheet_name: s for s in result.workbook.sheets}
    # ks candidate table regions per sheet (chunks that carry a cell_range)
    regions: dict[str, list[tuple[int, int, int, int]]] = defaultdict(list)
    for ch in result.chunks:
        if ch.cell_range is None:
            continue
        cr = ch.cell_range
        regions[ch.sheet_name].append(
            (cr.top_left.row, cr.top_left.col, cr.bottom_right.row, cr.bottom_right.col)
        )

    def header_rows_for(span: tuple[int, int, int, int], sheet_name: str) -> set[int]:
        sheet = sheets.get(sheet_name)
        if sheet is None:
            return set()
        cr = CellRange(
            top_left=CellCoord(row=span[0], col=span[1]),
            bottom_right=CellCoord(row=span[2], col=span[3]),
        )
        hs = find_header_span(sheet, cr)
        return set() if hs is None else set(range(hs.top, hs.bottom + 1))

    tbl_records = []
    for sheet_name, gtables in gt.items():
        cand = regions.get(sheet_name, [])
        for t in gtables:
            gbox = t["range"]
            best_iou = max((iou(gbox, c) for c in cand), default=0.0)
            best_region = max(cand, key=lambda c: iou(gbox, c), default=None) if cand else None
            frags = sum(1 for c in cand if _inter(gbox, c) > 0)
            gt_hrows = set()
            for h in t["headers"]:
                gt_hrows |= rows_of(h)
            # end-to-end: header span ks computes on its best-matched region
            ks_h_e2e = header_rows_for(best_region, sheet_name) if best_region else set()
            # isolated: header detector accuracy on the GT region itself
            ks_h_iso = header_rows_for(gbox, sheet_name)
            tbl_records.append({
                "sheet": sheet_name,
                "gt_range": gbox,
                "best_iou": round(best_iou, 4),
                "frags": frags,
                "has_gt_header": bool(gt_hrows),
                "gt_header_multirow": bool(gt_hrows) and (len({r for r in gt_hrows}) > 1),
                "gt_hrows": sorted(gt_hrows),
                "ks_hrows_e2e": sorted(ks_h_e2e),
                "ks_hrows_iso": sorted(ks_h_iso),
            })
    del result
    gc.collect()
    return {"file": os.path.basename(path), "parser": "ks", "tables": tbl_records}


# ───────────────────────────────────────────────────────────── docling scorer
_DOC_CONV = None


def _docling_converter():
    global _DOC_CONV
    if _DOC_CONV is None:
        from docling.document_converter import DocumentConverter

        _DOC_CONV = DocumentConverter()
    return _DOC_CONV


def score_docling(path: str, gt: dict[str, list[dict]]) -> dict:
    conv = _docling_converter()
    real_sheets = set(gt.keys())  # GT sheets are the real (non-annotation) sheets
    try:
        doc = conv.convert(path).document
    except Exception as exc:  # noqa: BLE001
        return {"file": os.path.basename(path), "parser": "docling",
                "error": f"{type(exc).__name__}: {exc}", "per_sheet": []}
    # map tables → sheet name via provenance page_no (1 page per sheet, in order)
    page_names = {}
    try:
        for pno, page in (doc.pages or {}).items():
            page_names[pno] = getattr(page, "name", None)
    except Exception:
        page_names = {}
    # fallback: page order maps to workbook sheet order
    sheet_order = list(_sheet_order(path))
    docling_counts: dict[str, int] = defaultdict(int)
    for t in doc.tables:
        sname = None
        try:
            prov = t.prov[0] if t.prov else None
            if prov is not None:
                pno = prov.page_no
                sname = page_names.get(pno)
                if sname is None and 1 <= pno <= len(sheet_order):
                    sname = sheet_order[pno - 1]
        except Exception:
            sname = None
        if sname is None:
            sname = "__unknown__"
        docling_counts[sname] += 1
    per_sheet = []
    for sheet_name, gtables in gt.items():
        per_sheet.append({
            "sheet": sheet_name,
            "gt_tables": len(gtables),
            "docling_tables": docling_counts.get(sheet_name, 0),
        })
    # if mapping failed entirely, also record total docling tables on real sheets
    total_docling = sum(v for k, v in docling_counts.items()
                        if k in real_sheets or k == "__unknown__")
    del doc
    gc.collect()
    return {"file": os.path.basename(path), "parser": "docling",
            "per_sheet": per_sheet, "total_docling_tables_unmapped": total_docling}


def _sheet_order(path: str) -> list[str]:
    # Full workbook sheet order (incl. annotation sheets) so docling's 1-based
    # page_no aligns exactly with the sheet at that position. Annotation sheets
    # are dropped at scoring time because they never appear in the GT.
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


# ───────────────────────────────────────────────────────────── runner
def iter_files(corpus: str, sample: int | None) -> list[str]:
    files = sorted(glob.glob(os.path.join(corpus, "*.xlsx")))
    if sample:
        # deterministic stride sample for reproducibility
        step = max(1, len(files) // sample)
        files = files[::step][:sample]
    return files


class _Timeout(Exception):
    pass


def run_parser(parser: str, files: list[str], out_dir: Path, timeout: int = 60) -> None:
    import signal

    def _on_alarm(signum, frame):
        raise _Timeout()

    signal.signal(signal.SIGALRM, _on_alarm)

    out_path = out_dir / f"{parser}.ndjson"
    scorer = score_ks if parser == "ks" else score_docling
    n = timeouts = errors = 0
    with out_path.open("w") as fh:
        for path in files:
            gt = load_gt(path)
            if not gt:
                continue
            signal.alarm(timeout)
            try:
                rec = scorer(path, gt)
            except _Timeout:
                timeouts += 1
                rec = {"file": os.path.basename(path), "parser": parser,
                       "error": f"timeout>{timeout}s"}
            except Exception as exc:  # noqa: BLE001
                errors += 1
                rec = {"file": os.path.basename(path), "parser": parser,
                       "error": f"{type(exc).__name__}: {exc}"}
            finally:
                signal.alarm(0)
            fh.write(json.dumps(rec) + "\n")
            fh.flush()
            n += 1
            if n % 25 == 0:
                print(f"  {parser}: {n}/{len(files)} scored "
                      f"({timeouts} timeout, {errors} err)", file=sys.stderr, flush=True)
    print(f"✓ {parser}: wrote {n} records ({timeouts} timeout, {errors} err) → {out_path}",
          file=sys.stderr)


# ───────────────────────────────────────────────────────────── report
def _prf(pred: set[int], gt: set[int]) -> tuple[int, int, int]:
    """Return (true_pos, pred_size, gt_size)."""
    return (len(pred & gt), len(pred), len(gt))


def report(out_dir: Path) -> None:
    ks_recs = _read_ndjson(out_dir / "ks.ndjson")
    doc_recs = _read_ndjson(out_dir / "docling.ndjson")

    # ── ks table + header aggregates
    n_tables = 0
    iou_sum = 0.0
    det50 = 0
    det30 = 0
    frag_sum = 0
    frag_vals: list[int] = []
    over_seg = 0          # GT table split across >1 ks region
    # header (only tables with a GT header)
    h_tables = 0
    h_multirow = 0
    tp_e2e = pred_e2e = gtsz_e2e = 0
    tp_iso = pred_iso = gtsz_iso = 0
    h_detected_e2e = 0    # ks emitted a non-empty header
    h_exact_iso = 0
    ks_timeouts = sum(1 for r in ks_recs if "timeout" in (r.get("error") or ""))
    ks_errors = sum(1 for r in ks_recs if r.get("error") and "timeout" not in r["error"])
    # per-file macro accumulators (dampen any single huge file, e.g. the 130-table UAMPS sheet)
    macro_iou: list[float] = []
    macro_hf1: list[float] = []
    for rec in ks_recs:
        f_iou: list[float] = []
        f_tp = f_pred = f_gt = 0
        for t in rec.get("tables", []):
            f_iou.append(t["best_iou"])
            if t["has_gt_header"]:
                e = set(t["ks_hrows_e2e"])
                g = set(t["gt_hrows"])
                f_tp += len(e & g)
                f_pred += len(e)
                f_gt += len(g)
        if f_iou:
            macro_iou.append(sum(f_iou) / len(f_iou))
            p = f_tp / f_pred if f_pred else 0.0
            r = f_tp / f_gt if f_gt else 0.0
            macro_hf1.append(2 * p * r / (p + r) if (p + r) else 0.0)
    for rec in ks_recs:
        for t in rec.get("tables", []):
            n_tables += 1
            iou_sum += t["best_iou"]
            det50 += t["best_iou"] >= 0.5
            det30 += t["best_iou"] >= 0.3
            frag_sum += t["frags"]
            frag_vals.append(t["frags"])
            over_seg += t["frags"] > 1
            if t["has_gt_header"]:
                h_tables += 1
                gt_h = set(t["gt_hrows"])
                if len(gt_h) > 1:
                    h_multirow += 1
                e2e = set(t["ks_hrows_e2e"])
                iso = set(t["ks_hrows_iso"])
                a, b, c = _prf(e2e, gt_h)
                tp_e2e += a
                pred_e2e += b
                gtsz_e2e += c
                h_detected_e2e += bool(e2e)
                a, b, c = _prf(iso, gt_h)
                tp_iso += a
                pred_iso += b
                gtsz_iso += c
                h_exact_iso += (iso == gt_h)

    def f1(tp, pred, gt):
        p = tp / pred if pred else 0.0
        r = tp / gt if gt else 0.0
        f = 2 * p * r / (p + r) if (p + r) else 0.0
        return p, r, f

    p_e2e, r_e2e, f_e2e = f1(tp_e2e, pred_e2e, gtsz_e2e)
    p_iso, r_iso, f_iso = f1(tp_iso, pred_iso, gtsz_iso)

    def _median(xs: list) -> float:
        if not xs:
            return 0.0
        s = sorted(xs)
        m = len(s) // 2
        return float(s[m]) if len(s) % 2 else (s[m - 1] + s[m]) / 2

    # ── docling tables-per-sheet
    d_sheets = 0
    d_abs_err = 0
    d_tables = 0
    d_table_vals: list[int] = []
    d_over = 0               # docling emitted MORE tables than GT (over-segment)
    d_under = 0              # docling emitted FEWER (collapse/miss)
    d_exact = 0
    d_errors = 0
    for rec in doc_recs:
        if rec.get("error"):
            d_errors += 1
            continue
        for s in rec.get("per_sheet", []):
            d_sheets += 1
            gt_n = s["gt_tables"]
            dn = s["docling_tables"]
            d_tables += dn
            d_table_vals.append(dn)
            d_abs_err += abs(dn - gt_n)
            if dn > gt_n:
                d_over += 1
            elif dn < gt_n:
                d_under += 1
            else:
                d_exact += 1

    lines = []
    lines.append("# DECO structural benchmark — ks vs docling\n")
    lines.append(f"Corpus: DECO `completed/` · GT tables scored: **{n_tables}** "
                 f"(across {len(ks_recs)} files)\n")
    lines.append(f"ks parse: {ks_timeouts} files timed out, {ks_errors} errored "
                 f"(excluded from metrics below).\n")
    lines.append("## Table-boundary detection (ks — needs A1 localisation)\n")
    lines.append("| metric | value |")
    lines.append("|---|---|")
    lines.append(f"| mean best-IoU vs GT table | {iou_sum / n_tables:.3f} |" if n_tables else "| mean best-IoU | n/a |")
    lines.append(f"| detected @ IoU≥0.5 | {det50}/{n_tables} ({100*det50/n_tables:.1f}%) |")
    lines.append(f"| detected @ IoU≥0.3 | {det30}/{n_tables} ({100*det30/n_tables:.1f}%) |")
    lines.append(f"| mean best-IoU, macro by file | {sum(macro_iou)/len(macro_iou):.3f} |" if macro_iou else "")
    lines.append(f"| ks regions overlapping one GT table (fragmentation) | mean {frag_sum / n_tables:.2f}, median {_median(frag_vals):.0f} |")
    lines.append(f"| GT tables split across >1 ks region | {over_seg}/{n_tables} ({100*over_seg/n_tables:.1f}%) |")
    lines.append("")
    lines.append("## Header-row detection (ks — shipped `find_header_span`)\n")
    lines.append(f"GT tables with a header: **{h_tables}** · of which multi-row: "
                 f"**{h_multirow}** ({100*h_multirow/h_tables:.1f}%)\n" if h_tables else "No GT headers.\n")
    lines.append("| metric | precision | recall | F1 |")
    lines.append("|---|---|---|---|")
    lines.append(f"| end-to-end (header on ks's own region) | {p_e2e:.3f} | {r_e2e:.3f} | {f_e2e:.3f} |")
    lines.append(f"| isolated (header on GT region) | {p_iso:.3f} | {r_iso:.3f} | {f_iso:.3f} |")
    if macro_hf1:
        lines.append(f"| end-to-end, macro F1 by file | | | {sum(macro_hf1)/len(macro_hf1):.3f} |")
    lines.append("")
    lines.append(f"- ks emitted a non-empty header for **{h_detected_e2e}/{h_tables}** "
                 f"({100*h_detected_e2e/h_tables:.1f}%) GT-headered tables (end-to-end).")
    lines.append(f"- exact header-row match (isolated): **{h_exact_iso}/{h_tables}** "
                 f"({100*h_exact_iso/h_tables:.1f}%).")
    lines.append("")
    lines.append("## Tables-per-sheet (docling — its only measurable axis)\n")
    lines.append("docling emits no A1 coordinates, so it can't be scored on IoU or "
                 "header rows. The one axis it exposes is how many table objects it "
                 "produces per sheet vs the GT count.\n")
    if d_sheets:
        lines.append("| metric | value |")
        lines.append("|---|---|")
        lines.append(f"| GT sheets scored | {d_sheets} |")
        lines.append(f"| docling tables per sheet | mean {d_tables / d_sheets:.2f}, median {_median(d_table_vals):.0f} |")
        lines.append(f"| mean \\|docling − GT\\| tables per sheet | {d_abs_err / d_sheets:.2f} |")
        lines.append(f"| sheets where docling = GT count | {d_exact}/{d_sheets} ({100*d_exact/d_sheets:.1f}%) |")
        lines.append(f"| sheets where docling **over**-segments (>GT) | {d_over}/{d_sheets} ({100*d_over/d_sheets:.1f}%) |")
        lines.append(f"| sheets where docling **under**-counts (<GT) | {d_under}/{d_sheets} ({100*d_under/d_sheets:.1f}%) |")
        lines.append(f"| docling convert errors | {d_errors} files |")
    else:
        lines.append("_No docling records (run `--parser docling` first)._")
    lines.append("")

    (out_dir / "summary.md").write_text("\n".join(lines))
    summary = {
        "n_tables": n_tables,
        "ks": {
            "mean_best_iou": iou_sum / n_tables if n_tables else None,
            "detect_at_0.5": det50 / n_tables if n_tables else None,
            "detect_at_0.3": det30 / n_tables if n_tables else None,
            "mean_fragments": frag_sum / n_tables if n_tables else None,
            "over_segmented_frac": over_seg / n_tables if n_tables else None,
            "header_tables": h_tables,
            "header_multirow": h_multirow,
            "header_e2e": {"precision": p_e2e, "recall": r_e2e, "f1": f_e2e},
            "header_iso": {"precision": p_iso, "recall": r_iso, "f1": f_iso},
            "header_detected_e2e_frac": h_detected_e2e / h_tables if h_tables else None,
            "header_exact_iso_frac": h_exact_iso / h_tables if h_tables else None,
        },
        "docling": {
            "sheets": d_sheets,
            "mean_tables_per_sheet": d_tables / d_sheets if d_sheets else None,
            "mean_abs_table_count_err": d_abs_err / d_sheets if d_sheets else None,
            "exact_frac": d_exact / d_sheets if d_sheets else None,
            "over_segment_frac": d_over / d_sheets if d_sheets else None,
            "under_count_frac": d_under / d_sheets if d_sheets else None,
            "convert_errors": d_errors,
        },
    }
    (out_dir / "summary.json").write_text(json.dumps(summary, indent=2))
    print("\n".join(lines))
    print(f"\n✓ report → {out_dir}/summary.md")


def _read_ndjson(path: Path) -> list[dict]:
    if not path.exists():
        return []
    out = []
    with path.open() as fh:
        for line in fh:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


# ───────────────────────────────────────────────────────────── main
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--corpus", default="data/corpora/deco/completed")
    ap.add_argument("--parser", choices=["ks", "docling"])
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--sample", type=int, default=None)
    ap.add_argument("--timeout", type=int, default=60, help="per-file timeout (s)")
    ap.add_argument("--out", required=True, help="output run directory")
    args = ap.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.report:
        report(out_dir)
        return 0
    if not args.parser:
        ap.error("specify --parser ks|docling or --report")

    files = iter_files(args.corpus, args.sample)
    print(f"{args.parser}: {len(files)} candidate files from {args.corpus}", file=sys.stderr)
    run_parser(args.parser, files, out_dir, timeout=args.timeout)
    return 0


if __name__ == "__main__":
    sys.exit(main())
