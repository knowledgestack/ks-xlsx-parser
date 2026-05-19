#!/usr/bin/env python3
"""Enrich a benchmark run with per-instance diagnostics for failure clustering.

`eval_retrieval.py --emit-failures` only sees *text-match* misses. For
citation-grade scoring (geometric recall@5 = 0.283) we also need to
classify instances where the answer text is in some chunk but no chunk's
A1 range covers the ground truth — those don't show up in failures.ndjson
at all.

This script re-parses each instance's input.xlsx with both ks-xlsx-parser
and openpyxl, then emits one row per FAILED instance (text-miss OR
geometric-miss) with diagnostic columns chosen so post-hoc clustering is
easy:

    instance_id           question id
    bucket_combined       both_miss / text_hit_geom_miss / text_miss_geom_hit
    answer_position       the GT spec from dataset.json
    gt_sheet              ground-truth sheet name (default: answer_sheet)
    gt_cell_raw           openpyxl raw value at the first GT cell
    gt_cell_formula       formula string if any
    gt_cell_data_only     cached computed value if any
    gt_in_workbook_sheets is gt_sheet in wb.sheetnames?
    gt_in_chunked_sheets  did the parser produce any chunk on gt_sheet?
    n_workbook_sheets     total sheets in the workbook (incl hidden)
    n_chunked_sheets      distinct sheets we emitted chunks for
    n_workbook_cells_in_gt   non-empty openpyxl cells in the GT range
    chunks_on_gt_sheet    how many chunks we emitted for gt_sheet
    chunk_bbox_on_gt_sheet bbox (min_r,min_c,max_r,max_c) over all chunks on gt_sheet
    gt_range_bbox         GT range as (r0,c0,r1,c1)
    text_match_rank       rank of first chunk whose text contains a GT value
    geom_match_rank       rank of first chunk whose A1 range overlaps GT

Usage:
    python scripts/enrich_failures.py <run-dir-or-results.ndjson>
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

# Hoist eval_retrieval helpers so we share the *exact* normalization
# logic — clustering by "answer present in chunk" must be apples-to-apples.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.eval_retrieval import (  # noqa: E402
    _matches_chunk_text,
    parse_a1,
    parse_position_spec,
    parse_range,
)

REPO_ROOT = Path(__file__).resolve().parent.parent


def find_run(arg: Path) -> Path:
    if arg.is_file():
        return arg.parent
    if arg.is_dir():
        if (arg / "results.ndjson").exists():
            return arg
        runs = sorted(p for p in arg.glob("*/results.ndjson"))
        if runs:
            return runs[-1].parent
    sys.exit(f"ERROR: no results.ndjson at {arg}")


def overlaps(box1: tuple[int, int, int, int], box2: tuple[int, int, int, int]) -> bool:
    r0a, c0a, r1a, c1a = box1
    r0b, c0b, r1b, c1b = box2
    return not (r1a < r0b or r0a > r1b or c1a < c0b or c0a > c1b)


def chunk_bbox(chunks) -> tuple[int, int, int, int] | None:
    boxes = []
    for c in chunks:
        tl = parse_a1(c.top_left_cell) if c.top_left_cell else None
        br = parse_a1(c.bottom_right_cell) if c.bottom_right_cell else None
        if tl and br:
            boxes.append((tl[0], tl[1], br[0], br[1]))
    if not boxes:
        return None
    return (
        min(b[0] for b in boxes),
        min(b[1] for b in boxes),
        max(b[2] for b in boxes),
        max(b[3] for b in boxes),
    )


def enrich(run_dir: Path, corpus: Path, out_path: Path) -> None:
    from openpyxl import load_workbook

    from ks_xlsx_parser.pipeline import parse_workbook

    # Load dataset.json once — we need question text + the original
    # answer_sheet attribution for instances where the rank scoring
    # already considers the spec parse-resolved.
    dataset = {str(d["id"]): d for d in
               json.loads((corpus / "dataset.json").read_text())}

    results_path = run_dir / "results.ndjson"
    rows: list[dict[str, Any]] = []
    for line in results_path.read_text().splitlines():
        if line.strip():
            rows.append(json.loads(line))

    out_rows: list[dict[str, Any]] = []
    n_failed = 0
    for rec in rows:
        if rec.get("error"):
            continue
        text_rank = rec.get("rank_of_text_match")
        geom_rank = rec.get("rank_of_first_overlap")
        text_hit = text_rank is not None and text_rank <= 5
        geom_hit = geom_rank is not None and geom_rank <= 5
        if text_hit and geom_hit:
            continue  # not a failure either way
        n_failed += 1

        inst_id = rec["instance_id"]
        meta = dataset.get(inst_id, {})
        instruction = meta.get("instruction", "")
        answer_sheet = meta.get("answer_sheet") or None
        if answer_sheet and "," in answer_sheet:
            answer_sheet = answer_sheet.split(",")[0].strip()
        answer_position = rec.get("answer_position") or meta.get("answer_position") or ""
        data_position = meta.get("data_position") or ""

        # Re-parse input.xlsx — cheap (50 ms median).
        inst_dir = corpus / "spreadsheet" / inst_id
        input_path = inst_dir / f"1_{inst_id}_input.xlsx"
        if not input_path.exists():
            continue

        try:
            result = parse_workbook(path=str(input_path))
            chunks = list(result.chunks)
        except Exception as exc:
            out_rows.append({
                "instance_id": inst_id,
                "bucket_combined": "parse_error",
                "error": f"{type(exc).__name__}: {exc}",
            })
            continue

        chunked_sheets = sorted({c.sheet_name for c in chunks if c.sheet_name})

        # openpyxl view — both formula and data_only passes so we can
        # tell "formula uncached" from "cell genuinely empty".
        try:
            wb_f = load_workbook(str(input_path), data_only=False, read_only=False)
            wb_d = load_workbook(str(input_path), data_only=True, read_only=False)
            wb_sheets = list(wb_f.sheetnames)
            hidden_sheets = [s for s in wb_sheets
                             if getattr(wb_f[s], "sheet_state", "visible") != "visible"]
        except Exception as exc:
            wb_sheets = []
            hidden_sheets = []
            wb_f = wb_d = None

        # Geometric overlap is scored against data_position (the input data
        # region the question is asking about); falls back to answer_position
        # when the dataset didn't fill data_position in (561 of 912 instances).
        # This mirrors eval_retrieval.py's geom_spec = data_pos or answer_pos.
        geom_spec = data_position or answer_position
        regions = parse_position_spec(geom_spec, answer_sheet)
        gt_sheet = regions[0][0] if regions else answer_sheet
        gt_range_bbox = regions[0][1] if regions else None

        # Separately track whether the *answer* region is empty in input.xlsx.
        # If so, the question is "compute X and write here" — the parser
        # cannot possibly contain the answer text, so this is a benchmark
        # construct, not a parser bug. We flag it as instruction_requires_execution.
        answer_regions = parse_position_spec(answer_position, answer_sheet)
        answer_sheet_resolved = (answer_regions[0][0]
                                 if answer_regions else answer_sheet)
        answer_range_bbox = answer_regions[0][1] if answer_regions else None
        n_input_cells_in_answer_range = 0
        if (wb_d and answer_sheet_resolved and answer_range_bbox
                and answer_sheet_resolved in wb_d.sheetnames):
            try:
                ws = wb_d[answer_sheet_resolved]
                r0, c0, r1, c1 = answer_range_bbox
                for row in ws.iter_rows(min_row=r0, max_row=r1, min_col=c0,
                                        max_col=c1, values_only=True):
                    for v in row:
                        if v is not None and str(v).strip():
                            n_input_cells_in_answer_range += 1
            except Exception:
                pass

        gt_cell_raw = None
        gt_cell_formula = None
        gt_cell_data_only = None
        n_workbook_cells_in_gt = 0
        if wb_f and gt_sheet and gt_sheet in wb_f.sheetnames and gt_range_bbox:
            ws_f = wb_f[gt_sheet]
            ws_d = wb_d[gt_sheet]
            r0, c0, r1, c1 = gt_range_bbox
            # First cell only — enough to know "formula vs. value".
            try:
                tl_cell_f = ws_f.cell(row=r0, column=c0)
                tl_cell_d = ws_d.cell(row=r0, column=c0)
                gt_cell_raw = tl_cell_f.value
                if isinstance(gt_cell_raw, str) and gt_cell_raw.startswith("="):
                    gt_cell_formula = gt_cell_raw
                gt_cell_data_only = tl_cell_d.value
            except Exception:
                pass
            # Count non-empty cells across the range
            try:
                for row in ws_d.iter_rows(min_row=r0, max_row=r1, min_col=c0,
                                          max_col=c1, values_only=True):
                    for v in row:
                        if v is not None and str(v).strip():
                            n_workbook_cells_in_gt += 1
            except Exception:
                pass

        chunks_on_gt = [c for c in chunks if gt_sheet and c.sheet_name == gt_sheet]
        gt_chunk_bbox = chunk_bbox(chunks_on_gt)

        if not text_hit and not geom_hit:
            bucket = "both_miss"
        elif text_hit and not geom_hit:
            bucket = "text_hit_geom_miss"
        elif geom_hit and not text_hit:
            bucket = "text_miss_geom_hit"
        else:
            bucket = "other"

        sheet_chunked = (gt_sheet in chunked_sheets) if gt_sheet else None
        sheet_in_wb = (gt_sheet in wb_sheets) if (gt_sheet and wb_sheets) else None
        sheet_hidden = (gt_sheet in hidden_sheets) if gt_sheet else False

        # Pre-named signal heuristics — informational only; clustering is
        # still done by reading. These flags help spot patterns FAST.
        flags: list[str] = []
        if sheet_in_wb is False:
            flags.append("gt_sheet_missing_from_workbook")
        elif sheet_chunked is False:
            flags.append("gt_sheet_present_but_not_chunked")
        if sheet_hidden:
            flags.append("gt_sheet_hidden")
        if gt_cell_formula and gt_cell_data_only in (None, ""):
            flags.append("gt_cell_uncached_formula")
        elif gt_cell_formula:
            flags.append("gt_cell_is_formula")
        if (gt_range_bbox and gt_chunk_bbox and
                not overlaps(gt_range_bbox, gt_chunk_bbox)):
            flags.append("gt_range_outside_chunk_bbox")
        if (gt_range_bbox and gt_chunk_bbox and
                overlaps(gt_range_bbox, gt_chunk_bbox)):
            # Inside the chunk bbox but no individual chunk overlaps?
            # That's "the parser saw the right area but split it wrong".
            any_overlap = False
            for c in chunks_on_gt:
                tl = parse_a1(c.top_left_cell) if c.top_left_cell else None
                br = parse_a1(c.bottom_right_cell) if c.bottom_right_cell else None
                if tl and br and overlaps(
                        gt_range_bbox, (tl[0], tl[1], br[0], br[1])):
                    any_overlap = True
                    break
            if not any_overlap:
                flags.append("gt_inside_bbox_but_no_chunk_overlap")
        if (n_workbook_cells_in_gt == 0 and gt_range_bbox):
            flags.append("gt_range_empty_in_workbook")
        # The big one: if answer_position is empty in input, the benchmark
        # is asking the system to WRITE the answer. Not a parser bug.
        if (answer_range_bbox and n_input_cells_in_answer_range == 0):
            flags.append("instruction_requires_execution")
        # cell rendered but truncated to a sub-range?
        if (gt_chunk_bbox and gt_range_bbox and
                gt_chunk_bbox[2] < gt_range_bbox[2]):
            flags.append("chunk_bbox_rows_truncated")
        if (gt_chunk_bbox and gt_range_bbox and
                gt_chunk_bbox[3] < gt_range_bbox[3]):
            flags.append("chunk_bbox_cols_truncated")

        out_rows.append({
            "instance_id": inst_id,
            "bucket_combined": bucket,
            "instruction": instruction[:200],
            "answer_position": answer_position,
            "answer_sheet": answer_sheet,
            "gt_sheet": gt_sheet,
            "gt_range_bbox": list(gt_range_bbox) if gt_range_bbox else None,
            "gt_cell_raw": str(gt_cell_raw)[:120] if gt_cell_raw is not None else None,
            "gt_cell_formula": gt_cell_formula,
            "gt_cell_data_only":
                str(gt_cell_data_only)[:120] if gt_cell_data_only is not None else None,
            "n_workbook_sheets": len(wb_sheets),
            "n_chunked_sheets": len(chunked_sheets),
            "wb_sheets": wb_sheets,
            "hidden_sheets": hidden_sheets,
            "chunked_sheets": chunked_sheets,
            "n_chunks_total": len(chunks),
            "n_chunks_on_gt_sheet": len(chunks_on_gt),
            "n_workbook_cells_in_gt": n_workbook_cells_in_gt,
            "chunk_bbox_on_gt_sheet": list(gt_chunk_bbox) if gt_chunk_bbox else None,
            "rank_of_text_match": text_rank,
            "rank_of_first_overlap": geom_rank,
            "flags": flags,
            "data_position": data_position,
            "answer_range_bbox": list(answer_range_bbox) if answer_range_bbox else None,
            "n_input_cells_in_answer_range": n_input_cells_in_answer_range,
        })

    out_path.write_text("\n".join(json.dumps(r, separators=(",", ":")) for r in out_rows) + "\n")
    print(f"Examined {len(rows)} instances, {n_failed} failed (text OR geom).")
    print(f"Wrote {len(out_rows)} enriched rows to {out_path}")

    # Quick histogram for sanity
    from collections import Counter
    bc = Counter(r["bucket_combined"] for r in out_rows)
    fc = Counter()
    for r in out_rows:
        for f in r.get("flags", []):
            fc[f] += 1
    print("\nCombined bucket counts:")
    for b, n in bc.most_common():
        print(f"  {b:<30s} {n}")
    print("\nDiagnostic flags (rows can have multiple):")
    for f, n in fc.most_common():
        print(f"  {f:<40s} {n}")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", type=Path,
                    help="run dir, results.ndjson, or parent reports dir")
    ap.add_argument("--corpus", type=Path,
                    default=REPO_ROOT / "data/corpora/spreadsheetbench/all_data_912_v0.1")
    ap.add_argument("--out", type=Path, default=None,
                    help="output path (default: <run>/enriched_failures.ndjson)")
    args = ap.parse_args(argv)

    run_dir = find_run(args.path)
    out = args.out or (run_dir / "enriched_failures.ndjson")
    enrich(run_dir, args.corpus, out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
