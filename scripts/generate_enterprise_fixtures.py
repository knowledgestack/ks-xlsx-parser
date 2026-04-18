"""Generate small, deterministic enterprise-style Excel fixtures.

These fixtures are used by enterprise scoring tests and corpus metrics.
They are intentionally lightweight so they can be generated at test time
without network access or large disk usage.
"""



from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parent.parent
TARGET_DIR = ROOT / "testBench" / "enterprise"


def _prepare_target() -> None:
    TARGET_DIR.mkdir(parents=True, exist_ok=True)


def create_financial_model() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Financial Model Q1 2026"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "ASSUMPTIONS"
    ws["A4"] = "Rent per unit"
    ws["B4"] = 2500
    ws["A5"] = "Units occupied"
    ws["B5"] = 42

    ws["A7"] = "RESULTS"
    ws["A8"] = "Total Revenue"
    ws["B8"] = "=B4*B5"

    wb.defined_names.add(DefinedName("UnitCount", attr_text="Model!$B$5"))
    wb.defined_names.add(DefinedName("RentPerUnit", attr_text="Model!$B$4"))

    return wb


def create_inventory_tracker() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Master"

    ws["A1"] = "SKU"
    ws["B1"] = "Description"
    ws["C1"] = "Qty"
    ws["D1"] = "Unit Cost"

    for i in range(2, 52):
        ws[f"A{i}"] = f"SKU-{i:04d}"
        ws[f"B{i}"] = f"Product {i}"
        ws[f"C{i}"] = i * 100
        ws[f"D{i}"] = i * 1.5

    tx = wb.create_sheet("Transactions")
    tx["A1"] = "SKU"
    tx["B1"] = "Qty"
    tx["C1"] = "Total"

    for i in range(2, 102):
        tx[f"A{i}"] = f"=Master!A{(i % 50) + 2}"
        tx[f"B{i}"] = (i % 10) + 1
        tx[f"C{i}"] = f"=VLOOKUP(A{i},Master!A:D,4,0)*B{i}"

    return wb


def create_forecast_model() -> Workbook:
    wb = Workbook()
    base = wb.active
    base.title = "Base"

    for month in range(1, 13):
        base[f"A{month}"] = f"Month {month}"
        base[f"B{month}"] = 10000 * (1 + month * 0.05)

    pess = wb.create_sheet("Pessimistic")
    opt = wb.create_sheet("Optimistic")
    for month in range(1, 13):
        pess[f"B{month}"] = f"=Base!B{month}*0.8"
        opt[f"B{month}"] = f"=Base!B{month}*1.2"

    return wb


def create_operations_tracker() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ops"

    ws["A1"] = "Project"
    ws["B1"] = "Status"
    ws["C1"] = "Budget"
    ws["D1"] = "Actual"
    ws["E1"] = "Variance %"

    statuses = ["Active", "Complete", "On Hold"]
    for i in range(2, 22):
        ws[f"A{i}"] = f"Project {i-1}"
        ws[f"B{i}"] = statuses[i % 3]
        ws[f"C{i}"] = i * 50000
        ws[f"D{i}"] = i * 50000 * (1 + (i % 5) * 0.1)
        ws[f"E{i}"] = f"=(D{i}-C{i})/C{i}"

    ref = wb.create_sheet("Reference", 1)
    ref.sheet_state = "hidden"
    ref["A1"] = "Rate"
    ref["A2"] = 1.05

    return wb


def _write_workbook(name: str, builder: Callable[[], Workbook]) -> Path:
    _prepare_target()
    path = TARGET_DIR / name
    if path.exists():
        return path
    wb = builder()
    wb.save(path)
    return path


def generate_all() -> list[Path]:
    """Generate all enterprise fixtures and return their paths."""
    fixtures = [
        ("financial_model.xlsx", create_financial_model),
        ("inventory_tracker.xlsx", create_inventory_tracker),
        ("forecast_model.xlsx", create_forecast_model),
        ("operations_tracker.xlsx", create_operations_tracker),
    ]

    return [_write_workbook(name, builder) for name, builder in fixtures]


if __name__ == "__main__":
    paths = generate_all()
    for p in paths:
        print(f"✓ Generated {p.relative_to(ROOT)}")
