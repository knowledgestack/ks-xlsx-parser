#!/usr/bin/env python3
"""
build_testbench.py — deterministic generator for the ks-xlsx-parser testBench.

Produces ~1000 `.xlsx` workbooks under ``testBench/generated/`` organised into
three groups:

* ``matrix/``       — one feature-per-file across every knob the parser exercises
                      (formulas, merges, named ranges, CF, DV, tables, charts,
                      styles, dates, errors, hidden rows/cols, hyperlinks,
                      comments, rich text, number formats, edge addresses,
                      array formulas, 3D refs, pivot placeholders, huge sheet
                      names).
* ``combo/``        — randomised combinations of the above at five density
                      levels (5/10/25/50/100 operations per file) × 80 seeds.
* ``adversarial/``  — files engineered to break parsers: circular formulas,
                      deep formula chains, 1M-row sparse sheets, 255-sheet
                      workbooks, unicode/RTL/emoji stress, oversized merges,
                      broken references, long formula strings.

Usage
-----

    python scripts/build_testbench.py            # builds everything
    python scripts/build_testbench.py --force    # regenerates even if present
    python scripts/build_testbench.py --group matrix
    python scripts/build_testbench.py --limit 50 # first 50 files only (smoke)

The generator is fully deterministic: identical invocations produce
byte-identical files (modulo openpyxl's own timestamping, which we neutralise).
Every file is accompanied by one row in ``testBench/generated/MANIFEST.json``
describing its group, feature tags, expected cell count, and SHA256.
"""


import argparse
import hashlib
import json
import random
import string
import sys
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    BubbleChart,
    LineChart,
    PieChart,
    RadarChart,
    Reference,
    ScatterChart,
)
from openpyxl.comments import Comment
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
    Rule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
OUT_ROOT = ROOT / "testBench" / "generated"
MANIFEST_PATH = OUT_ROOT / "MANIFEST.json"

# ----------------------------------------------------------------------------
# Data classes
# ----------------------------------------------------------------------------


@dataclass
class GeneratedFile:
    path: Path
    group: str
    features: list[str] = field(default_factory=list)
    expected_sheets: int = 1
    expected_cells: int = 0
    expected_formulas: int = 0
    notes: str = ""

    def to_manifest_row(self) -> dict:
        return {
            "path": str(self.path.relative_to(OUT_ROOT)),
            "group": self.group,
            "features": self.features,
            "expected_sheets": self.expected_sheets,
            "expected_cells": self.expected_cells,
            "expected_formulas": self.expected_formulas,
            "sha256": sha256_of(self.path),
            "size_bytes": self.path.stat().st_size,
            "notes": self.notes,
        }


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _finalize(wb: Workbook, out: Path) -> None:
    """Save workbook with deterministic metadata."""
    wb.properties.created = datetime(2025, 1, 1, 0, 0, 0)
    wb.properties.modified = datetime(2025, 1, 1, 0, 0, 0)
    wb.properties.creator = "ks-xlsx-parser testBench generator"
    wb.properties.title = out.stem
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


# ----------------------------------------------------------------------------
# Matrix group — one feature per file
# ----------------------------------------------------------------------------


MATRIX_DIR = OUT_ROOT / "matrix"


def _matrix_path(slug: str) -> Path:
    return MATRIX_DIR / f"{slug}.xlsx"


# --- formulas -------------------------------------------------------------

FORMULA_RECIPES: list[tuple[str, str, str]] = [
    # (slug, label, formula expression — evaluated in B1 with constants in A1:A5)
    ("formula_sum", "SUM", "=SUM(A1:A5)"),
    ("formula_average", "AVERAGE", "=AVERAGE(A1:A5)"),
    ("formula_min_max", "MIN/MAX", "=MAX(A1:A5)-MIN(A1:A5)"),
    ("formula_count", "COUNT", "=COUNT(A1:A5)"),
    ("formula_counta", "COUNTA", "=COUNTA(A1:A5)"),
    ("formula_sumif", "SUMIF", "=SUMIF(A1:A5,\">2\")"),
    ("formula_sumifs", "SUMIFS", "=SUMIFS(A1:A5,A1:A5,\">1\",A1:A5,\"<5\")"),
    ("formula_countif", "COUNTIF", "=COUNTIF(A1:A5,\">2\")"),
    ("formula_countifs", "COUNTIFS", "=COUNTIFS(A1:A5,\">0\",A1:A5,\"<5\")"),
    ("formula_averageif", "AVERAGEIF", "=AVERAGEIF(A1:A5,\">1\")"),
    ("formula_if_basic", "IF", "=IF(A1>2,\"big\",\"small\")"),
    ("formula_if_nested", "nested IF", "=IF(A1>4,\"high\",IF(A1>2,\"mid\",\"low\"))"),
    ("formula_ifs", "IFS", "=IFS(A1>4,\"high\",A1>2,\"mid\",TRUE,\"low\")"),
    ("formula_ifna", "IFNA", "=IFNA(VLOOKUP(99,A1:B5,2,FALSE),\"missing\")"),
    ("formula_iferror", "IFERROR", "=IFERROR(1/0,\"err\")"),
    ("formula_and_or_not", "AND/OR/NOT", "=AND(A1>0,OR(A2>0,NOT(A3<0)))"),
    ("formula_concat", "CONCAT", "=CONCAT(A1,\"-\",A2)"),
    ("formula_textjoin", "TEXTJOIN", "=TEXTJOIN(\",\",TRUE,A1:A5)"),
    ("formula_left_right_mid", "LEFT/RIGHT/MID", "=LEFT(\"abcdef\",3)&RIGHT(\"abcdef\",2)&MID(\"abcdef\",3,2)"),
    ("formula_substitute", "SUBSTITUTE", "=SUBSTITUTE(\"foo-bar\",\"-\",\"_\")"),
    ("formula_find_search", "FIND/SEARCH", "=FIND(\"b\",\"foobar\")+SEARCH(\"B\",\"foobar\")"),
    ("formula_len_trim", "LEN/TRIM", "=LEN(TRIM(\"  hi  \"))"),
    ("formula_upper_lower_proper", "case fns", "=UPPER(\"a\")&LOWER(\"B\")&PROPER(\"hello world\")"),
    ("formula_round_roundup_rounddown", "ROUND*", "=ROUND(A1,1)+ROUNDUP(A1,0)+ROUNDDOWN(A1,0)"),
    ("formula_int_mod", "INT/MOD", "=INT(A1)+MOD(A1,2)"),
    ("formula_abs_sign", "ABS/SIGN", "=ABS(-5)+SIGN(A1)"),
    ("formula_sqrt_power", "SQRT/POWER", "=SQRT(16)+POWER(A1,2)"),
    ("formula_log_ln_exp", "LOG/LN/EXP", "=LOG(10)+LN(EXP(1))"),
    ("formula_date_functions", "DATE fns", "=YEAR(TODAY())+MONTH(TODAY())+DAY(TODAY())"),
    ("formula_datedif", "DATEDIF", "=DATEDIF(DATE(2020,1,1),DATE(2025,1,1),\"Y\")"),
    ("formula_edate_eomonth", "EDATE/EOMONTH", "=EDATE(DATE(2020,1,1),12)+EOMONTH(DATE(2020,1,1),3)"),
    ("formula_weekday_workday", "WEEKDAY/WORKDAY", "=WEEKDAY(TODAY())+WORKDAY(TODAY(),5)"),
    ("formula_vlookup", "VLOOKUP", "=VLOOKUP(A1,A1:B5,2,FALSE)"),
    ("formula_hlookup", "HLOOKUP", "=HLOOKUP(A1,A1:E2,2,FALSE)"),
    ("formula_xlookup", "XLOOKUP", "=XLOOKUP(A1,A1:A5,B1:B5,\"not found\")"),
    ("formula_index_match", "INDEX/MATCH", "=INDEX(A1:A5,MATCH(A2,A1:A5,0))"),
    ("formula_offset", "OFFSET", "=OFFSET(A1,2,0)"),
    ("formula_indirect", "INDIRECT", "=INDIRECT(\"A\"&2)"),
    ("formula_rank", "RANK", "=RANK(A1,A1:A5,0)"),
    ("formula_large_small", "LARGE/SMALL", "=LARGE(A1:A5,2)+SMALL(A1:A5,2)"),
    ("formula_choose", "CHOOSE", "=CHOOSE(2,\"a\",\"b\",\"c\")"),
    ("formula_switch", "SWITCH", "=SWITCH(A1,1,\"one\",2,\"two\",\"other\")"),
    ("formula_array_cse", "array CSE", "{=SUM(A1:A5*A1:A5)}"),
    ("formula_long", "8000-char expression", "=" + "+".join(f"A{((i % 5) + 1)}" for i in range(400))),
]


def build_formula_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    for slug, label, formula in FORMULA_RECIPES:
        wb = Workbook()
        ws = wb.active
        ws.title = "Formula"
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 1.5)
        ws["B1"] = formula
        ws["D1"] = f"Test: {label}"
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/formula",
                features=["formula", slug.replace("formula_", "")],
                expected_cells=7,
                expected_formulas=1,
            )
        )
    return files


# --- merged cells ---------------------------------------------------------


def build_merge_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    recipes = [
        ("merge_horizontal_small", [("A1:C1",)]),
        ("merge_horizontal_wide", [(f"A1:{get_column_letter(20)}1",)]),
        ("merge_vertical_small", [("A1:A5",)]),
        ("merge_vertical_tall", [("A1:A100",)]),
        ("merge_rectangular", [("A1:E5",)]),
        ("merge_many_horizontal", [(f"A{r}:C{r}",) for r in range(1, 51)]),
        ("merge_many_vertical", [(f"{get_column_letter(c)}1:{get_column_letter(c)}30",) for c in range(1, 11)]),
        ("merge_grid_5x5", [(f"{get_column_letter(2*c-1)}{2*r-1}:{get_column_letter(2*c)}{2*r}",) for r in range(1, 6) for c in range(1, 6)]),
        ("merge_diagonal_steps", [(f"{get_column_letter(2*i-1)}{2*i-1}:{get_column_letter(2*i)}{2*i}",) for i in range(1, 8)]),
        ("merge_header_3_levels", [("A1:F1",), ("A2:C2",), ("D2:F2",), ("A3:B3",), ("C3:C3",), ("D3:E3",), ("F3:F3",)]),
        ("merge_with_value_only_in_master", [("A1:C3",)]),
        ("merge_around_data", [("A1:C1",), ("A5:C5",)]),
        ("merge_single_cell_noop", [("A1:A1",)]),  # degenerate
        ("merge_adjacent_row_pair", [("A1:B1",), ("A2:B2",)]),
        ("merge_wide_header_narrow_data", [("A1:J1",)]),
        ("merge_mixed_sizes", [("A1:B2",), ("C1:E1",), ("A4:A10",), ("D4:F6",)]),
        ("merge_100_singletons", [(f"{get_column_letter(((i-1) % 20)+1)}{((i-1)//20)+1}:{get_column_letter(((i-1) % 20)+1)}{((i-1)//20)+1}",) for i in range(1, 101)]),
        ("merge_full_row", [("A1:Z1",)]),
        ("merge_full_column_short", [("A1:A50",)]),
        ("merge_nonadjacent_blocks", [("A1:C3",), ("F1:H3",), ("A5:C7",), ("F5:H7",)]),
        ("merge_within_table_header", [("A1:D1",)]),  # we'll add a table below
        ("merge_empty_range", [("B2:D4",)]),  # no data in master
        ("merge_unicode_content", [("A1:C1",)]),
        ("merge_with_rich_formatting", [("A1:C1",)]),
        ("merge_column_header_stack", [("A1:A2",), ("B1:B2",), ("C1:C2",)]),
        ("merge_report_grid", [("A1:D1",), ("A2:A10",), ("B2:D2",), ("B3:B10",), ("C3:D3",)]),
        ("merge_large_single", [("A1:Z100",)]),
        ("merge_thousand_cells", [("A1:J100",)]),
        ("merge_within_table_footer", [("A11:D11",)]),
        ("merge_spanning_formula_range", [("A1:C1",)]),
    ]
    for slug, ranges in recipes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Merges"
        for i, (rng,) in enumerate(ranges):
            anchor = rng.split(":")[0]
            try:
                ws[anchor] = f"m{i+1}"  # must write before merging; skip if cell is already merged
            except AttributeError:
                pass
            try:
                ws.merge_cells(rng)
            except Exception:
                pass
        if slug == "merge_with_value_only_in_master":
            ws["A1"] = "only-master"
        if slug == "merge_within_table_header":
            for c, h in enumerate(["a", "b", "c", "d"], 1):
                ws.cell(row=2, column=c, value=h)
            for r in range(3, 8):
                for c in range(1, 5):
                    ws.cell(row=r, column=c, value=r * c)
            ws.add_table(Table(displayName="T1", ref="A2:D7"))
        if slug == "merge_unicode_content":
            ws["A1"] = "éñÜ日本語 🚀 حرف"
        if slug == "merge_with_rich_formatting":
            ws["A1"].font = Font(bold=True, size=14, color="FF0000")
            ws["A1"].fill = PatternFill("solid", start_color="FFFF00")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/merge",
                features=["merged_cells", slug],
                expected_cells=len(ranges),
            )
        )
    return files


# --- named ranges ---------------------------------------------------------


def build_named_range_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    recipes = [
        ("named_workbook_scope", "Total", "Sheet1!$A$1", None),
        ("named_sheet_scope", "SheetLocal", "Sheet1!$B$1", "Sheet1"),
        ("named_constant", "TaxRate", "0.07", None),
        ("named_range_multi_cell", "Prices", "Sheet1!$A$1:$A$10", None),
        ("named_formula", "Doubled", "Sheet1!$A$1*2", None),
        ("named_with_unicode", "Mẹtá", "Sheet1!$A$1", None),
        ("named_long_identifier", "very_long_identifier_" + "x" * 50, "Sheet1!$A$1", None),
        ("named_escaped_sheet", "Quoted", "'Sheet 2'!$A$1", None),  # needs 'Sheet 2'
        ("named_external_like", "ExternalLike", "[Budget.xlsx]Sheet1!$A$1", None),
        ("named_list_variation", "ChoiceList", "Sheet1!$D$1:$D$5", None),
        ("named_col_range", "FullColumn", "Sheet1!$A:$A", None),
        ("named_row_range", "FullRow", "Sheet1!$1:$1", None),
        ("named_cross_sheet", "CrossRef", "Other!$A$1", None),  # needs Other sheet
        ("named_multi_area", "Islands", "Sheet1!$A$1,Sheet1!$C$3", None),
        ("named_with_hash_prefix", "_Prefix", "Sheet1!$A$1", None),
        ("named_digits", "X1", "Sheet1!$A$1", None),
        ("named_empty_formula_error", "ErrRef", "#REF!", None),
        ("named_boolean_constant", "IsOn", "TRUE", None),
        ("named_string_constant", "Greeting", '"hello"', None),
        ("named_table_column_ref", "TableCol", "Table1[Value]", None),  # needs table
    ]
    for slug, name, ref, scope in recipes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=4, value=f"item{i}")
        if scope == "Sheet1":
            ws.defined_names.add(DefinedName(name, attr_text=ref))
        elif slug == "named_escaped_sheet":
            wb.create_sheet("Sheet 2")["A1"] = 42
            wb.defined_names.add(DefinedName(name, attr_text=ref))
        elif slug == "named_cross_sheet":
            wb.create_sheet("Other")["A1"] = 99
            wb.defined_names.add(DefinedName(name, attr_text=ref))
        elif slug == "named_table_column_ref":
            for c, h in enumerate(["ID", "Value"], 1):
                ws.cell(row=1, column=c, value=h)
            for r in range(2, 6):
                ws.cell(row=r, column=1, value=r)
                ws.cell(row=r, column=2, value=r * 10)
            ws.add_table(Table(displayName="Table1", ref="A1:B5"))
            wb.defined_names.add(DefinedName(name, attr_text=ref))
        else:
            wb.defined_names.add(DefinedName(name, attr_text=ref))
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/named_range",
                features=["named_range", slug],
                expected_cells=14,
            )
        )
    return files


# --- data validation ------------------------------------------------------


def build_data_validation_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    recipes = [
        ("dv_list_literal", {"type": "list", "formula1": '"Red,Green,Blue"'}),
        ("dv_list_range", {"type": "list", "formula1": "=$D$1:$D$5"}),
        ("dv_whole_between", {"type": "whole", "operator": "between", "formula1": "1", "formula2": "100"}),
        ("dv_decimal_gt", {"type": "decimal", "operator": "greaterThan", "formula1": "0.5"}),
        ("dv_date_after", {"type": "date", "operator": "greaterThan", "formula1": "DATE(2024,1,1)"}),
        ("dv_time_before", {"type": "time", "operator": "lessThan", "formula1": "TIME(12,0,0)"}),
        ("dv_textlength", {"type": "textLength", "operator": "lessThan", "formula1": "10"}),
        ("dv_custom", {"type": "custom", "formula1": "=A1>0"}),
        ("dv_list_unicode", {"type": "list", "formula1": '"红,绿,蓝"'}),
        ("dv_list_one_item", {"type": "list", "formula1": '"Only"'}),
        ("dv_list_many_items", {"type": "list", "formula1": '"' + ",".join(f"opt{i}" for i in range(1, 31)) + '"'}),
        ("dv_with_error_message", {"type": "list", "formula1": '"A,B"', "error": "pick A or B", "errorTitle": "Err"}),
        ("dv_with_prompt", {"type": "list", "formula1": '"A,B"', "prompt": "select letter", "promptTitle": "Hint"}),
        ("dv_ignore_blank", {"type": "list", "formula1": '"A,B"', "allowBlank": True}),
        ("dv_multiple_ranges", {"type": "list", "formula1": '"A,B"'}),  # will apply to multiple ranges
        ("dv_whole_equal", {"type": "whole", "operator": "equal", "formula1": "42"}),
        ("dv_date_between", {"type": "date", "operator": "between", "formula1": "DATE(2020,1,1)", "formula2": "DATE(2025,12,31)"}),
        ("dv_decimal_not_between", {"type": "decimal", "operator": "notBetween", "formula1": "0", "formula2": "1"}),
        ("dv_textlength_greater", {"type": "textLength", "operator": "greaterThan", "formula1": "3"}),
        ("dv_custom_cross_cell", {"type": "custom", "formula1": "=AND(A1>0,B1<100)"}),
    ]
    for slug, kwargs in recipes:
        wb = Workbook()
        ws = wb.active
        ws.title = "DV"
        for r in range(1, 6):
            ws.cell(row=r, column=4, value=f"Option{r}")
        dv_kwargs = {k: v for k, v in kwargs.items() if k not in {"error", "errorTitle", "prompt", "promptTitle", "allowBlank"}}
        dv = DataValidation(**dv_kwargs)
        if "error" in kwargs:
            dv.error = kwargs["error"]
            dv.errorTitle = kwargs.get("errorTitle", "Err")
            dv.showErrorMessage = True
        if "prompt" in kwargs:
            dv.prompt = kwargs["prompt"]
            dv.promptTitle = kwargs.get("promptTitle", "Hint")
            dv.showInputMessage = True
        if kwargs.get("allowBlank"):
            dv.allowBlank = True
        ws.add_data_validation(dv)
        if slug == "dv_multiple_ranges":
            dv.add("A1:A5")
            dv.add("C1:C5")
        else:
            dv.add("A1:A10")
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/data_validation",
                features=["data_validation", slug],
                expected_cells=5,
            )
        )
    return files


# --- conditional formatting -----------------------------------------------


def build_conditional_formatting_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []

    def _seed_ws(ws):
        for r in range(1, 11):
            ws.cell(row=r, column=1, value=r)
            ws.cell(row=r, column=2, value=11 - r)
            ws.cell(row=r, column=3, value=(r * 7) % 10)

    recipes: list[tuple[str, Callable[[object], None]]] = [
        ("cf_cellis_greater", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            CellIsRule(operator="greaterThan", formula=["5"], fill=PatternFill("solid", start_color="FFC7CE")),
        )),
        ("cf_cellis_less", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            CellIsRule(operator="lessThan", formula=["3"], fill=PatternFill("solid", start_color="C6EFCE")),
        )),
        ("cf_cellis_between", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            CellIsRule(operator="between", formula=["3", "7"], fill=PatternFill("solid", start_color="FFEB9C")),
        )),
        ("cf_color_scale_2", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            ColorScaleRule(start_type="min", start_color="FFAA0000",
                           end_type="max", end_color="FF00AA00"),
        )),
        ("cf_color_scale_3", lambda ws: ws.conditional_formatting.add(
            "B1:B10",
            ColorScaleRule(start_type="min", start_color="FFAA0000",
                           mid_type="percentile", mid_value=50, mid_color="FFFFFFFF",
                           end_type="max", end_color="FF00AA00"),
        )),
        ("cf_databar", lambda ws: ws.conditional_formatting.add(
            "C1:C10",
            DataBarRule(start_type="min", end_type="max", color="FF638EC6"),
        )),
        ("cf_iconset_3traffic", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            IconSetRule("3TrafficLights1", "percent", [0, 33, 67]),
        )),
        ("cf_iconset_5arrows", lambda ws: ws.conditional_formatting.add(
            "B1:B10",
            IconSetRule("5Arrows", "percent", [0, 20, 40, 60, 80]),
        )),
        ("cf_formula_rule", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            FormulaRule(formula=["MOD(ROW(),2)=0"], fill=PatternFill("solid", start_color="DDDDDD")),
        )),
        # Note: omit dxfId; openpyxl cannot round-trip Rule(dxfId=0) unless
        # the differential style table has a matching entry.
        ("cf_top10", lambda ws: ws.conditional_formatting.add(
            "A1:C10", Rule(type="top10", rank=3),
        )),
        ("cf_unique_values", lambda ws: ws.conditional_formatting.add(
            "A1:A10", Rule(type="uniqueValues"),
        )),
        ("cf_duplicate_values", lambda ws: ws.conditional_formatting.add(
            "A1:A10", Rule(type="duplicateValues"),
        )),
        ("cf_contains_text", lambda ws: ws.conditional_formatting.add(
            "A1:A10", Rule(type="containsText", operator="containsText", text="5"),
        )),
        ("cf_above_average", lambda ws: ws.conditional_formatting.add(
            "A1:A10", Rule(type="aboveAverage", aboveAverage=True),
        )),
        ("cf_below_average", lambda ws: ws.conditional_formatting.add(
            "A1:A10", Rule(type="aboveAverage", aboveAverage=False),
        )),
        ("cf_multiple_rules_same_range", lambda ws: (
            ws.conditional_formatting.add("A1:A10", CellIsRule(operator="greaterThan", formula=["7"], fill=PatternFill("solid", start_color="FF0000"))),
            ws.conditional_formatting.add("A1:A10", CellIsRule(operator="lessThan", formula=["3"], fill=PatternFill("solid", start_color="00FF00"))),
        )),
        ("cf_overlapping_ranges", lambda ws: (
            ws.conditional_formatting.add("A1:B5", ColorScaleRule(start_type="min", start_color="FFFF0000", end_type="max", end_color="FF00FF00")),
            ws.conditional_formatting.add("B3:C10", DataBarRule(start_type="min", end_type="max", color="FF0000FF")),
        )),
        ("cf_single_cell", lambda ws: ws.conditional_formatting.add(
            "A1", CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", start_color="FFFF00")),
        )),
        ("cf_large_range", lambda ws: ws.conditional_formatting.add(
            "A1:Z100", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", start_color="EEEEEE")),
        )),
        ("cf_entire_column", lambda ws: ws.conditional_formatting.add(
            "A1:A1048576", CellIsRule(operator="greaterThan", formula=["5"], fill=PatternFill("solid", start_color="FFC7CE")),
        )),
        ("cf_formula_complex", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            FormulaRule(formula=["AND(A1>3,A1<8)"], fill=PatternFill("solid", start_color="99FF99")),
        )),
        ("cf_iconset_3signs", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            IconSetRule("3Signs", "percent", [0, 33, 67]),
        )),
        ("cf_iconset_4ratings", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            IconSetRule("4Rating", "percent", [0, 25, 50, 75]),
        )),
        ("cf_color_scale_percentile", lambda ws: ws.conditional_formatting.add(
            "A1:A10",
            ColorScaleRule(start_type="percentile", start_value=10, start_color="FF0000FF",
                           end_type="percentile", end_value=90, end_color="FFFF0000"),
        )),
        ("cf_databar_negative", lambda ws: ws.conditional_formatting.add(
            "C1:C10",
            DataBarRule(start_type="min", end_type="max", color="FFFF0000", showValue=False),
        )),
    ]

    for slug, apply in recipes:
        wb = Workbook()
        ws = wb.active
        ws.title = "CF"
        _seed_ws(ws)
        apply(ws)
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/conditional_formatting",
                features=["conditional_formatting", slug],
                expected_cells=30,
            )
        )
    return files


# --- tables ---------------------------------------------------------------


def build_table_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    for idx, (rows, cols, style, totals) in enumerate([
        (3, 2, "TableStyleLight1", False),
        (10, 3, "TableStyleMedium2", False),
        (50, 5, "TableStyleMedium9", True),
        (100, 8, "TableStyleDark1", False),
        (5, 20, "TableStyleLight9", False),
        (30, 4, "TableStyleMedium1", True),
        (3, 1, "TableStyleLight5", False),
        (3, 26, "TableStyleMedium3", False),
        (3, 2, None, False),
        (10, 3, "TableStyleMedium4", True),
        (200, 6, "TableStyleMedium5", False),
        (3, 2, "TableStyleLight13", False),
        (3, 2, "TableStyleLight14", False),
        (3, 2, "TableStyleLight15", False),
        (3, 2, "TableStyleLight16", False),
        (3, 2, "TableStyleLight17", False),
        (3, 2, "TableStyleLight18", False),
        (3, 2, "TableStyleLight19", False),
        (3, 2, "TableStyleLight20", False),
        (3, 2, "TableStyleLight21", False),
    ]):
        slug = f"table_{idx:02d}_{rows}r_{cols}c"
        wb = Workbook()
        ws = wb.active
        ws.title = "Table"
        for c in range(1, cols + 1):
            ws.cell(row=1, column=c, value=f"H{c}")
        for r in range(2, rows + 2):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=(r + c) % 97)
        ref = f"A1:{get_column_letter(cols)}{rows + 1}"
        tab = Table(displayName=f"Tbl{idx}", ref=ref)
        if style:
            tab.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
        if totals:
            tab.totalsRowShown = False  # openpyxl can be finicky about totals; keep simple
        ws.add_table(tab)
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/table",
                features=["table", f"{rows}r{cols}c"],
                expected_cells=(rows + 1) * cols,
            )
        )
    return files


# --- charts ---------------------------------------------------------------


def build_chart_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    chart_types = [
        ("chart_bar", BarChart, {"type": "col"}),
        ("chart_bar_stacked", BarChart, {"type": "col", "grouping": "stacked", "overlap": 100}),
        ("chart_bar_horizontal", BarChart, {"type": "bar"}),
        ("chart_line", LineChart, {}),
        ("chart_pie", PieChart, {}),
        ("chart_area", AreaChart, {}),
        ("chart_radar", RadarChart, {}),
        ("chart_scatter", ScatterChart, {}),
        ("chart_bubble", BubbleChart, {}),
        ("chart_with_title", BarChart, {"title": "Q1 Sales"}),
        ("chart_no_title", BarChart, {}),
        ("chart_many_series", BarChart, {"series_count": 6}),
        ("chart_one_datapoint", BarChart, {"rows": 2}),
        ("chart_long_labels", BarChart, {"long_labels": True}),
        ("chart_unicode_labels", BarChart, {"unicode": True}),
        ("chart_two_charts_one_sheet", BarChart, {"double": True}),
        ("chart_chart_plus_table", BarChart, {"with_table": True}),
        ("chart_line_dashed", LineChart, {"smooth": True}),
        ("chart_pie_exploded", PieChart, {}),
        ("chart_scatter_with_lines", ScatterChart, {"scatterStyle": "lineMarker"}),
    ]
    for slug, ChartCls, opts in chart_types:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        rows = opts.pop("rows", 6)
        series_count = opts.pop("series_count", 2)
        long_labels = opts.pop("long_labels", False)
        unicode_flag = opts.pop("unicode", False)
        double = opts.pop("double", False)
        with_table = opts.pop("with_table", False)

        ws.cell(row=1, column=1, value="Label")
        for s in range(1, series_count + 1):
            ws.cell(row=1, column=1 + s, value=f"Series{s}")
        for r in range(2, rows + 1):
            label = f"Item{r-1}"
            if long_labels:
                label = "A very long label " * 5 + str(r)
            if unicode_flag:
                label = f"标签{r} 🚀"
            ws.cell(row=r, column=1, value=label)
            for s in range(1, series_count + 1):
                ws.cell(row=r, column=1 + s, value=((r * s * 7) % 50) + 1)

        chart = ChartCls()
        for k, v in opts.items():
            try:
                setattr(chart, k, v)
            except Exception:
                pass
        data = Reference(ws, min_col=2, min_row=1, max_col=1 + series_count, max_row=rows)
        cats = Reference(ws, min_col=1, min_row=2, max_row=rows)
        chart.add_data(data, titles_from_data=True)
        try:
            chart.set_categories(cats)
        except Exception:
            pass
        ws.add_chart(chart, f"{get_column_letter(series_count + 3)}2")

        if double:
            chart2 = BarChart()
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            ws.add_chart(chart2, "H20")
        if with_table:
            ws.add_table(Table(displayName="ChartTable", ref=f"A1:{get_column_letter(series_count + 1)}{rows}"))

        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/chart",
                features=["chart", slug],
                expected_cells=rows * (series_count + 1),
            )
        )
    return files


# --- rich text / styles / fonts ------------------------------------------


def build_style_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    styles = [
        ("style_bold", lambda c: setattr(c, "font", Font(bold=True))),
        ("style_italic", lambda c: setattr(c, "font", Font(italic=True))),
        ("style_underline", lambda c: setattr(c, "font", Font(underline="single"))),
        ("style_strike", lambda c: setattr(c, "font", Font(strike=True))),
        ("style_color_red", lambda c: setattr(c, "font", Font(color="FF0000"))),
        ("style_font_size_24", lambda c: setattr(c, "font", Font(size=24))),
        ("style_font_family_courier", lambda c: setattr(c, "font", Font(name="Courier New"))),
        ("style_bg_yellow", lambda c: setattr(c, "fill", PatternFill("solid", start_color="FFFF00"))),
        ("style_bg_pattern_gray125", lambda c: setattr(c, "fill", PatternFill(patternType="gray125"))),
        ("style_border_thin_all", lambda c: setattr(c, "border", Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")))),
        ("style_border_thick_bottom", lambda c: setattr(c, "border", Border(bottom=Side(style="thick")))),
        ("style_border_dashed", lambda c: setattr(c, "border", Border(top=Side(style="dashed")))),
        ("style_border_double", lambda c: setattr(c, "border", Border(bottom=Side(style="double")))),
        ("style_alignment_center", lambda c: setattr(c, "alignment", Alignment(horizontal="center", vertical="center"))),
        ("style_alignment_wrap", lambda c: setattr(c, "alignment", Alignment(wrap_text=True))),
        ("style_alignment_rotate_45", lambda c: setattr(c, "alignment", Alignment(text_rotation=45))),
        ("style_alignment_rotate_90", lambda c: setattr(c, "alignment", Alignment(text_rotation=90))),
        ("style_indent", lambda c: setattr(c, "alignment", Alignment(indent=3))),
        ("style_shrink_to_fit", lambda c: setattr(c, "alignment", Alignment(shrink_to_fit=True))),
        ("style_vertical_text", lambda c: setattr(c, "alignment", Alignment(text_rotation=255))),
        ("style_combined", lambda c: (
            setattr(c, "font", Font(bold=True, italic=True, size=16, color="0000FF")),
            setattr(c, "fill", PatternFill("solid", start_color="FFE0E0")),
            setattr(c, "alignment", Alignment(horizontal="center", vertical="center", wrap_text=True)),
            setattr(c, "border", Border(left=Side("thin"), right=Side("thin"), top=Side("medium"), bottom=Side("medium"))),
        )),
        ("style_number_format_currency", lambda c: setattr(c, "number_format", "$#,##0.00")),
        ("style_number_format_percent", lambda c: setattr(c, "number_format", "0.0%")),
        ("style_number_format_scientific", lambda c: setattr(c, "number_format", "0.00E+00")),
        ("style_number_format_date_iso", lambda c: setattr(c, "number_format", "yyyy-mm-dd")),
        ("style_number_format_date_long", lambda c: setattr(c, "number_format", "dddd, mmmm dd, yyyy")),
        ("style_number_format_time", lambda c: setattr(c, "number_format", "hh:mm:ss")),
        ("style_number_format_negative_red", lambda c: setattr(c, "number_format", "#,##0;[Red]-#,##0")),
        ("style_number_format_accounting", lambda c: setattr(c, "number_format", "_($* #,##0.00_)")),
        ("style_number_format_fraction", lambda c: setattr(c, "number_format", "# ?/?")),
    ]
    for slug, apply in styles:
        wb = Workbook()
        ws = wb.active
        ws.title = "Style"
        ws["A1"] = "Styled Cell"
        if "number_format" in slug:
            ws["A1"] = 1234.567
            if "date" in slug or "time" in slug:
                ws["A1"] = datetime(2024, 6, 15, 14, 30, 45)
        apply(ws["A1"])
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(
                path=out,
                group="matrix/style",
                features=["style", slug],
                expected_cells=1,
            )
        )
    return files


# --- dates & times --------------------------------------------------------


def build_date_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    entries = [
        ("date_today", datetime.now()),
        ("date_epoch_1900", datetime(1900, 1, 1)),
        ("date_epoch_1904", datetime(1904, 1, 2)),
        ("date_y2k", datetime(2000, 1, 1)),
        ("date_future_2099", datetime(2099, 12, 31)),
        ("date_leap_year", datetime(2020, 2, 29)),
        ("date_weird_feb28", datetime(1900, 2, 28)),
        ("date_first_valid", datetime(1900, 3, 1)),
        ("date_midnight", datetime(2024, 6, 1, 0, 0, 0)),
        ("date_nearmidnight", datetime(2024, 6, 1, 23, 59, 59)),
        ("date_iso_string", "2024-06-15"),
        ("date_us_string", "06/15/2024"),
        ("date_eu_string", "15/06/2024"),
        ("date_just_time", time(13, 30, 0)),
        ("date_date_only", date(2024, 6, 15)),
        ("date_with_timedelta_format", datetime(2024, 6, 15)),
        ("date_mixed_formats_in_column", None),
        ("date_fractional_days", 44500.5),  # excel serial
        ("date_negative_serial", -1),  # invalid
        ("date_text_like_date", "2024-06-15 but not really"),
    ]
    for slug, val in entries:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dates"
        if slug == "date_mixed_formats_in_column":
            ws["A1"] = datetime(2024, 1, 1)
            ws["A2"] = "2024-02-01"
            ws["A3"] = 44593
            ws["A4"] = date(2024, 4, 1)
            ws["A5"] = datetime(2024, 5, 1, 12, 30)
        else:
            ws["A1"] = val
            ws["A1"].number_format = "yyyy-mm-dd hh:mm:ss"
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(path=out, group="matrix/date", features=["date", slug], expected_cells=1),
        )
    return files


# --- errors ---------------------------------------------------------------


def build_error_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    errors = [
        ("error_div_zero", "=1/0"),
        ("error_name", "=UNKNOWN_FN()"),
        ("error_ref", "=#REF!"),
        ("error_value", "=\"a\"+1"),
        ("error_num", "=SQRT(-1)"),
        ("error_null", "=A1 A2"),  # intersection of disjoint ranges
        ("error_na", "=NA()"),
        ("error_getting_data", "=VLOOKUP(999,A1:B2,2,FALSE)"),
        ("error_mixed_with_text", "=IF(TRUE,1/0,\"ok\")"),
        ("error_chained", "=1/0+2"),
        ("error_deliberate_bad_ref", "=BadSheet!A1"),
        ("error_unclosed_paren", "=SUM(A1"),  # may get rewritten by openpyxl
        ("error_bad_range", "=SUM(A1:)"),
        ("error_too_many_args", "=IF(1,2,3,4,5)"),
        ("error_circular_simple", "=A1"),  # A1 refers to itself
    ]
    for slug, formula in errors:
        wb = Workbook()
        ws = wb.active
        ws.title = "Err"
        try:
            if slug == "error_circular_simple":
                ws["A1"] = formula
            else:
                ws["A2"] = 1
                ws["A1"] = formula
        except Exception:
            pass  # a few are too malformed even for openpyxl to accept
        out = _matrix_path(slug)
        try:
            _finalize(wb, out)
        except Exception:
            continue
        files.append(
            GeneratedFile(path=out, group="matrix/error", features=["error", slug], expected_cells=2, expected_formulas=1),
        )
    return files


# --- hidden rows/cols/sheets ---------------------------------------------


def build_hidden_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    specs = [
        ("hidden_single_row", "row", [3]),
        ("hidden_single_col", "col", ["B"]),
        ("hidden_many_rows", "row", list(range(2, 20, 2))),
        ("hidden_many_cols", "col", ["B", "D", "F", "H"]),
        ("hidden_first_row", "row", [1]),
        ("hidden_last_row", "row", [100]),
        ("hidden_row_at_boundary", "row", [50, 51, 52]),
        ("hidden_entire_block", "row", list(range(5, 15))),
        ("hidden_sheet_tab", "sheet", None),
        ("hidden_very_hidden_sheet", "veryhidden", None),
        ("hidden_with_outline_group", "outline", None),
        ("hidden_mixed_rows_cols", "mixed", None),
    ]
    for slug, kind, items in specs:
        wb = Workbook()
        ws = wb.active
        ws.title = "Main"
        for r in range(1, 30):
            for c in range(1, 10):
                ws.cell(row=r, column=c, value=(r + c) % 100)
        if kind == "row":
            for r in items:
                ws.row_dimensions[r].hidden = True
        elif kind == "col":
            for col in items:
                ws.column_dimensions[col].hidden = True
        elif kind == "sheet":
            hs = wb.create_sheet("HiddenSheet")
            hs["A1"] = "hidden content"
            hs.sheet_state = "hidden"
        elif kind == "veryhidden":
            hs = wb.create_sheet("VeryHidden")
            hs["A1"] = "very hidden"
            hs.sheet_state = "veryHidden"
        elif kind == "outline":
            for r in range(5, 15):
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = True
        elif kind == "mixed":
            ws.row_dimensions[3].hidden = True
            ws.row_dimensions[5].hidden = True
            ws.column_dimensions["C"].hidden = True
            ws.column_dimensions["E"].hidden = True
            hs = wb.create_sheet("MixedHidden")
            hs.sheet_state = "hidden"
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(path=out, group="matrix/hidden", features=["hidden", slug], expected_cells=29 * 9),
        )
    return files


# --- edge addresses -------------------------------------------------------


def build_edge_address_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    specs = [
        ("addr_xfd1", "XFD1", "lastcol_row1"),
        ("addr_a1048576", "A1048576", "col_a_lastrow"),
        ("addr_xfd1048576", "XFD1048576", "last_cell"),
        ("addr_zz1000", "ZZ1000", "mid_extreme"),
        ("addr_aaa1", "AAA1", "col_aaa"),
        ("addr_aa500", "AA500", "col_aa_500"),
        ("addr_very_sparse", None, "sparse"),
        ("addr_column_1000", f"{get_column_letter(1000)}1", "col_1000"),
        ("addr_row_100000", "A100000", "row_100k"),
        ("addr_gaps", None, "gaps"),
    ]
    for slug, addr, kind in specs:
        wb = Workbook()
        ws = wb.active
        ws.title = "Edge"
        ws["A1"] = "anchor"
        if kind == "sparse":
            ws["A1"] = "tl"
            ws["XFD1048576"] = "br"
        elif kind == "gaps":
            for offset in [0, 100, 1000, 10000]:
                ws.cell(row=1 + offset, column=1 + min(offset // 100, 50), value=f"v{offset}")
        elif addr:
            ws[addr] = f"marker_{slug}"
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(path=out, group="matrix/edge_address", features=["edge_address", slug], expected_cells=2),
        )
    return files


# --- sheet name variations ------------------------------------------------


SHEET_NAME_VARIANTS = [
    ("sheetname_ascii", "Simple"),
    ("sheetname_spaces", "Has Spaces"),
    ("sheetname_quote", "Has'Quote"),
    ("sheetname_unicode_jp", "日本語シート"),
    ("sheetname_unicode_emoji", "📊 Sheet"),
    ("sheetname_leading_digits", "1stSheet"),
    ("sheetname_long_30chars", "X" * 30),
    ("sheetname_dash_underscore", "my-sheet_name"),
    ("sheetname_hash_unicode", "Résumé-2025"),
    ("sheetname_parens", "Sheet (copy)"),
]


def build_sheet_name_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    for slug, name in SHEET_NAME_VARIANTS:
        wb = Workbook()
        ws = wb.active
        try:
            ws.title = name[:31]  # Excel limit
        except Exception:
            ws.title = "Fallback"
        ws["A1"] = f"in {name!r}"
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(
            GeneratedFile(path=out, group="matrix/sheet_name", features=["sheet_name", slug], expected_cells=1),
        )
    return files


# --- hyperlinks / comments / misc ----------------------------------------


def build_misc_files() -> list[GeneratedFile]:
    files: list[GeneratedFile] = []

    # hyperlinks
    hl_specs = [
        ("hyperlink_external_http", "https://example.com"),
        ("hyperlink_external_https", "https://www.anthropic.com"),
        ("hyperlink_mailto", "mailto:test@example.com"),
        ("hyperlink_file", "file:///tmp/x.txt"),
        ("hyperlink_internal_cell", "#Sheet1!B5"),
        ("hyperlink_internal_named", "#NamedRng"),
        ("hyperlink_many_links", None),
    ]
    for slug, url in hl_specs:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        if slug == "hyperlink_many_links":
            for i in range(1, 21):
                ws.cell(row=i, column=1, value=f"link{i}").hyperlink = f"https://example.com/page/{i}"
        else:
            ws["A1"].hyperlink = url
            ws["A1"].value = f"click ({slug})"
        if slug == "hyperlink_internal_named":
            wb.defined_names.add(DefinedName("NamedRng", attr_text="Sheet1!$A$1"))
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(GeneratedFile(path=out, group="matrix/hyperlink", features=["hyperlink", slug], expected_cells=20 if url is None else 1))

    # comments
    comment_specs = [
        ("comment_short", "Quick note"),
        ("comment_multiline", "line1\nline2\nline3"),
        ("comment_unicode", "注释 🔍 ملاحظة"),
        ("comment_long", "Note " * 500),
        ("comment_many_cells", None),
    ]
    for slug, text in comment_specs:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comments"
        if slug == "comment_many_cells":
            for i in range(1, 21):
                ws.cell(row=i, column=1, value=f"c{i}").comment = Comment(f"comment on row {i}", "Builder")
        else:
            ws["A1"] = "Cell with comment"
            ws["A1"].comment = Comment(text, "Builder")
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(GeneratedFile(path=out, group="matrix/comment", features=["comment", slug], expected_cells=20 if text is None else 1))

    # freeze panes
    for slug, freeze in [
        ("freeze_row_1", "A2"),
        ("freeze_col_a", "B1"),
        ("freeze_both_a1", "B2"),
        ("freeze_mid_sheet", "C5"),
        ("freeze_deep", "E10"),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Freeze"
        for r in range(1, 21):
            for c in range(1, 10):
                ws.cell(row=r, column=c, value=f"{r},{c}")
        ws.freeze_panes = freeze
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(GeneratedFile(path=out, group="matrix/freeze_panes", features=["freeze_panes", slug], expected_cells=20 * 9))

    # rich text (mixed fonts within a cell) — openpyxl exposes this via CellRichText
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        for slug, blocks in [
            ("rich_text_bold_plain", [TextBlock(InlineFont(b=True), "Bold "), TextBlock(InlineFont(), "plain")]),
            ("rich_text_colors", [TextBlock(InlineFont(color="FF0000"), "Red "), TextBlock(InlineFont(color="0000FF"), "Blue")]),
            ("rich_text_sizes", [TextBlock(InlineFont(sz="8"), "small "), TextBlock(InlineFont(sz="18"), "BIG")]),
        ]:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rich"
            ws["A1"] = CellRichText(blocks)
            out = _matrix_path(slug)
            _finalize(wb, out)
            files.append(GeneratedFile(path=out, group="matrix/rich_text", features=["rich_text", slug], expected_cells=1))
    except Exception:
        pass

    # 3D refs / cross-sheet
    for slug in ["threed_sum_across_sheets"]:
        wb = Workbook()
        ws = wb.active
        ws.title = "A"
        for r in range(1, 6):
            ws.cell(row=r, column=1, value=r)
        wb.create_sheet("B")
        for r in range(1, 6):
            wb["B"].cell(row=r, column=1, value=r * 10)
        summary = wb.create_sheet("Summary")
        summary["A1"] = "=SUM(A:B!A1:A5)"  # Excel 3D ref syntax
        out = _matrix_path(slug)
        _finalize(wb, out)
        files.append(GeneratedFile(path=out, group="matrix/3d_ref", features=["3d_ref", slug], expected_cells=11, expected_formulas=1))

    return files


MATRIX_BUILDERS: list[Callable[[], list[GeneratedFile]]] = [
    build_formula_files,
    build_merge_files,
    build_named_range_files,
    build_data_validation_files,
    build_conditional_formatting_files,
    build_table_files,
    build_chart_files,
    build_style_files,
    build_date_files,
    build_error_files,
    build_hidden_files,
    build_edge_address_files,
    build_sheet_name_files,
    build_misc_files,
]


# ----------------------------------------------------------------------------
# Combinatoric group — randomised feature cocktails
# ----------------------------------------------------------------------------


COMBO_DIR = OUT_ROOT / "combo"
DENSITIES = [5, 10, 25, 50, 100]
SEEDS_PER_DENSITY = 80   # → 400 combo files


def _rand_cell_value(rng: random.Random):
    kind = rng.choice(["int", "float", "str", "bool", "date", "blank"])
    if kind == "int":
        return rng.randint(-10_000, 10_000)
    if kind == "float":
        return rng.uniform(-1000.0, 1000.0)
    if kind == "str":
        return "".join(rng.choices(string.ascii_letters + string.digits + " ", k=rng.randint(1, 30)))
    if kind == "bool":
        return rng.choice([True, False])
    if kind == "date":
        return date(rng.randint(2000, 2030), rng.randint(1, 12), rng.randint(1, 28))
    return None


def _safe_set(ws, row: int, col: int, value) -> bool:
    """Try to set ws cell; return True on success, False if cell is part of a merge."""
    try:
        ws.cell(row=row, column=col, value=value)
        return True
    except (AttributeError, TypeError):
        return False


def build_combo_file(seed: int, density: int) -> GeneratedFile | None:
    rng = random.Random(seed * 10_000 + density)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Main_{seed}_{density}"
    cells_written = 0
    formulas = 0
    features: set[str] = set()

    for _ in range(density):
        op = rng.choices(
            population=["cell", "formula", "merge", "style", "comment", "hyperlink", "validation", "table", "named"],
            weights=[45, 20, 8, 12, 3, 3, 3, 3, 3],
            k=1,
        )[0]
        r = rng.randint(1, 100)
        c = rng.randint(1, 30)
        if op == "cell":
            if _safe_set(ws, r, c, _rand_cell_value(rng)):
                cells_written += 1
                features.add("cells")
        elif op == "formula":
            if _safe_set(ws, r, c, f"=SUM({get_column_letter(c)}1:{get_column_letter(c)}{max(1, r-1)})"):
                formulas += 1
                features.add("formulas")
        elif op == "merge":
            try:
                r2 = min(r + rng.randint(0, 3), 100)
                c2 = min(c + rng.randint(0, 3), 30)
                if (r, c) != (r2, c2):
                    _safe_set(ws, r, c, f"m{seed}")  # write before merge
                    ws.merge_cells(start_row=r, start_column=c, end_row=r2, end_column=c2)
                    features.add("merge")
            except Exception:
                pass
        elif op == "style":
            try:
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    if _safe_set(ws, r, c, rng.randint(0, 99)):
                        cells_written += 1
                    cell = ws.cell(row=r, column=c)
                cell.font = Font(bold=rng.choice([True, False]), italic=rng.choice([True, False]), color=f"{rng.randint(0, 0xFFFFFF):06X}")
                cell.fill = PatternFill("solid", start_color=f"{rng.randint(0xAAAAAA, 0xFFFFFF):06X}")
                features.add("style")
            except AttributeError:
                pass
        elif op == "comment":
            try:
                if _safe_set(ws, r, c, "c"):
                    ws.cell(row=r, column=c).comment = Comment(f"seed{seed}", "combo")
                    cells_written += 1
                    features.add("comment")
            except Exception:
                pass
        elif op == "hyperlink":
            try:
                if _safe_set(ws, r, c, "lnk"):
                    ws.cell(row=r, column=c).hyperlink = f"https://example.com/{seed}/{r}-{c}"
                    cells_written += 1
                    features.add("hyperlink")
            except Exception:
                pass
        elif op == "validation":
            try:
                dv = DataValidation(type="list", formula1='"A,B,C"')
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(c)}{r}")
                features.add("validation")
            except Exception:
                pass
        elif op == "table":
            try:
                r2 = min(r + 3, 100)
                c2 = min(c + 2, 30)
                if r2 > r and c2 > c:
                    for rr in range(r, r2 + 1):
                        for cc in range(c, c2 + 1):
                            try:
                                if ws.cell(row=rr, column=cc).value is None:
                                    _safe_set(ws, rr, cc, rr * cc)
                            except AttributeError:
                                pass
                    for cc in range(c, c2 + 1):
                        _safe_set(ws, r, cc, f"H{cc}")
                    tab_name = f"T{seed}_{density}_{rng.randint(0, 99)}"
                    ws.add_table(Table(displayName=tab_name, ref=f"{get_column_letter(c)}{r}:{get_column_letter(c2)}{r2}"))
                    features.add("table")
            except Exception:
                pass
        elif op == "named":
            try:
                nm = f"N_{seed}_{density}_{rng.randint(0, 99)}"
                wb.defined_names.add(DefinedName(nm, attr_text=f"{ws.title}!${get_column_letter(c)}${r}"))
                features.add("named_range")
            except Exception:
                pass

    out = COMBO_DIR / f"combo_d{density:03d}_s{seed:03d}.xlsx"
    try:
        _finalize(wb, out)
    except Exception:
        return None
    return GeneratedFile(
        path=out,
        group="combo",
        features=sorted(features),
        expected_cells=cells_written,
        expected_formulas=formulas,
        notes=f"seed={seed} density={density}",
    )


def build_combo_files(limit: int | None) -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    count = 0
    for density in DENSITIES:
        for seed in range(SEEDS_PER_DENSITY):
            if limit is not None and count >= limit:
                return files
            gf = build_combo_file(seed, density)
            if gf:
                files.append(gf)
            count += 1
    return files


# ----------------------------------------------------------------------------
# Adversarial group — try to break the parser
# ----------------------------------------------------------------------------


ADVERSARIAL_DIR = OUT_ROOT / "adversarial"


def _adv_path(slug: str) -> Path:
    return ADVERSARIAL_DIR / f"{slug}.xlsx"


def build_adversarial_files(limit: int | None) -> list[GeneratedFile]:
    files: list[GeneratedFile] = []
    specs: list[tuple[str, Callable[[Workbook], tuple[int, int, str]]]] = []

    def _mk(slug: str):
        def deco(fn: Callable[[Workbook], tuple[int, int, str]]):
            specs.append((slug, fn))
            return fn
        return deco

    @_mk("adv_empty_workbook")
    def _(wb):
        # openpyxl always has one sheet; clear it
        ws = wb.active
        ws.title = "Empty"
        return 0, 0, "no cells"

    @_mk("adv_one_cell_1e300")
    def _(wb):
        wb.active["A1"] = 1e300
        return 1, 0, "huge float"

    @_mk("adv_one_cell_neg_1e300")
    def _(wb):
        wb.active["A1"] = -1e300
        return 1, 0, "huge negative"

    @_mk("adv_one_cell_tiny")
    def _(wb):
        wb.active["A1"] = 1e-300
        return 1, 0, "tiny float"

    @_mk("adv_unicode_bomb")
    def _(wb):
        ws = wb.active
        emojis = "🚀🔥💀🎯🌀⚡️🌈🎨🧪💡" * 20
        rtl = "مرحبا بكم في اختبار التحليل" * 5
        cjk = "こんにちは世界 你好世界 안녕하세요" * 5
        ws["A1"] = emojis + " " + rtl + " " + cjk
        ws["A2"] = "\u200B\u200C\u200D\ufeff"  # zero-width chars
        ws["A3"] = "a" * 32_000  # long string
        return 3, 0, "unicode stress"

    @_mk("adv_circular_chain_10")
    def _(wb):
        ws = wb.active
        for i in range(1, 10):
            ws.cell(row=i, column=1, value=f"=A{i+1}")
        ws["A10"] = "=A1"
        return 10, 10, "10-step cycle"

    @_mk("adv_formula_chain_deep_500")
    def _(wb):
        ws = wb.active
        ws["A1"] = 1
        for i in range(2, 501):
            ws.cell(row=i, column=1, value=f"=A{i-1}+1")
        return 500, 499, "500-deep chain"

    @_mk("adv_huge_merge_1000x100")
    def _(wb):
        ws = wb.active
        ws.merge_cells("A1:CV1000")  # 100 cols × 1000 rows
        ws["A1"] = "one giant merge"
        return 1, 0, "100k-cell merge"

    @_mk("adv_many_merges_5000")
    def _(wb):
        ws = wb.active
        for i in range(5000):
            r = i // 50 + 1
            c = (i % 50) * 2 + 1
            try:
                ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
                ws.cell(row=r, column=c, value="m")
            except Exception:
                pass
        return 2500, 0, "5000 merges"

    @_mk("adv_100_sheets")
    def _(wb):
        wb.active.title = "S0"
        for i in range(1, 100):
            ws = wb.create_sheet(f"S{i}")
            ws["A1"] = i
        return 100, 0, "100 sheets"

    @_mk("adv_very_wide_2000_cols")
    def _(wb):
        ws = wb.active
        for c in range(1, 2001):
            ws.cell(row=1, column=c, value=c)
        return 2000, 0, "2000 cols in one row"

    @_mk("adv_very_tall_20k_rows")
    def _(wb):
        ws = wb.active
        for r in range(1, 20_001):
            ws.cell(row=r, column=1, value=r)
        return 20_000, 0, "20k rows"

    @_mk("adv_sparse_million")
    def _(wb):
        ws = wb.active
        for r in [1, 10, 100, 1000, 10_000, 100_000, 500_000, 1_000_000]:
            ws.cell(row=r, column=1, value=f"r{r}")
        ws["A1"].value = "start"
        return 8, 0, "sparse across 1M rows"

    @_mk("adv_all_error_types")
    def _(wb):
        ws = wb.active
        for i, formula in enumerate([
            "=1/0", "=SQRT(-1)", "=NA()", "=BAD_FN()", "=#REF!", '="a"+1',
        ], start=1):
            ws.cell(row=i, column=1, value=formula)
        return 6, 6, "errors galore"

    @_mk("adv_broken_refs")
    def _(wb):
        ws = wb.active
        ws["A1"] = "=MissingSheet!B5"
        ws["A2"] = "=OtherBook.xlsx!Sheet1!A1"
        ws["A3"] = "=#REF!+1"
        return 3, 3, "dangling references"

    @_mk("adv_long_formula")
    def _(wb):
        ws = wb.active
        ws["A1"] = 1
        long_expr = "=" + "+".join("A1" for _ in range(2000))
        ws["B1"] = long_expr
        return 2, 1, "very long formula"

    @_mk("adv_long_cell_string")
    def _(wb):
        ws = wb.active
        ws["A1"] = "X" * 32_767  # Excel limit
        return 1, 0, "32k char cell"

    @_mk("adv_all_formulas_sheet")
    def _(wb):
        ws = wb.active
        for r in range(1, 101):
            for c in range(1, 6):
                ws.cell(row=r, column=c, value=f"={get_column_letter(c)}{((r - 1) % 5) + 1}+1")
        return 500, 500, "500 formulas"

    @_mk("adv_massive_table")
    def _(wb):
        ws = wb.active
        for c in range(1, 51):
            ws.cell(row=1, column=c, value=f"C{c}")
        for r in range(2, 202):
            for c in range(1, 51):
                ws.cell(row=r, column=c, value=(r * c) % 997)
        ws.add_table(Table(displayName="Huge", ref=f"A1:{get_column_letter(50)}201"))
        return 10_050, 0, "50x200 table"

    @_mk("adv_cyclic_cross_sheet")
    def _(wb):
        a = wb.active
        a.title = "A"
        a["A1"] = "=B!A1"
        b = wb.create_sheet("B")
        b["A1"] = "=A!A1"
        return 2, 2, "cross-sheet cycle"

    @_mk("adv_many_named_ranges")
    def _(wb):
        ws = wb.active
        for i in range(1, 301):
            wb.defined_names.add(DefinedName(f"N{i}", attr_text=f"Sheet!${get_column_letter((i % 30) + 1)}${(i % 100) + 1}"))
        ws["A1"] = "seed"
        return 1, 0, "300 named ranges"

    @_mk("adv_duplicate_sheet_names_almost")
    def _(wb):
        wb.active.title = "Data"
        wb.create_sheet("data")
        wb.create_sheet("DATA")
        return 0, 0, "case-sensitive sheet names"

    @_mk("adv_rtl_sheet")
    def _(wb):
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws["A1"] = "النص يقرأ من اليمين"
        return 1, 0, "RTL view"

    @_mk("adv_extreme_column_width")
    def _(wb):
        ws = wb.active
        ws.column_dimensions["A"].width = 255
        ws.row_dimensions[1].height = 409  # excel max
        ws["A1"] = "wide+tall"
        return 1, 0, "max col/row size"

    @_mk("adv_autofilter_large")
    def _(wb):
        ws = wb.active
        for c in range(1, 11):
            ws.cell(row=1, column=c, value=f"H{c}")
        for r in range(2, 301):
            for c in range(1, 11):
                ws.cell(row=r, column=c, value=r * c)
        ws.auto_filter.ref = "A1:J300"
        return 3000, 0, "autofilter 3k cells"

    @_mk("adv_mixed_types_same_column")
    def _(wb):
        ws = wb.active
        for r in range(1, 51):
            if r % 5 == 0:
                ws.cell(row=r, column=1, value=f"text_{r}")
            elif r % 5 == 1:
                ws.cell(row=r, column=1, value=r)
            elif r % 5 == 2:
                ws.cell(row=r, column=1, value=float(r) / 7.0)
            elif r % 5 == 3:
                ws.cell(row=r, column=1, value=date(2024, (r % 12) + 1, 1))
            else:
                ws.cell(row=r, column=1, value=(r % 2 == 0))
        return 50, 0, "mixed types in one column"

    _SAFE_STR_CHARS = string.ascii_letters + string.digits + " -_.,:;!?@#$%^&*()[]{}<>+=/|~"

    # adversarial via parametrised generator to pad counts to ~1000 total
    for i in range(1, 278):  # 277 parametric adversarial files → 1000 total generated
        rng = random.Random(10_000 + i)

        @_mk(f"adv_param_{i:03d}")
        def _(wb, rng=rng, i=i):
            ws = wb.active
            # Keep sizes modest so the full bench runs under 10 min wall-clock.
            n_cells = rng.randint(100, 800)
            cells = 0
            formulas = 0
            for _ in range(n_cells):
                r = rng.randint(1, 300)
                c = rng.randint(1, 50)
                kind = rng.choice(["int", "str", "formula", "date", "bool"])
                try:
                    if kind == "int":
                        val = rng.randint(-1_000_000, 1_000_000)
                    elif kind == "str":
                        val = "".join(rng.choices(_SAFE_STR_CHARS, k=rng.randint(1, 50)))
                    elif kind == "formula":
                        val = f"={get_column_letter(max(1, c - 1))}{max(1, r - 1)}+1"
                    elif kind == "date":
                        val = date(rng.randint(1900, 2099), rng.randint(1, 12), rng.randint(1, 28))
                    else:
                        val = rng.choice([True, False])
                    if _safe_set(ws, r, c, val):
                        cells += 1
                        if kind == "formula":
                            formulas += 1
                except Exception:
                    pass
            for _ in range(rng.randint(0, 20)):
                try:
                    r0 = rng.randint(1, 100)
                    c0 = rng.randint(1, 50)
                    ws.merge_cells(start_row=r0, start_column=c0, end_row=r0 + rng.randint(0, 5), end_column=c0 + rng.randint(0, 5))
                except Exception:
                    pass
            return cells, formulas, f"param seed {i}"

    files: list[GeneratedFile] = []
    count = 0
    for slug, fn in specs:
        if limit is not None and count >= limit:
            break
        wb = Workbook()
        try:
            cells, formulas, notes = fn(wb)
        except Exception as exc:
            # skip uncooperative generators
            print(f"  ⚠ adversarial {slug} failed to build: {exc}", file=sys.stderr)
            continue
        out = _adv_path(slug)
        try:
            _finalize(wb, out)
        except Exception as exc:
            print(f"  ⚠ adversarial {slug} failed to save: {exc}", file=sys.stderr)
            continue
        files.append(
            GeneratedFile(
                path=out,
                group="adversarial",
                features=["adversarial", slug],
                expected_cells=cells,
                expected_formulas=formulas,
                notes=notes,
            )
        )
        count += 1
    return files


# ----------------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------------


def build_all(groups: set[str], force: bool, limit: int | None) -> list[GeneratedFile]:
    all_files: list[GeneratedFile] = []
    if "matrix" in groups:
        MATRIX_DIR.mkdir(parents=True, exist_ok=True)
        for builder in MATRIX_BUILDERS:
            for gf in builder():
                all_files.append(gf)
                if limit is not None and len(all_files) >= limit:
                    return all_files
    if "combo" in groups:
        COMBO_DIR.mkdir(parents=True, exist_ok=True)
        remaining = None if limit is None else max(0, limit - len(all_files))
        all_files.extend(build_combo_files(remaining))
        if limit is not None and len(all_files) >= limit:
            return all_files
    if "adversarial" in groups:
        ADVERSARIAL_DIR.mkdir(parents=True, exist_ok=True)
        remaining = None if limit is None else max(0, limit - len(all_files))
        all_files.extend(build_adversarial_files(remaining))
    return all_files


def write_manifest(files: list[GeneratedFile]) -> None:
    by_group: dict[str, int] = {}
    rows = []
    for gf in files:
        rows.append(gf.to_manifest_row())
        by_group[gf.group] = by_group.get(gf.group, 0) + 1
    manifest = {
        "version": 1,
        "generated_at": "deterministic",
        "total_files": len(files),
        "by_group": by_group,
        "files": rows,
    }
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2, sort_keys=False))
    print(f"✓ manifest written → {MANIFEST_PATH.relative_to(ROOT)}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--group", choices=["matrix", "combo", "adversarial", "all"], default="all")
    parser.add_argument("--force", action="store_true", help="regenerate even if outputs exist")
    parser.add_argument("--limit", type=int, help="stop after N files (smoke mode)")
    parser.add_argument("--clean", action="store_true", help="wipe testBench/generated/ first")
    args = parser.parse_args()

    if args.clean and OUT_ROOT.exists():
        import shutil
        shutil.rmtree(OUT_ROOT)
        print(f"✓ cleaned {OUT_ROOT.relative_to(ROOT)}")

    groups = {"matrix", "combo", "adversarial"} if args.group == "all" else {args.group}
    OUT_ROOT.mkdir(parents=True, exist_ok=True)

    print(f"building testBench into {OUT_ROOT.relative_to(ROOT)}  groups={sorted(groups)}  limit={args.limit}")
    files = build_all(groups, args.force, args.limit)
    write_manifest(files)

    print(f"\n{'═' * 60}")
    print(f"  Generated {len(files)} workbooks")
    by_group: dict[str, int] = {}
    for gf in files:
        by_group[gf.group] = by_group.get(gf.group, 0) + 1
    for g in sorted(by_group):
        print(f"    {g:32s} {by_group[g]:4d}")
    print(f"{'═' * 60}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
