"""
Chunk-quality benchmark on SpreadsheetBench.

Uses ``dataset.json`` (instruction → answer_sheet!answer_position) as
ground truth — each of the 912 instances provides a natural-language
question and the cell range where the answer lives.

For each parser × instance, we:
  1. Parse the input.xlsx and obtain a list of (chunk_id, sheet, range, text)
  2. Embed all chunks + the instruction with sentence-transformers
  3. Rank chunks by cosine similarity to the instruction
  4. Check: does the top-k include a chunk that overlaps the ground-truth
     ``answer_sheet!answer_position`` range?
  5. Score table-integrity: how many chunks span the answer region?
     (1 = clean, >1 = the answer table was fragmented across chunks)

Output: per-instance NDJSON + aggregate JSON with recall@1/3/5 +
fragmentation distribution per parser.

Usage:
    python scripts/eval_retrieval.py \\
        --corpus data/corpora/spreadsheetbench/all_data_912_v0.1 \\
        --out tests/benchmarks/reports/retrieval \\
        --parsers ks,docling \\
        [--sample 100]
"""

from __future__ import annotations

import argparse
import contextlib
import json
import re
import signal
import sys
import time
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parent.parent
# Keep ``import scripts.X`` style imports working when invoked as
# ``python scripts/eval_retrieval.py``. We no longer need ``src`` on the
# path — excel_parser is a properly-installed package now.
sys.path.insert(0, str(REPO_ROOT))


def _normalize_value_for_match(s: str) -> set[str]:
    """Produce a set of equivalent string forms of ``s`` for substring
    matching against chunk text.

    Different parsers emit the same datum in different shapes:
      - 1272 → "1272" (ks raw), "1,272.00" (Excel display), "1272.0" (Python repr)
      - 2021-09-01 → "2021-09-01", "2021-09-01 00:00:00", "9/1/2021"
      - 0.06 → "0.06", "6%", "6.0%"

    We can't predict every parser's choice, so we generate every plausible
    rendering we'd accept as a hit, then the caller checks if ANY appears
    in the chunk text. This makes the metric fair across formatting
    conventions without giving any parser undeserved credit.
    """
    s = s.strip()
    if not s:
        return set()
    forms = {s}

    # Numeric? Strip commas + Excel currency/percent decorations, normalize.
    raw = s.replace(",", "").lstrip("$€£¥").rstrip("%")
    try:
        f = float(raw)
        # ``inf``/``nan`` can sneak through if a cell stores "Infinity";
        # they have no useful canonical form for substring match.
        import math
        if math.isfinite(f):
            if f == int(f) and abs(f) < 1e16:
                forms.add(str(int(f)))
                forms.add(f"{int(f)}.0")
            else:
                forms.add(f"{f:.10g}")
                forms.add(str(f))
    except ValueError:
        pass

    # Date with time-component? Add the bare date form.
    # Common shapes: "2021-09-01 00:00:00", "2021-09-01T00:00:00"
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        forms.add(s[:10])

    # Booleans: Excel/answer.xlsx surfaces these as the python ``True`` /
    # ``False`` literals; parsers render uppercase ``TRUE`` / ``FALSE``.
    low = s.lower()
    if low in {"true", "false"}:
        forms.add(low.upper())
        forms.add(low.capitalize())

    return forms


def _matches_chunk_text(values: list[str], chunk_text: str) -> bool:
    """True if any normalized form of any expected value appears in chunk_text."""
    if not values or not chunk_text:
        return False
    for v in values:
        for form in _normalize_value_for_match(v):
            if form and len(form) >= 2 and form in chunk_text:
                return True
    return False


class _TimeoutError(Exception):
    pass


@contextlib.contextmanager
def _alarm_timeout(seconds: float):
    """SIGALRM-based wall-clock timeout. UNIX-only.

    Docling occasionally hangs indefinitely on pathological workbooks
    (large layout / table-recognition inference loops). Without a timeout
    a single bad file blocks the whole 912-instance run. SIGALRM is a
    blunt tool but adequate here — we always run from the main thread
    in this script.
    """
    if seconds <= 0:
        yield
        return

    def _handler(signum, frame):
        raise _TimeoutError(f"parser exceeded {seconds:.0f}s")

    old = signal.signal(signal.SIGALRM, _handler)
    signal.setitimer(signal.ITIMER_REAL, seconds)
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, old)

A1_RE = re.compile(r"^([A-Z]+)(\d+)$", re.IGNORECASE)
RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", re.IGNORECASE)
# Match optional `Sheet!` or `'Sheet'!` prefix + an A1 range.
# SpreadsheetBench is sloppy: sometimes only one quote ("Sheet1'!A1:B2"),
# sometimes none ("Sheet1!A1:B2"). We accept all forms.
SHEET_RANGE_RE = re.compile(
    r"""(?P<quote>'?)             # optional opening quote
        (?P<sheet>[^'!,]+?)        # sheet name (non-greedy, no '!' or ',')
        (?P=quote)                 # matching closing quote (may be empty)
        !                          # required sheet separator
        (?P<range>[A-Z]+\d+(?::[A-Z]+\d+)?)
    """,
    re.IGNORECASE | re.VERBOSE,
)


# ────────────────────────────────────────────────────────────── chunk record


@dataclass
class Chunk:
    """Parser-agnostic chunk for retrieval scoring."""

    parser: str
    sheet: str | None
    top_left: tuple[int, int] | None    # (row, col) 1-indexed
    bottom_right: tuple[int, int] | None
    text: str
    chunk_id: str = ""

    def overlaps(self, sheet: str, range_box: tuple[int, int, int, int]) -> bool:
        """True if this chunk's range overlaps the given (r0,c0,r1,c1) on `sheet`.

        When ``sheet`` is unspecified (None/empty) the match is geometry-only:
        ~62% of SpreadsheetBench instances omit the sheet name (single-sheet
        workbooks), and a real chunk's sheet name can never equal "". Requiring
        equality there silently denied geometric credit to correct chunks.
        """
        if sheet and self.sheet is not None and self.sheet != sheet:
            return False
        if self.top_left is None or self.bottom_right is None:
            # Parser didn't surface a range — fall back to text match
            return False
        r0, c0, r1, c1 = range_box
        cr0, cc0 = self.top_left
        cr1, cc1 = self.bottom_right
        return not (cr1 < r0 or cr0 > r1 or cc1 < c0 or cc0 > c1)


# ────────────────────────────────────────────────────────────── A1 helpers


def col_letter_to_number(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def parse_a1(a1: str) -> tuple[int, int] | None:
    m = A1_RE.match(a1.strip())
    if not m:
        return None
    return (int(m.group(2)), col_letter_to_number(m.group(1)))


def parse_range(rng: str) -> tuple[int, int, int, int] | None:
    """Parse 'A1:D10' → (r0, c0, r1, c1). Single cell 'A1' → (1,1,1,1).

    Tolerates whitespace around the colon ('A1: D10') and the fullwidth
    colon ('A1：D10') — both shapes appear in SpreadsheetBench
    `data_position` / `answer_position` strings.
    """
    rng = rng.strip().replace("：", ":")
    # Collapse any whitespace around the colon: 'A1 : D10' -> 'A1:D10'
    rng = re.sub(r"\s*:\s*", ":", rng)
    m = RANGE_RE.match(rng)
    if m:
        r0 = int(m.group(2))
        c0 = col_letter_to_number(m.group(1))
        r1 = int(m.group(4))
        c1 = col_letter_to_number(m.group(3))
        return (min(r0, r1), min(c0, c1), max(r0, r1), max(c0, c1))
    p = parse_a1(rng)
    if p:
        return (p[0], p[1], p[0], p[1])
    return None


def parse_position_spec(
    spec: str, default_sheet: str | None,
) -> list[tuple[str | None, tuple[int, int, int, int]]]:
    """Parse SpreadsheetBench's free-form `data_position` / `answer_position`.

    Examples that appear in the wild:
      "A1:D10"                            → [(default_sheet, A1:D10)]
      "'Sheet1'!A1:D10"                   → [("Sheet1", A1:D10)]
      "Sheet1'!A1:D10"                    → [("Sheet1", A1:D10)]   (typo)
      "'Sheet1!'A1:D10"                   → [("Sheet1", A1:D10)]   (typo)
      "'A'!B2:C3,'B'!D4"                  → [("A", B2:C3), ("B", D4:D4)]
      "Sheet1!A1:B2,Sheet2!C3:D4"         → [("Sheet1",…), ("Sheet2",…)]
      "G12：J15"                          → [(default_sheet, G12:J15)]  (CJK colon)

    Returns a list of (sheet_or_None, range_box). Empty list if unparseable.
    """
    if not spec:
        return []
    # Fullwidth colon `：` (U+FF1A) appears in a handful of CJK-locale dataset
    # entries where Excel/WPS substituted it for the ASCII `:`.
    spec = spec.strip().replace("：", ":")

    out: list[tuple[str | None, tuple[int, int, int, int]]] = []

    # First try the strict regex — handles `'Sheet'!A1:B2` and bare
    # `Sheet!A1:B2` cleanly. If it covers the whole spec we're done.
    for m in SHEET_RANGE_RE.finditer(spec):
        sheet = m.group("sheet").strip().strip("'")
        rng = parse_range(m.group("range"))
        if rng is not None:
            out.append((sheet or default_sheet, rng))

    if out:
        return out

    # Fallback for malformed quoted forms the strict regex rejects:
    #   "Sheet1'!A1"        — stray closing quote, no opening
    #   "'Sheet1!'A1"       — quote placed AFTER the `!` separator
    # Split on `,` for multi-region, then on `!` for sheet/range, then
    # strip stray apostrophes adjacent to either side of the separator.
    for piece in spec.split(","):
        piece = piece.strip()
        if not piece:
            continue
        if "!" in piece:
            sheet_part, _, range_part = piece.partition("!")
            sheet = sheet_part.strip().strip("'").strip() or None
            range_str = range_part.strip().strip("'").strip()
        else:
            sheet = None
            range_str = piece.strip("'").strip()
        rng = parse_range(range_str)
        if rng is not None:
            out.append((sheet or default_sheet, rng))
    return out


# ────────────────────────────────────────────────────────────── ks adapter


def extract_chunks_ks(path: Path) -> list[Chunk]:
    from excel_parser.pipeline import parse_workbook

    result = parse_workbook(path=str(path))
    out: list[Chunk] = []
    for c in result.chunks:
        tl = parse_a1(c.top_left_cell) if c.top_left_cell else None
        br = parse_a1(c.bottom_right_cell) if c.bottom_right_cell else None
        out.append(Chunk(
            parser="excel-parser",
            sheet=c.sheet_name,
            top_left=tl,
            bottom_right=br,
            text=c.render_text or "",
            chunk_id=c.chunk_id or "",
        ))
    return out


# ────────────────────────────────────────────────────────────── docling adapter


# Docling is run in a long-lived child subprocess so we can hard-kill it
# on hangs without paying the model-load cost (~5–10s) per file. SIGALRM
# doesn't work — docling's table-recognition path is in PyTorch C-land
# and holds the GIL through tight inference loops, ignoring Python signal
# handlers. A separate process is the only reliable timeout boundary.
#
# Protocol (one persistent worker per script run):
#   parent -> worker (stdin):  one line  {"path":"..."}
#   worker -> parent (stdout): one line  [{"text":"...","id":"..."}, ...]
# If the worker doesn't respond within ``timeout_s``, we SIGKILL it and
# the next call respawns a fresh one (re-paying the model-load cost).

_DOCLING_WORKER_SCRIPT = r"""
import json, sys
from docling.document_converter import DocumentConverter

conv = DocumentConverter()
sys.stdout.write(json.dumps({"event":"ready"}) + "\n")
sys.stdout.flush()

for line in sys.stdin:
    line = line.strip()
    if not line:
        continue
    try:
        msg = json.loads(line)
        path = msg["path"]
        result = conv.convert(path)
        doc = result.document
        chunks = []
        for i, table in enumerate(doc.tables):
            try:
                md = table.export_to_dataframe(doc).to_markdown(index=False)
            except Exception:
                try:
                    md = table.export_to_html(doc)
                except Exception:
                    md = ""
            chunks.append({"text": md, "id": f"table-{i}"})
        for j, txt in enumerate(doc.texts):
            t = (txt.text or "").strip()
            if t:
                chunks.append({"text": t, "id": f"text-{j}"})
        sys.stdout.write(json.dumps(chunks) + "\n")
    except Exception as exc:
        sys.stdout.write(json.dumps({"error": f"{type(exc).__name__}: {exc}"}) + "\n")
    sys.stdout.flush()
"""


class _DoclingWorker:
    """Persistent docling subprocess with hard-kill timeout."""

    def __init__(self, timeout_s: float = 60.0):
        self.timeout_s = timeout_s
        self._proc = None
        self._worker_path = None

    def _ensure_alive(self):
        import subprocess
        import tempfile

        if self._proc is not None and self._proc.poll() is None:
            return
        if self._worker_path is None:
            self._worker_path = Path(tempfile.gettempdir()) / "_eval_docling_worker.py"
            self._worker_path.write_text(_DOCLING_WORKER_SCRIPT)

        self._proc = subprocess.Popen(
            [sys.executable, str(self._worker_path)],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
            bufsize=1,
        )
        # Wait for handshake
        ready = self._read_with_timeout(120.0)  # model load can take long on cold start
        if ready is None or '"ready"' not in ready:
            self._kill()
            raise RuntimeError("docling worker handshake failed")

    def _kill(self):
        if self._proc is not None:
            try:
                self._proc.kill()
                self._proc.wait(timeout=2)
            except Exception:
                pass
            self._proc = None

    def stop(self):
        self._kill()

    def _read_with_timeout(self, deadline_s: float) -> str | None:
        import select
        if self._proc is None or self._proc.stdout is None:
            return None
        fd = self._proc.stdout.fileno()
        import time as _t
        deadline = _t.monotonic() + deadline_s
        buf = b""
        while True:
            remaining = deadline - _t.monotonic()
            if remaining <= 0:
                return None
            r, _, _ = select.select([fd], [], [], remaining)
            if not r:
                return None
            import os as _os
            chunk = _os.read(fd, 8192)
            if not chunk:
                return None
            buf += chunk
            if b"\n" in buf:
                line, _, _ = buf.partition(b"\n")
                return line.decode("utf-8", errors="replace")

    def extract(self, path: Path) -> list[Chunk]:
        self._ensure_alive()
        assert self._proc is not None and self._proc.stdin is not None
        self._proc.stdin.write(json.dumps({"path": str(path)}) + "\n")
        self._proc.stdin.flush()
        line = self._read_with_timeout(self.timeout_s)
        if line is None:
            self._kill()  # respawn on next call
            raise RuntimeError(f"docling timeout (>{self.timeout_s:.0f}s)")
        raw = json.loads(line)
        if isinstance(raw, dict) and "error" in raw:
            raise RuntimeError(f"docling worker error: {raw['error']}")
        return [
            Chunk(parser="docling", sheet=None, top_left=None, bottom_right=None,
                  text=c["text"], chunk_id=c["id"])
            for c in raw
        ]


_DOCLING_WORKER: _DoclingWorker | None = None


def extract_chunks_docling(path: Path) -> list[Chunk]:
    """Use a persistent docling subprocess. Hard-killed if a file hangs."""
    global _DOCLING_WORKER
    if _DOCLING_WORKER is None:
        _DOCLING_WORKER = _DoclingWorker(timeout_s=60.0)
    return _DOCLING_WORKER.extract(path)


# ────────────────────────────────────────────────────────────── retrieval scoring


@dataclass
class InstanceResult:
    instance_id: str
    parser: str
    n_chunks: int
    parse_ms: float
    data_position: str
    answer_position: str
    data_regions: int                         # parsed regions in data_position
    chunks_overlapping_data: int              # table-integrity: <=1 is good
    rank_of_first_overlap: int | None         # by similarity, 1-indexed
    rank_of_text_match: int | None            # fallback: answer-value substring match
    error: str | None = None
    extra: dict[str, Any] = field(default_factory=dict)


# ────────────────────────────────────────────────────────── failure triage
#
# recall@5 sitting at ~0.70 means ~30% of questions miss. A single recall
# number can't tell you WHY — and the fix is completely different per cause:
#
#   * answer_absent_from_chunks → the answer value is in NO chunk at all.
#     The parser dropped/garbled the cell. Fix the EXTRACTION pipeline.
#   * present_but_ranked_low    → a chunk DOES contain the answer, but the
#     embedding model ranked it past k. Fix CHUNKING (smaller/cleaner
#     chunks rank better) or the embedding step — NOT the parser.
#   * wrong_sheet               → answer is on a sheet we produced no chunk
#     for, or mis-attributed the sheet. Fix sheet enumeration.
#   * geometric_no_overlap      → no chunk's A1 range overlaps ground truth
#     even though text may match. Fix RANGE bookkeeping (top_left /
#     bottom_right anchors drift during merge/split).
#   * no_chunks / parse_error   → upstream parser failure.
#
# Splitting the miss population into these buckets turns "recall is low"
# into a ranked, actionable worklist. This is the single most useful thing
# for an agent iterating on recall — see scripts/triage_recall.py.

FAILURE_BUCKETS = (
    "answer_absent_from_chunks",
    "present_but_ranked_low",
    "wrong_sheet",
    "geometric_no_overlap",
    "no_chunks",
    "parse_error",
    "unparseable_ground_truth",
)


def classify_text_failure(rec: dict[str, Any]) -> str | None:
    """Bucket a single result record for the TEXT-match recall@5 metric.

    Returns None if the instance was a recall@5 hit (rank ≤ 5) or was not
    scoreable (no ground-truth values to match against).
    """
    if rec.get("error"):
        return "no_chunks" if rec.get("n_chunks", 0) == 0 else "parse_error"
    if not rec.get("had_answer_values", True):
        return None  # not scoreable for text-match — exclude, don't penalise
    if rec.get("n_chunks", 0) == 0:
        return "no_chunks"
    rank = rec.get("rank_of_text_match")
    if rank is not None and rank <= 5:
        return None  # hit
    if rank is None:
        # Not in any chunk. Distinguish "value never extracted" from
        # "value extracted but on a sheet we never chunked".
        if rec.get("answer_on_unchunked_sheet"):
            return "wrong_sheet"
        return "answer_absent_from_chunks"
    return "present_but_ranked_low"  # rank > 5


def score_instance(
    *,
    parser_name: str,
    extract_fn,
    input_path: Path,
    instruction: str,
    data_position: str,
    answer_position: str,
    default_sheet: str | None,
    answer_cell_values: list[str],
    answer_regions: list[tuple[str | None, tuple[int, int, int, int]]] | None = None,
    model,
    per_parser_timeout_s: float = 60.0,
    emit_failures: bool = False,
) -> InstanceResult:
    import numpy as np

    inst_id = input_path.parent.name

    t0 = time.perf_counter()
    try:
        with _alarm_timeout(per_parser_timeout_s):
            chunks = extract_fn(input_path)
    except _TimeoutError as exc:
        return InstanceResult(
            instance_id=inst_id,
            parser=parser_name,
            n_chunks=0,
            parse_ms=(time.perf_counter() - t0) * 1000.0,
            data_position=data_position,
            answer_position=answer_position,
            data_regions=0,
            chunks_overlapping_data=0,
            rank_of_first_overlap=None,
            rank_of_text_match=None,
            error=str(exc),
        )
    except Exception as exc:  # noqa: BLE001
        return InstanceResult(
            instance_id=inst_id,
            parser=parser_name,
            n_chunks=0,
            parse_ms=(time.perf_counter() - t0) * 1000.0,
            data_position=data_position,
            answer_position=answer_position,
            data_regions=0,
            chunks_overlapping_data=0,
            rank_of_first_overlap=None,
            rank_of_text_match=None,
            error=f"{type(exc).__name__}: {exc}",
        )

    parse_ms = (time.perf_counter() - t0) * 1000.0

    data_regions = parse_position_spec(data_position, default_sheet)

    if not chunks:
        return InstanceResult(
            instance_id=inst_id,
            parser=parser_name,
            n_chunks=0,
            parse_ms=parse_ms,
            data_position=data_position,
            answer_position=answer_position,
            data_regions=len(data_regions),
            chunks_overlapping_data=0,
            rank_of_first_overlap=None,
            rank_of_text_match=None,
            error="no chunks produced",
        )

    # Table-integrity: how many chunks overlap any of the input data regions?
    overlap_idxs: list[int] = []
    for i, c in enumerate(chunks):
        for sheet, box in data_regions:
            if c.overlaps(sheet or "", box):
                overlap_idxs.append(i)
                break

    # Embed chunks + query
    texts = [c.text or " " for c in chunks]
    embs = model.encode(texts, convert_to_numpy=True, normalize_embeddings=True,
                        show_progress_bar=False)
    q_emb = model.encode([instruction], convert_to_numpy=True,
                         normalize_embeddings=True, show_progress_bar=False)[0]
    sims = embs @ q_emb
    ranking = np.argsort(-sims)  # best first

    # rank of first overlap (1-indexed)
    rank_overlap: int | None = None
    if overlap_idxs:
        for r, idx in enumerate(ranking, start=1):
            if idx in overlap_idxs:
                rank_overlap = r
                break

    # rank of first chunk that contains any of the expected cell values.
    # Uses parser-agnostic numeric/date normalization on both sides so
    # "1272" matches a chunk rendering "1,272.00" and vice-versa.
    rank_text: int | None = None
    if answer_cell_values:
        for r, idx in enumerate(ranking, start=1):
            text = chunks[idx].text or ""
            if _matches_chunk_text(answer_cell_values, text):
                rank_text = r
                break

    # Did the parser produce ANY chunk on the sheet(s) the answer lives on?
    # If not, a text miss is a sheet-enumeration bug, not a cell-drop bug.
    answer_on_unchunked_sheet = False
    if answer_regions:
        gt_sheets = {s for s, _ in answer_regions if s}
        chunk_sheets = {c.sheet for c in chunks if c.sheet}
        if gt_sheets and chunk_sheets and not (gt_sheets & chunk_sheets):
            answer_on_unchunked_sheet = True

    extra: dict[str, Any] = {
        "answer_on_unchunked_sheet": answer_on_unchunked_sheet,
        "had_answer_values": bool(answer_cell_values),
    }
    if emit_failures:
        # Dump the top-8 ranked chunks so a human/agent can eyeball WHY the
        # answer was missed without re-running the (expensive) benchmark.
        top: list[dict[str, Any]] = []
        for r, idx in enumerate(ranking[:8], start=1):
            c = chunks[idx]
            top.append({
                "rank": r,
                "sheet": c.sheet,
                "range": (
                    f"{c.top_left}:{c.bottom_right}"
                    if c.top_left else None
                ),
                "contains_answer": _matches_chunk_text(answer_cell_values, c.text or ""),
                "text": (c.text or "")[:280],
            })
        extra["top_chunks"] = top
        extra["answer_values"] = answer_cell_values[:20]

    return InstanceResult(
        instance_id=inst_id,
        parser=parser_name,
        n_chunks=len(chunks),
        parse_ms=parse_ms,
        data_position=data_position,
        answer_position=answer_position,
        data_regions=len(data_regions),
        chunks_overlapping_data=len(overlap_idxs),
        rank_of_first_overlap=rank_overlap,
        rank_of_text_match=rank_text,
        extra=extra,
    )


# ────────────────────────────────────────────────────────────── answer values


def classify_execution_required(
    input_xlsx: Path,
    answer_regions: list[tuple[str | None, tuple[int, int, int, int]]],
) -> bool:
    """Return True if the INPUT spreadsheet has no non-empty cells in any
    `answer_regions` cell — i.e. the question is asking the system to
    *compute and write* the answer, not retrieve it.

    Parser-independent: only depends on the dataset + input.xlsx. Mirrors
    the `instruction_requires_execution` flag emitted by enrich_failures.py
    so the two stay consistent.

    Returns False when in doubt (unreadable workbook, missing sheet) — the
    in-scope number is the one we want to inflate the LEAST, so we err
    toward "scoreable, even if dubious".
    """
    if not answer_regions:
        return False
    try:
        from openpyxl import load_workbook

        # data_only=True so the classifier matches the
        # `instruction_requires_execution` flag emitted by
        # ``scripts/enrich_failures.py`` (which uses the cached-value pass).
        # An uncached formula at the answer cell means the parser has no
        # value to retrieve — the question genuinely requires execution.
        # See docs/planning/recall-90/05-out-of-scope-execution-instances.md
        # acceptance criterion (2): "classifier reproducibly identifies
        # ≥ 120 of the 127 currently-flagged instances".
        wb = load_workbook(str(input_xlsx), data_only=True, read_only=True)
        try:
            for sheet_name, (r0, c0, r1, c1) in answer_regions:
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                elif wb.worksheets:
                    ws = wb.worksheets[0]
                else:
                    continue
                for row in ws.iter_rows(min_row=r0, max_row=r1, min_col=c0,
                                        max_col=c1, values_only=True):
                    for v in row:
                        if v is not None and str(v).strip():
                            return False
            return True
        finally:
            wb.close()
    except Exception:
        return False


def read_answer_cell_values(
    answer_xlsx: Path,
    regions: list[tuple[str | None, tuple[int, int, int, int]]],
) -> list[str]:
    """Read distinct non-empty cell values across all `regions` of `answer_xlsx`.

    These become the ground-truth string tokens that should appear in
    the chunk a parser surfaces (text-match metric). Complementary to
    geometric overlap, which docling can't satisfy because it doesn't
    expose A1 anchors.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(answer_xlsx), data_only=True, read_only=True)
        seen: set[str] = set()
        values: list[str] = []
        for sheet_name, (r0, c0, r1, c1) in regions:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.worksheets:
                ws = wb.worksheets[0]
            else:
                continue
            for row in ws.iter_rows(min_row=r0, max_row=r1, min_col=c0,
                                    max_col=c1, values_only=True):
                for v in row:
                    if v is None:
                        continue
                    s = str(v).strip()
                    if len(s) >= 2 and s not in seen:
                        seen.add(s)
                        values.append(s)
                    if len(values) >= 50:
                        wb.close()
                        return values
        wb.close()
        return values
    except Exception:
        return []


# ────────────────────────────────────────────────────────────── aggregation


def aggregate(
    results: list[InstanceResult],
    execution_required: dict[str, bool] | None = None,
) -> dict[str, Any]:
    """Aggregate per-instance results into per-parser headline metrics.

    When ``execution_required`` is supplied, emit *both* `recall_*@k` (all
    instances — the long-standing number) and `recall_*@k_in_scope`
    (denominator excludes instances the parser cannot possibly satisfy
    because the answer cells are empty in the input). The latter is the
    metric the recall-90 roadmap actually gates on. See
    `docs/planning/recall-90/05-out-of-scope-execution-instances.md`.
    """
    exec_map = execution_required or {}
    by_parser: dict[str, list[InstanceResult]] = {}
    for r in results:
        by_parser.setdefault(r.parser, []).append(r)

    summary: dict[str, Any] = {}
    for parser, recs in by_parser.items():
        total = len(recs)
        errors = sum(1 for r in recs if r.error)
        ok = total - errors
        n_out_of_scope = sum(
            1 for r in recs if exec_map.get(str(r.instance_id))
        )
        in_scope_recs = [
            r for r in recs if not exec_map.get(str(r.instance_id))
        ]

        def _recall_at(
            k: int,
            key: str,
            subset: list[InstanceResult],
            require_values: bool = False,
        ) -> float:
            hits = 0
            denom = 0
            for r in subset:
                if r.error:
                    continue
                # Text-match recall is undefined for instances whose answer
                # region holds no scoreable values (empty / uncached-formula
                # answer.xlsx cells). The bucket histogram already excludes
                # these via `had_answer_values`; the recall denominator must be
                # consistent and not count an unscoreable instance as a miss.
                if require_values and not r.extra.get("had_answer_values", True):
                    continue
                rank = getattr(r, key)
                if rank is None:
                    denom += 1  # parser produced chunks but missed the answer
                    continue
                denom += 1
                if rank <= k:
                    hits += 1
            return hits / denom if denom else 0.0

        # fragmentation: among instances where the input data region is
        # covered, how many chunks does it span? 1 = clean, >1 = fragmented.
        # We only count single-region instances so n_chunks_overlap is
        # directly comparable; multi-region instances would inflate by
        # design.
        frags = [r.chunks_overlapping_data for r in recs
                 if not r.error and r.data_regions == 1
                 and r.chunks_overlapping_data > 0]
        n_with_overlap = len(frags)
        n_clean = sum(1 for f in frags if f == 1)
        n_frag = n_with_overlap - n_clean
        frag_rate = (n_frag / n_with_overlap) if n_with_overlap else 0.0

        parse_times = [r.parse_ms for r in recs if not r.error]

        # Failure-bucket histogram for the text-match recall@5 metric.
        # Only counts instances that HAD ground-truth values to match
        # against (others aren't scoreable). See classify_text_failure.
        buckets = dict.fromkeys(FAILURE_BUCKETS, 0)
        for r in recs:
            bucket = classify_text_failure({
                "error": r.error,
                "n_chunks": r.n_chunks,
                "rank_of_text_match": r.rank_of_text_match,
                "had_answer_values": r.extra.get("had_answer_values", True),
                "answer_on_unchunked_sheet": r.extra.get("answer_on_unchunked_sheet"),
            })
            if bucket is not None:
                buckets[bucket] += 1

        summary[parser] = {
            "instances": total,
            "ok": ok,
            "errors": errors,
            "out_of_scope_instances": n_out_of_scope,
            "in_scope_instances": total - n_out_of_scope,
            "recall_geometric@1": _recall_at(1, "rank_of_first_overlap", recs),
            "recall_geometric@3": _recall_at(3, "rank_of_first_overlap", recs),
            "recall_geometric@5": _recall_at(5, "rank_of_first_overlap", recs),
            "recall_text@1": _recall_at(1, "rank_of_text_match", recs, require_values=True),
            "recall_text@3": _recall_at(3, "rank_of_text_match", recs, require_values=True),
            "recall_text@5": _recall_at(5, "rank_of_text_match", recs, require_values=True),
            # Recall over the subset of instances the parser can possibly
            # satisfy (execution-required questions excluded). This is the
            # metric the recall-90 roadmap gates on; see cluster-05 doc.
            "recall_geometric@1_in_scope":
                _recall_at(1, "rank_of_first_overlap", in_scope_recs),
            "recall_geometric@3_in_scope":
                _recall_at(3, "rank_of_first_overlap", in_scope_recs),
            "recall_geometric@5_in_scope":
                _recall_at(5, "rank_of_first_overlap", in_scope_recs),
            "recall_text@1_in_scope":
                _recall_at(1, "rank_of_text_match", in_scope_recs, require_values=True),
            "recall_text@3_in_scope":
                _recall_at(3, "rank_of_text_match", in_scope_recs, require_values=True),
            "recall_text@5_in_scope":
                _recall_at(5, "rank_of_text_match", in_scope_recs, require_values=True),
            "table_integrity_clean": n_clean,
            "table_integrity_fragmented": n_frag,
            "table_fragmentation_rate": round(frag_rate, 4),
            "mean_parse_ms": round(sum(parse_times) / len(parse_times), 2)
            if parse_times else None,
            "p50_parse_ms": round(sorted(parse_times)[len(parse_times) // 2], 2)
            if parse_times else None,
            # Why the text-match recall@5 misses happen, bucketed. The
            # biggest bucket is the highest-leverage thing to fix next.
            "failure_buckets": buckets,
        }

    return summary


# ────────────────────────────────────────────────────────────── main


def iter_instances(corpus: Path) -> Iterable[dict[str, Any]]:
    ds = corpus / "dataset.json"
    if not ds.exists():
        raise FileNotFoundError(f"dataset.json not found in {corpus}")
    data = json.loads(ds.read_text())
    if not isinstance(data, list):
        raise ValueError("dataset.json should be a list of instances")
    yield from data


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--corpus", type=Path,
                        default=REPO_ROOT / "data" / "corpora" / "spreadsheetbench"
                        / "all_data_912_v0.1")
    parser.add_argument("--out", type=Path,
                        default=REPO_ROOT / "tests" / "benchmarks"
                        / "reports" / "retrieval")
    parser.add_argument("--parsers", type=str, default="ks,docling")
    parser.add_argument("--sample", type=int, default=None,
                        help="Random-sample N instances (seeded).")
    parser.add_argument("--seed", type=int, default=1337)
    parser.add_argument("--model", type=str, default="BAAI/bge-small-en-v1.5")
    parser.add_argument("--test-case", type=int, default=1,
                        help="Which of the (typically 3) test cases per instance "
                             "to score on. We use one to keep eval costs bounded.")
    parser.add_argument("--emit-failures", action="store_true",
                        help="Also write failures.ndjson — one row per "
                             "recall@5 miss with the top-8 ranked chunks and "
                             "ground-truth values, for failure triage.")
    parser.add_argument("--per-parser-timeout", type=float, default=60.0,
                        help="Wall-clock seconds before a parser is "
                             "considered hung on a single file (docling can "
                             "loop forever on pathological table layouts).")
    args = parser.parse_args(argv)

    instances = list(iter_instances(args.corpus))
    if args.sample is not None and args.sample < len(instances):
        import random
        rng = random.Random(args.seed)
        instances = rng.sample(instances, args.sample)
    sys.stderr.write(f"Scoring {len(instances)} SpreadsheetBench instances\n")

    selected = {p.strip() for p in args.parsers.split(",")}
    parser_fns: dict[str, Any] = {}
    if "ks" in selected:
        parser_fns["excel-parser"] = extract_chunks_ks
    if "docling" in selected:
        parser_fns["docling"] = extract_chunks_docling
    if not parser_fns:
        sys.stderr.write("no valid parsers selected\n")
        return 2

    # Load embedding model
    sys.stderr.write(f"Loading embedding model: {args.model}\n")
    from sentence_transformers import SentenceTransformer
    model = SentenceTransformer(args.model)

    from datetime import UTC, datetime
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%S")
    out_dir = args.out / stamp
    out_dir.mkdir(parents=True, exist_ok=True)
    ndjson_path = out_dir / "results.ndjson"

    results: list[InstanceResult] = []
    failure_rows: list[dict[str, Any]] = []
    # Map instance_id → True if input.xlsx has nothing at answer_position
    # (i.e. the question asks the system to compute and write a value).
    # Classified once per instance — parser-independent, cheap reuse.
    execution_required: dict[str, bool] = {}
    n = len(instances) * len(parser_fns)
    done = 0

    with ndjson_path.open("w") as f:
        for inst in instances:
            inst_id = str(inst["id"])
            instr = inst["instruction"]
            data_pos = inst.get("data_position") or ""
            answer_pos = inst.get("answer_position") or ""
            default_sheet = inst.get("answer_sheet") or None
            if default_sheet and "," in default_sheet:
                # answer_sheet is multi-sheet; pick the first as default,
                # the per-region parsers will override anyway.
                default_sheet = default_sheet.split(",")[0].strip()

            inst_dir = args.corpus / "spreadsheet" / inst_id
            input_path = inst_dir / f"{args.test_case}_{inst_id}_input.xlsx"
            answer_path = inst_dir / f"{args.test_case}_{inst_id}_answer.xlsx"

            if not input_path.exists() or not answer_path.exists():
                done += len(parser_fns)
                sys.stderr.write(f"\r[{done}/{n}] skipped (files missing): {inst_id}\n")
                continue

            # Geometric ground truth: 561/912 instances leave data_position
            # empty. For those, the question targets the answer region in
            # the input file (the answer cells already exist there as a
            # template the system rewrites). Fall back to answer_position
            # so we exercise every instance.
            geom_spec = data_pos or answer_pos
            data_pos_for_record = geom_spec

            # Cell values come from answer.xlsx in the answer regions —
            # that's what the question is asking the system to surface.
            answer_regions = parse_position_spec(answer_pos, default_sheet)
            answer_values = (
                read_answer_cell_values(answer_path, answer_regions)
                if answer_regions else []
            )

            # Cluster-05 in-scope classifier: if the *input* spreadsheet has
            # nothing at answer_position the question requires execution and
            # no parser can retrieve the answer.
            execution_required[inst_id] = classify_execution_required(
                input_path, answer_regions
            )

            for parser_name, extract_fn in parser_fns.items():
                res = score_instance(
                    parser_name=parser_name,
                    extract_fn=extract_fn,
                    input_path=input_path,
                    instruction=instr,
                    data_position=geom_spec,
                    answer_position=answer_pos,
                    default_sheet=default_sheet,
                    answer_cell_values=answer_values,
                    answer_regions=answer_regions,
                    model=model,
                    per_parser_timeout_s=args.per_parser_timeout,
                    emit_failures=args.emit_failures,
                )
                results.append(res)
                bucket = classify_text_failure({
                    "error": res.error,
                    "n_chunks": res.n_chunks,
                    "rank_of_text_match": res.rank_of_text_match,
                    "had_answer_values": res.extra.get("had_answer_values", True),
                    "answer_on_unchunked_sheet": res.extra.get(
                        "answer_on_unchunked_sheet"),
                })
                f.write(json.dumps({
                    "instance_id": res.instance_id,
                    "parser": res.parser,
                    "n_chunks": res.n_chunks,
                    "parse_ms": res.parse_ms,
                    "data_position": res.data_position,
                    "answer_position": res.answer_position,
                    "data_regions": res.data_regions,
                    "chunks_overlapping_data": res.chunks_overlapping_data,
                    "rank_of_first_overlap": res.rank_of_first_overlap,
                    "rank_of_text_match": res.rank_of_text_match,
                    "failure_bucket": bucket,
                    "error": res.error,
                }, separators=(",", ":")) + "\n")
                if args.emit_failures and bucket is not None:
                    failure_rows.append({
                        "instance_id": res.instance_id,
                        "parser": res.parser,
                        "failure_bucket": bucket,
                        "instruction": instr,
                        "answer_position": answer_pos,
                        "answer_values": res.extra.get("answer_values", []),
                        "rank_of_text_match": res.rank_of_text_match,
                        "n_chunks": res.n_chunks,
                        "top_chunks": res.extra.get("top_chunks", []),
                        "error": res.error,
                    })
                done += 1
                if done % 10 == 0:
                    sys.stderr.write(f"\r[{done}/{n}] ")
                    sys.stderr.flush()

    sys.stderr.write(f"\nWrote {ndjson_path}\n")

    if args.emit_failures:
        fail_path = out_dir / "failures.ndjson"
        with fail_path.open("w") as ff:
            for row in failure_rows:
                ff.write(json.dumps(row, separators=(",", ":")) + "\n")
        sys.stderr.write(f"Wrote {fail_path} ({len(failure_rows)} failure rows)\n")

    summary = aggregate(results, execution_required=execution_required)
    summary_path = out_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, indent=2))
    sys.stderr.write(f"Wrote {summary_path}\n")

    # Audit log of which instances the in-scope filter excluded. Diff this
    # across runs to spot classifier drift; argue with it if you disagree.
    out_of_scope_ids = sorted(
        iid for iid, is_exec in execution_required.items() if is_exec
    )
    (out_dir / "out_of_scope.txt").write_text(
        f"# {len(out_of_scope_ids)} instances classified as "
        "instruction_requires_execution\n"
        "# (input.xlsx has no non-empty cells at answer_position — "
        "answer must be computed, not retrieved)\n"
        + "\n".join(out_of_scope_ids) + ("\n" if out_of_scope_ids else "")
    )

    # Human-readable summary
    md_lines = ["# Retrieval-recall benchmark (SpreadsheetBench)\n"]
    md_lines.append(f"- Corpus: `{args.corpus}`")
    md_lines.append(f"- Instances scored: {len(instances)}")
    md_lines.append(f"- Embedding model: `{args.model}`")
    md_lines.append("")
    parsers = sorted(summary.keys())
    md_lines.append("| Metric | " + " | ".join(parsers) + " |")
    md_lines.append("|---|" + "|".join(["---"] * len(parsers)) + "|")
    metrics = [
        ("recall_geometric@1", "Recall@1 (geometric, all)"),
        ("recall_geometric@3", "Recall@3 (geometric, all)"),
        ("recall_geometric@5", "Recall@5 (geometric, all)"),
        ("recall_text@1", "Recall@1 (text-match, all)"),
        ("recall_text@3", "Recall@3 (text-match, all)"),
        ("recall_text@5", "Recall@5 (text-match, all)"),
        ("recall_geometric@5_in_scope", "Recall@5 (geometric, in-scope) **"),
        ("recall_text@5_in_scope", "Recall@5 (text-match, in-scope) **"),
        ("in_scope_instances", "In-scope instances"),
        ("out_of_scope_instances", "Out-of-scope (execution-required)"),
        ("table_fragmentation_rate", "Fragmentation rate"),
        ("mean_parse_ms", "Mean parse ms"),
        ("p50_parse_ms", "P50 parse ms"),
        ("errors", "Errors"),
    ]
    for key, label in metrics:
        row = [label]
        for p in parsers:
            v = summary[p].get(key)
            if v is None:
                row.append("—")
            elif isinstance(v, float):
                row.append(f"{v:.3f}")
            else:
                row.append(str(v))
        md_lines.append("| " + " | ".join(row) + " |")
    md_lines.append("")
    md_lines.append("**Geometric overlap** = chunk's reported A1 range overlaps the "
                    "ground-truth `answer_position`. Requires the parser to surface "
                    "(sheet, range) per chunk — docling does not, so its geometric "
                    "recall is structurally 0.")
    md_lines.append("")
    md_lines.append("** **In-scope** excludes instances where the input spreadsheet "
                    "has nothing at `answer_position` (the question asks the system to "
                    "*compute and write* the answer; a retrieval parser cannot help). "
                    "The recall-90 roadmap gates on the in-scope number. See "
                    "`docs/planning/recall-90/05-out-of-scope-execution-instances.md`.")
    md_lines.append("")
    md_lines.append("## Failure buckets (text-match recall@5 misses)")
    md_lines.append("")
    md_lines.append("Why each miss happened — the biggest bucket is the highest-")
    md_lines.append("leverage fix. `answer_absent_from_chunks` → fix extraction; ")
    md_lines.append("`present_but_ranked_low` → fix chunking/embedding.")
    md_lines.append("")
    md_lines.append("| Bucket | " + " | ".join(parsers) + " |")
    md_lines.append("|---|" + "|".join(["---"] * len(parsers)) + "|")
    for b in FAILURE_BUCKETS:
        row = [b]
        for p in parsers:
            row.append(str(summary[p].get("failure_buckets", {}).get(b, 0)))
        md_lines.append("| " + " | ".join(row) + " |")
    md_lines.append("")
    md_lines.append("**Text-match** = the answer cell's actual string value appears "
                    "as a substring of the chunk's text. Parser-agnostic; this is "
                    "the apples-to-apples retrieval comparison.")
    md_lines.append("")
    (out_dir / "summary.md").write_text("\n".join(md_lines) + "\n")
    sys.stderr.write(f"Wrote {out_dir / 'summary.md'}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
